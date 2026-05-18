"""Geração dos arquivos de saída: CSVs e Excel formatado."""

import csv
import os

from config import ANOS, OUT, LINHAS_ESTADO, LINHAS_CAPITAL
from utils import fmt_mi, pct, get, rcl6


# ---------------------------------------------------------------------------
# CSV – Quadro de decomposição da dívida
# ---------------------------------------------------------------------------

def gerar_quadro_csv(linhas, dados, nome_arquivo):
    """Salva a tabela de decomposição da dívida em CSV (valores em R$ milhões)."""
    caminho = os.path.join(OUT, nome_arquivo)
    header = ["Conta/Indicador"] + [str(a) for a in ANOS]
    rows = [["Unidade: R$ milhões"], header]

    for label, chave, _ in linhas:
        if chave is None:
            rows.append([label] + [""] * len(ANOS))
            continue
        valores = []
        for ano in ANOS:
            d = dados[ano]
            if chave == "rcl_ajustada":
                if "rcl_ajustada" in d:
                    v = d["rcl_ajustada"]
                elif d.get("transf_emendas", 0) == 0:
                    v = get(d, "rcl")
                else:
                    v = None
            else:
                v = get(d, chave)
                if v == 0.0 and chave not in d:
                    valores.append("–")
                    continue
            valores.append(fmt_mi(v) if v is not None else "–")
        rows.append([label] + valores)

    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, delimiter=";").writerows(rows)
    print(f"  Salvo: {nome_arquivo}")


# ---------------------------------------------------------------------------
# CSV – Indicadores de endividamento e liquidez
# ---------------------------------------------------------------------------

def gerar_indicadores_csv(dados_estado, dados_capital, liq_estado, nome_arquivo, servico_divida=None):
    """Salva o quadro de indicadores em CSV."""
    caminho = os.path.join(OUT, nome_arquivo)
    header = ["Indicador", "Ente"] + [str(a) for a in ANOS]
    rows_out = [header]

    def linha(label, ente, valores):
        rows_out.append([label, ente] + [
            f"{v:.2f}" if v is not None else "–" for v in valores
        ])

    # Endividamento (Anexo 2)
    for ente, dados, limite in [("MS (Estado)", dados_estado, 200), ("Campo Grande", dados_capital, 120)]:
        vals = [pct(get(dados[a], "dcl"), rcl6(dados[a])) for a in ANOS]
        linha(f"DCL / RCL Ajustada (%) — limite LRF: {limite}%", ente, vals)

    if servico_divida:
        for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [pct(servico_divida[ente][a]["servico_divida"], rcl6(dados[a])) for a in ANOS]
            linha("Serviço da Dívida / RCL Ajustada (%)", ente, vals)
        for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [pct(servico_divida[ente][a]["pessoal_encargos"], get(dados[a], "rcl")) for a in ANOS]
            linha("Despesa com Pessoal / RCL (%)", ente, vals)

    for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
        vals = [pct(get(dados[a], "passivo_atuarial"), rcl6(dados[a]))
                if get(dados[a], "passivo_atuarial") else None for a in ANOS]
        linha("Passivo Atuarial (RPPS) / RCL Ajustada (%)", ente, vals)

    for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
        vals = [pct(get(dados[a], "rp_nao_processados"), rcl6(dados[a]))
                if get(dados[a], "rp_nao_processados") else None for a in ANOS]
        linha("RP Não-Processados / RCL Ajustada (%)", ente, vals)

    for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
        vals = [pct(get(dados[a], "aprops_dep_judiciais"), rcl6(dados[a]))
                if get(dados[a], "aprops_dep_judiciais") else None for a in ANOS]
        linha("Apropr. Depósitos Judiciais / RCL Ajustada (%)", ente, vals)

    # Liquidez (Anexo 5) – apenas MS Estado
    rows_out.append(["--- INDICADORES DE LIQUIDEZ (Anexo 5) – MS Estado ---", ""] + [""] * len(ANOS))
    for chave_liq, label_liq in [
        ("disp_bruta_a5",  "Disp. de Caixa Bruta / RCL Ajustada (%)"),
        ("disp_liq_antes", "Disp. de Caixa Líquida (antes RP) / RCL Ajustada (%)"),
        ("disp_liq_apos",  "Disp. de Caixa Líquida (após RP) / RCL Ajustada (%)"),
    ]:
        vals = [pct(liq_estado[a].get(chave_liq, 0), rcl6(dados_estado[a]))
                if liq_estado[a].get(chave_liq) is not None else None for a in ANOS]
        linha(label_liq, "MS (Estado)", vals)

    # Valores absolutos (referência)
    rows_out.append([])
    for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
        rows_out.append(["DCL (R$ milhões)", ente]
                        + [f"{get(dados[a], 'dcl')/1e6:.3f}" for a in ANOS])
    for ente, dados in [("MS (Estado)", dados_estado), ("Campo Grande", dados_capital)]:
        rows_out.append(["RCL Ajustada (R$ milhões)", ente]
                        + [f"{rcl6(dados[a])/1e6:.3f}" for a in ANOS])
    if servico_divida:
        for ente in ["MS (Estado)", "Campo Grande"]:
            rows_out.append(["Serviço da Dívida (R$ milhões)", ente]
                            + [f"{servico_divida[ente][a]['servico_divida']/1e6:.3f}" for a in ANOS])
            rows_out.append(["  Juros e Encargos da Dívida (R$ milhões)", ente]
                            + [f"{servico_divida[ente][a]['juros_encargos']/1e6:.3f}" for a in ANOS])
            rows_out.append(["  Amortização da Dívida (R$ milhões)", ente]
                            + [f"{servico_divida[ente][a]['amortizacao_divida']/1e6:.3f}" for a in ANOS])
            rows_out.append(["Despesa com Pessoal (R$ milhões)", ente]
                            + [f"{servico_divida[ente][a]['pessoal_encargos']/1e6:.3f}" for a in ANOS])
    for chave_liq, label_liq in [
        ("disp_bruta_a5", "Disp. de Caixa Bruta – Anexo 5 (R$ mi)"),
        ("disp_liq_apos", "Disp. de Caixa Líquida (após RP) – Anexo 5 (R$ mi)"),
    ]:
        rows_out.append([label_liq, "MS (Estado)"] + [
            f"{liq_estado[a].get(chave_liq, 0)/1e6:.3f}"
            if liq_estado[a].get(chave_liq) is not None else "–"
            for a in ANOS
        ])

    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, delimiter=";").writerows(rows_out)
    print(f"  Salvo: {nome_arquivo}")


# ---------------------------------------------------------------------------
# Excel – workbook formatado
# ---------------------------------------------------------------------------

def gerar_excel(dados_estado, dados_capital, liq_estado, nome_arquivo, servico_divida=None):
    """Salva o Excel consolidado com 3 abas formatadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COL_HEADER = "1E3A5F"
    COL_TOTAL  = "D6E4F0"
    COL_WARN   = "FFF2CC"
    COL_LIQ    = "E8F5E9"
    COL_REF    = "F5F5F5"

    def _border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    # ── aba de decomposição ──────────────────────────────────────────────────
    def _aba_quadro(wb, titulo, linhas, dados):
        ws = wb.create_sheet(title=titulo)
        ws.append([""])
        hrow = ["Conta/Indicador (R$ milhões)"] + [str(a) for a in ANOS]
        ws.append(hrow)
        for ci, val in enumerate(hrow, 1):
            c = ws.cell(row=2, column=ci, value=val)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = _fill(COL_HEADER)
            c.alignment = Alignment(horizontal="center")
            c.border    = _border()

        ws.column_dimensions["A"].width = 44
        for i in range(2, len(ANOS) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 14

        rn = 3
        for label, chave, negrito in linhas:
            if chave is None:
                ws.append([label])
                c = ws.cell(row=rn, column=1)
                c.font = Font(bold=True, italic=True)
                c.fill = _fill("E8E8E8")
                rn += 1
                continue

            vals = []
            for ano in ANOS:
                d = dados[ano]
                if chave == "rcl_ajustada":
                    if "rcl_ajustada" in d:
                        v = d["rcl_ajustada"]
                    elif d.get("transf_emendas", 0) == 0:
                        v = get(d, "rcl")
                    else:
                        v = None
                else:
                    v = get(d, chave)
                    if v == 0.0 and chave not in d:
                        v = None
                vals.append(v / 1e6 if v is not None else None)

            row_data = [label] + [v if v is not None else "" for v in vals]
            ws.append(row_data)
            fill = _fill(COL_TOTAL) if negrito else None
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=rn, column=ci)
                if fill:
                    cell.fill = fill
                if negrito:
                    cell.font = Font(bold=True)
                if ci > 1 and val != "":
                    cell.number_format = "#,##0.000"
                    cell.alignment = Alignment(horizontal="right")
                cell.border = _border()
            rn += 1

    # ── aba de indicadores ───────────────────────────────────────────────────
    def _aba_indicadores(wb):
        ws = wb.create_sheet(title="Indicadores")
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 18
        for i in range(3, len(ANOS) + 3):
            ws.column_dimensions[get_column_letter(i)].width = 12

        header = ["Indicador", "Ente"] + [str(a) for a in ANOS]
        ws.append(header)
        for ci, val in enumerate(header, 1):
            c = ws.cell(row=1, column=ci, value=val)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = _fill(COL_HEADER)
            c.alignment = Alignment(horizontal="center")
            c.border    = _border()

        rn = 2

        def _linha(label, ente, vals, fmt=".2f", fill_color=None, bold=False):
            nonlocal rn
            row_data = [label, ente] + [
                (round(v, 2) if fmt == ".2f" else round(v, 3)) if v is not None else ""
                for v in vals
            ]
            ws.append(row_data)
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=rn, column=ci)
                if fill_color:
                    cell.fill = _fill(fill_color)
                if bold:
                    cell.font = Font(bold=True)
                if ci > 2 and val != "":
                    cell.number_format = "#,##0.00" if fmt == ".2f" else "#,##0.000"
                    cell.alignment = Alignment(horizontal="right")
                cell.border = _border()
            rn += 1

        def _sep(texto=""):
            nonlocal rn
            ws.append([texto])
            ws.cell(row=rn, column=1).font = Font(bold=True, italic=True)
            ws.cell(row=rn, column=1).fill = _fill("E8E8E8")
            for ci in range(1, len(ANOS) + 3):
                ws.cell(row=rn, column=ci).border = _border()
            rn += 1

        # Indicadores de endividamento (Anexo 2)
        _sep("── ENDIVIDAMENTO (Anexo 2 RGF) ──")
        for ente, dados, limite in [
            ("MS – Estado", dados_estado, 200),
            ("Campo Grande", dados_capital, 120),
        ]:
            vals = [pct(get(dados[a], "dcl"), rcl6(dados[a])) for a in ANOS]
            _linha(f"1. DCL / RCL Ajustada (%) – limite LRF: {limite}%",
                   ente, vals, fill_color=COL_TOTAL, bold=True)

        if servico_divida:
            for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
                chave_ente = "MS (Estado)" if ente.startswith("MS") else "Campo Grande"
                vals = [pct(servico_divida[chave_ente][a]["servico_divida"], rcl6(dados[a])) for a in ANOS]
                _linha("2. Serviço da Dívida / RCL Ajustada (%)", ente, vals, fill_color=COL_TOTAL, bold=True)
            for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
                chave_ente = "MS (Estado)" if ente.startswith("MS") else "Campo Grande"
                vals = [pct(servico_divida[chave_ente][a]["pessoal_encargos"], get(dados[a], "rcl")) for a in ANOS]
                _linha("3. Despesa com Pessoal / RCL (%)", ente, vals, fill_color=COL_LIQ)

        for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [pct(get(dados[a], "passivo_atuarial"), rcl6(dados[a]))
                    if get(dados[a], "passivo_atuarial") else None for a in ANOS]
            _linha("4. Passivo Atuarial (RPPS) / RCL Ajustada (%)", ente, vals, fill_color=COL_WARN)

        for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [pct(get(dados[a], "rp_nao_processados"), rcl6(dados[a]))
                    if get(dados[a], "rp_nao_processados") else None for a in ANOS]
            _linha("5. RP Não-Processados / RCL Ajustada (%)", ente, vals, fill_color=COL_WARN)

        for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [pct(get(dados[a], "aprops_dep_judiciais"), rcl6(dados[a]))
                    if get(dados[a], "aprops_dep_judiciais") else None for a in ANOS]
            _linha("6. Apropr. Depósitos Judiciais / RCL Ajustada (%)", ente, vals, fill_color=COL_WARN)

        # Indicadores de liquidez (Anexo 5)
        _sep("── LIQUIDEZ (Anexo 5 RGF) – MS Estado ──")
        for num, (chave_liq, label_liq) in enumerate([
            ("disp_bruta_a5",  "Disp. de Caixa Bruta / RCL Ajustada (%)"),
            ("disp_liq_antes", "Disp. de Caixa Líquida (antes RP) / RCL Ajustada (%)"),
            ("disp_liq_apos",  "Disp. de Caixa Líquida (após RP) / RCL Ajustada (%)"),
        ], start=7):
            vals = [pct(liq_estado[a].get(chave_liq, 0), rcl6(dados_estado[a]))
                    if liq_estado[a].get(chave_liq) is not None else None for a in ANOS]
            _linha(f"{num}. {label_liq}", "MS – Estado", vals,
                   fill_color=COL_LIQ, bold=(num == 5))

        # Valores absolutos (referência)
        _sep("── VALORES ABSOLUTOS (R$ milhões) ──")
        for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [round(get(dados[a], "dcl") / 1e6, 3) for a in ANOS]
            _linha("DCL (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
        for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [round(rcl6(dados[a]) / 1e6, 3) for a in ANOS]
            _linha("RCL Ajustada (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
        if servico_divida:
            for ente in ["MS – Estado", "Campo Grande"]:
                chave_ente = "MS (Estado)" if ente.startswith("MS") else "Campo Grande"
                vals = [round(servico_divida[chave_ente][a]["servico_divida"] / 1e6, 3) for a in ANOS]
                _linha("Serviço da Dívida (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
                vals = [round(servico_divida[chave_ente][a]["juros_encargos"] / 1e6, 3) for a in ANOS]
                _linha("  Juros e Encargos da Dívida (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
                vals = [round(servico_divida[chave_ente][a]["amortizacao_divida"] / 1e6, 3) for a in ANOS]
                _linha("  Amortização da Dívida (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
                vals = [round(servico_divida[chave_ente][a]["pessoal_encargos"] / 1e6, 3) for a in ANOS]
                _linha("Despesa com Pessoal (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
        for ente, dados in [("MS – Estado", dados_estado), ("Campo Grande", dados_capital)]:
            vals = [round(get(dados[a], "passivo_atuarial") / 1e6, 3)
                    if get(dados[a], "passivo_atuarial") else None for a in ANOS]
            _linha("Passivo Atuarial RPPS (R$ mi)", ente, vals, fmt=".3f", fill_color=COL_REF)
        for chave_liq, label_liq in [
            ("disp_bruta_a5", "Disp. de Caixa Bruta – Anexo 5 (R$ mi)"),
            ("disp_liq_apos", "Disp. de Caixa Líquida (após RP) – Anexo 5 (R$ mi)"),
        ]:
            vals = [round(liq_estado[a].get(chave_liq, 0) / 1e6, 3)
                    if liq_estado[a].get(chave_liq) is not None else None for a in ANOS]
            _linha(label_liq, "MS – Estado", vals, fmt=".3f", fill_color=COL_REF)

    # ── montagem e salvamento ────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)
    _aba_quadro(wb, "Decomp.Dívida MS-Estado",   LINHAS_ESTADO,   dados_estado)
    _aba_quadro(wb, "Decomp.Dívida Campo Grande", LINHAS_CAPITAL,  dados_capital)
    _aba_indicadores(wb)

    caminho = os.path.join(OUT, nome_arquivo)
    wb.save(caminho)
    print(f"  Salvo: {nome_arquivo}")
