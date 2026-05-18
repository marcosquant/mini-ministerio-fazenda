"""Extração de dados dos arquivos-fonte do SICONFI.

  Anexo 2 RGF – xlsx  → extrair_dados_uf()
  Anexo 5 RGF – csv   → extrair_liquidez_uf()
"""

import csv
import os
import unicodedata
from collections import defaultdict

import openpyxl

from config import ANOS, BASE, CONTA_MAP, ANEXO5_TOTAL_CONTAS


def _normaliza(nome):
    return nome.strip() if nome else ""


def _normaliza_ascii(nome):
    texto = unicodedata.normalize("NFKD", str(nome or ""))
    return "".join(c for c in texto if ord(c) < 128).upper().strip()


def extrair_dados_uf(escopo, uf, quadrimestre="3"):
    """Lê os xlsx do Anexo 2 e retorna {ano: {conta_key: valor}}.

    Parâmetros
    ----------
    escopo      : 'Estados.DF' ou 'Capitais'
    uf          : sigla da UF (ex: 'MS')
    quadrimestre: '1', '2' ou '3' (padrão = 3º, posição de dezembro)
    """
    resultado = {}
    for ano in ANOS:
        fname = os.path.join(BASE, f"Dívida Consolidada Líquida - {escopo} - {ano}.xlsx")
        wb = openpyxl.load_workbook(fname, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        filtradas = [
            r for r in rows[6:]
            if r[2] == uf
            and r[5]
            and quadrimestre in str(r[5])
            and "Quadrimestre" in str(r[5])
        ]

        contas = {}
        # Rastreia conta-pai para resolver a ambiguidade de "Internos/Externos",
        # que aparecem tanto em Empréstimos quanto em Financiamentos.
        pai_atual = None
        for r in filtradas:
            nome  = _normaliza(r[6])
            valor = r[8] if r[8] is not None else 0.0
            chave = CONTA_MAP.get(nome)
            if chave is None:
                continue

            if chave == "__ambig_internos":
                chave = f"{pai_atual or 'outro'}_internos".replace("emprestimos", "emprest")
            elif chave == "__ambig_externos":
                chave = f"{pai_atual or 'outro'}_externos".replace("emprestimos", "emprest")
            else:
                if chave in ("emprestimos", "financiamentos", "parcelamento",
                             "reestruturacao", "contratual"):
                    pai_atual = chave

            contas[chave] = valor

        resultado[ano] = contas
        wb.close()
    return resultado


def extrair_liquidez_uf(uf):
    """Lê os CSV do Anexo 5 e retorna {ano: {indicador: valor}}.

    Indicadores produzidos
    ----------------------
    disp_bruta_a5  – Disponibilidade de Caixa Bruta (a), consolidado
    disp_liq_antes – Disp. Líquida ANTES inscrição RP não-processados
    disp_liq_apos  – Disp. Líquida APÓS  inscrição RP não-processados
    rp_exercicio   – RP não-processados do exercício inscritos
    """
    resultado = {}
    for ano in ANOS:
        fname = os.path.join(BASE, f"finbraRGF-{ano}.csv")
        with open(fname, encoding="cp1252", errors="replace") as f:
            linhas = f.readlines()

        reader = csv.DictReader(linhas[5:], delimiter=";")
        totais = [
            r for r in reader
            if r.get("UF") == uf and r.get("Conta") in ANEXO5_TOTAL_CONTAS
        ]

        soma = defaultdict(float)
        for r in totais:
            coluna = r["Coluna"]
            try:
                v = float(r["Valor"].replace(",", "."))
            except (ValueError, KeyError):
                continue

            if coluna.startswith("DISPONIBILIDADE DE CAIXA BRUTA"):
                soma["disp_bruta_a5"] += v
            elif "ANTES" in coluna and "INSCRIÇÃO" in coluna:
                soma["disp_liq_antes"] += v
            elif "APÓS" in coluna and "INSCRIÇÃO" in coluna:
                soma["disp_liq_apos"] += v
            elif coluna.startswith("RESTOS A PAGAR EMPENHADOS E NÃO LIQUIDADOS DO EXERCÍCIO"):
                soma["rp_exercicio"] += v

        resultado[ano] = dict(soma)
    return resultado


def extrair_servico_divida_etapa1():
    """Le a planilha da Etapa 1 e retorna o servico da divida em reais.

    Servico da divida = Juros e Encargos da Divida + Amortizacao da Divida.
    A planilha da Etapa 1 esta em R$ milhoes; o retorno e convertido para reais
    para manter a mesma unidade dos dados extraidos do SICONFI na Etapa 2.
    """
    caminho = os.path.join(
        os.path.dirname(__file__),
        "..",
        "Etapa1",
        "demonstrativos_fiscais_MS_2019_2024.xlsx",
    )
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    def linha(ws, label):
        wanted = _normaliza_ascii(label)
        for row in ws.iter_rows(values_only=True):
            if _normaliza_ascii(row[0]) == wanted:
                return [0.0 if v is None else float(v) * 1_000_000 for v in row[1:7]]
        raise KeyError(label)

    saida = {}
    sheets = {
        "MS (Estado)": ("Dem2_Estado_MS", "Dem1_Estado_MS"),
        "Campo Grande": ("Dem2_Capital_CG", "Dem1_Capital_CG"),
    }
    for ente, (sheet_financeiro, sheet_orcamentario) in sheets.items():
        ws_financeiro = wb[sheet_financeiro]
        ws_orcamentario = wb[sheet_orcamentario]
        juros = linha(ws_financeiro, "Juros e Encargos da Divida")
        amortizacao = linha(ws_financeiro, "Amortizacao da Divida")
        despesas = linha(ws_financeiro, "DESPESAS FINANCEIRAS")
        pessoal = linha(ws_orcamentario, "Pessoal e Encargos Sociais")
        saida[ente] = {}
        for idx, ano in enumerate(ANOS):
            servico = juros[idx] + amortizacao[idx]
            # Usa a soma detalhada; a linha agregada fica como conferencia.
            saida[ente][ano] = {
                "juros_encargos": juros[idx],
                "amortizacao_divida": amortizacao[idx],
                "servico_divida": servico,
                "despesas_financeiras": despesas[idx],
                "pessoal_encargos": pessoal[idx],
            }
    wb.close()
    return saida
