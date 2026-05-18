from __future__ import annotations

import sys
import unicodedata
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from reportlab.platypus import KeepTogether

ROOT = Path(__file__).resolve().parent
ETAPA1_DIR = ROOT.parent / "Etapa1"
sys.path.insert(0, str(ETAPA1_DIR))

import generate_relatorio_pdf as tpl
from config import LINHAS_CAPITAL, LINHAS_ESTADO
from extrair import extrair_servico_divida_etapa1


INPUT_XLSX = ROOT / "Etapa2_Diagnostico_Endividamento_MS_CG.xlsx"
BUILD_DIR = ROOT / "build_relatorio_endividamento"
CHART_DIR = BUILD_DIR / "charts"
OUTPUT_PDF = ROOT / "Relatorio_Diagnostico_Endividamento_MS_Campo_Grande.pdf"

IMG_MS = ETAPA1_DIR / "img_ms_natureza.jpg"
IMG_CG = ETAPA1_DIR / "img_campo_grande.jpg"

TEAM_MEMBERS = tpl.TEAM_MEMBERS
ANOS = [2019, 2020, 2021, 2022, 2023, 2024]


def normalize(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", str(text)) if ord(c) < 128).upper().strip()


LABEL_ALIASES = {
    "RP PROCESSADOS": ["RESTOS A PAGAR PROCESSADOS"],
    "RP NAO-PROCESSADOS": ["RP NAO-PROCESSADOS", "RESTOS A PAGAR NAO PROCESSADOS"],
    "REESTRUTURACAO": ["REESTRUTURACAO DA DIVIDA", "REESTRUTURACAO DA DIVIDA DE ESTADOS E MUNICIPIOS"],
    "DCL": ["DIVIDA CONSOLIDADA LIQUIDA", "DCL"],
    "RCL": ["RECEITA CORRENTE LIQUIDA", "RCL"],
    "RCL AJUSTADA": ["RCL AJUSTADA", "RECEITA CORRENTE LIQUIDA AJUSTADA"],
    "CONTRIBUICOES PREVIDENCIARIAS": ["CONTRIBUICOES PREVIDENCIARIAS", "DE CONTRIBUICOES PREVIDENCIARIAS"],
    "DEMAIS CONTRIBUICOES SOCIAIS": ["DEMAIS CONTRIBUICOES SOCIAIS", "DE DEMAIS CONTRIBUICOES SOCIAIS"],
}


def search_terms(label: str) -> list[str]:
    wanted = normalize(label)
    terms = [wanted]
    for key, aliases in LABEL_ALIASES.items():
        if wanted == key or wanted in key or key in wanted:
            terms.extend(aliases)
    return [normalize(term) for term in terms]


def fmt_million(value: float | None, decimals: int = 2) -> str:
    if value is None:
        return "–"
    return tpl.br_money_millions(value, decimals)


def fmt_percent(value: float | None, decimals: int = 2) -> str:
    if value is None:
        return "n.d."
    return tpl.br_percent(value, decimals)


def pct_change_opt(values: list[float | None]) -> float | None:
    if values[0] in (None, 0) or values[-1] is None:
        return None
    return (values[-1] - values[0]) / values[0] * 100


def nominal_change_opt(values: list[float | None]) -> float | None:
    if values[0] is None or values[-1] is None:
        return None
    return values[-1] - values[0]


def ratio(part: float | None, total: float | None) -> float | None:
    if part is None or total in (None, 0):
        return None
    return part / total * 100


def series_max(years: list[int], values: list[float | None]) -> tuple[int, float]:
    pairs = [(year, value) for year, value in zip(years, values) if value is not None]
    return max(pairs, key=lambda item: item[1])


def series_min(years: list[int], values: list[float | None]) -> tuple[int, float]:
    pairs = [(year, value) for year, value in zip(years, values) if value is not None]
    return min(pairs, key=lambda item: item[1])


def ensure_dirs() -> None:
    BUILD_DIR.mkdir(exist_ok=True)
    CHART_DIR.mkdir(exist_ok=True)


class MatrixSheet:
    def __init__(self, workbook, sheet_name: str) -> None:
        ws = workbook[sheet_name]
        self.years = [int(ws.cell(2, column).value) for column in range(2, 8)]
        self.rows: dict[str, tuple[str, list[float | None]]] = {}
        for row in range(3, ws.max_row + 1):
            label = ws.cell(row, 1).value
            if not label:
                continue
            values = []
            for column in range(2, 8):
                value = ws.cell(row, column).value
                values.append(None if value is None else float(value))
            self.rows[normalize(label)] = (str(label), values)

    def get(self, label: str) -> list[float | None]:
        for wanted in search_terms(label):
            for key, (_, values) in self.rows.items():
                if wanted == key or wanted in key or key in wanted:
                    return values
        raise KeyError(label)

    def label(self, label: str) -> str:
        for wanted in search_terms(label):
            for key, (raw_label, _) in self.rows.items():
                if wanted == key or wanted in key or key in wanted:
                    return raw_label
        raise KeyError(label)


class IndicatorSheet:
    def __init__(self, workbook, sheet_name: str) -> None:
        ws = workbook[sheet_name]
        self.years = [int(ws.cell(1, column).value) for column in range(3, 9)]
        self.rows: dict[tuple[str, str], list[float | None]] = {}
        for row in range(2, ws.max_row + 1):
            indicator = ws.cell(row, 1).value
            ente = ws.cell(row, 2).value
            if not indicator or not ente:
                continue
            values = []
            for column in range(3, 9):
                value = ws.cell(row, column).value
                values.append(None if value is None else float(value))
            self.rows[(normalize(indicator), normalize(ente))] = values

    def get(self, indicator: str, ente: str) -> list[float | None]:
        wanted_indicator = normalize(indicator)
        wanted_ente = normalize(ente)
        for (indicator_key, ente_key), values in self.rows.items():
            if (wanted_indicator == indicator_key or wanted_indicator in indicator_key or indicator_key in wanted_indicator) and (
                wanted_ente == ente_key or wanted_ente in ente_key or ente_key in wanted_ente
            ):
                return values
        raise KeyError((indicator, ente))


class DebtCoverPage(tpl.Flowable):
    def __init__(self, title: str, subtitle: str, image_path: Path, members: list[str]) -> None:
        super().__init__()
        self.title = title
        self.subtitle = subtitle
        self.image_path = image_path
        self.members = members

    def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:
        return tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT

    def draw(self) -> None:
        canvas = self.canv
        tpl.draw_cover_image(canvas, self.image_path, 0, 0, tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT)
        canvas.saveState()
        canvas.setFillColor(tpl.colors.Color(0.05, 0.15, 0.24, alpha=0.86))
        canvas.rect(0, 0, tpl.PAGE_WIDTH * 0.42, tpl.PAGE_HEIGHT, fill=1, stroke=0)
        canvas.setFillColor(tpl.WHITE)
        canvas.setFont(tpl.FONT_BOLD, 11)
        canvas.drawString(32, tpl.PAGE_HEIGHT - 42, str(canvas.getPageNumber()))
        canvas.setFont(tpl.FONT_REGULAR, 9)
        canvas.drawString(56, tpl.PAGE_HEIGHT - 42, "Boletim de Diagnóstico do Endividamento 2026")
        canvas.drawRightString(tpl.PAGE_WIDTH - 28, tpl.PAGE_HEIGHT - 42, "Projeto Mini Ministério da Fazenda")
        canvas.setFillColor(tpl.LIME)
        canvas.rect(32, tpl.PAGE_HEIGHT - 156, 92, 8, fill=1, stroke=0)

        title_text = canvas.beginText(32, tpl.PAGE_HEIGHT - 188)
        title_text.setFont(tpl.FONT_BOLD, 24)
        title_text.setFillColor(tpl.WHITE)
        for line in self.title.split("\n"):
            title_text.textLine(line)
        canvas.drawText(title_text)

        subtitle_text = canvas.beginText(32, tpl.PAGE_HEIGHT - 300)
        subtitle_text.setFont(tpl.FONT_REGULAR, 12)
        subtitle_text.setFillColor(tpl.colors.HexColor("#DCE7EF"))
        for line in self.subtitle.split("\n"):
            subtitle_text.textLine(line)
        canvas.drawText(subtitle_text)

        member_header = canvas.beginText(32, tpl.PAGE_HEIGHT - 372)
        member_header.setFont(tpl.FONT_BOLD, 10.5)
        member_header.setFillColor(tpl.WHITE)
        member_header.textLine("Equipe")
        canvas.drawText(member_header)

        member_text = canvas.beginText(32, tpl.PAGE_HEIGHT - 390)
        member_text.setFont(tpl.FONT_REGULAR, 9.8)
        member_text.setFillColor(tpl.colors.HexColor("#DCE7EF"))
        for member in self.members:
            member_text.textLine(member)
        canvas.drawText(member_text)

        canvas.setFillColor(tpl.WHITE)
        canvas.rect(0, 0, tpl.PAGE_WIDTH, 56, fill=1, stroke=0)
        canvas.setFillColor(tpl.NAVY)
        canvas.setFont(tpl.FONT_BOLD, 11)
        canvas.drawString(32, 34, "Pontifícia Universidade Católica de Minas Gerais")
        canvas.setFont(tpl.FONT_REGULAR, 9.5)
        canvas.drawString(32, 19, "Graduação em Ciências Econômicas | Eixo 4 - Projeto Mini Ministério da Fazenda")
        canvas.drawRightString(tpl.PAGE_WIDTH - 32, 27, "Etapa 2 | Abril de 2026")
        canvas.restoreState()


class DebtSectionPage(tpl.Flowable):
    def __init__(self, number: str, title: str, subtitle: str, image_path: Path, bookmark_name: str) -> None:
        super().__init__()
        self.number = number
        self.title = title
        self.subtitle = subtitle
        self.image_path = image_path
        self.toc_level = 0
        self.toc_title = title
        self.bookmark_name = bookmark_name

    def wrap(self, avail_width: float, avail_height: float) -> tuple[float, float]:
        return tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT

    def draw(self) -> None:
        canvas = self.canv
        tpl.draw_cover_image(canvas, self.image_path, 0, 0, tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT)
        canvas.saveState()
        canvas.setFillColor(tpl.colors.Color(0.04, 0.12, 0.20, alpha=0.68))
        canvas.rect(0, 0, tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT, fill=1, stroke=0)
        canvas.setFillColor(tpl.GOLD)
        canvas.rect(0, tpl.PAGE_HEIGHT - 86, tpl.PAGE_WIDTH, 10, fill=1, stroke=0)
        canvas.setFillColor(tpl.WHITE)
        canvas.setFont(tpl.FONT_BOLD, 34)
        canvas.drawString(42, tpl.PAGE_HEIGHT - 176, self.number)
        canvas.setFont(tpl.FONT_REGULAR, 9)
        canvas.drawRightString(tpl.PAGE_WIDTH - 28, tpl.PAGE_HEIGHT - 42, "Relatório temático")
        title_text = canvas.beginText(42, tpl.PAGE_HEIGHT - 214)
        title_text.setFont(tpl.FONT_BOLD, 24)
        title_text.setFillColor(tpl.WHITE)
        for line in self.title.split("\n"):
            title_text.textLine(line)
        canvas.drawText(title_text)
        subtitle_text = canvas.beginText(42, tpl.PAGE_HEIGHT - 266)
        subtitle_text.setFont(tpl.FONT_REGULAR, 11)
        subtitle_text.setFillColor(tpl.colors.HexColor("#E0E8EE"))
        for line in self.subtitle.split("\n"):
            subtitle_text.textLine(line)
        canvas.drawText(subtitle_text)
        canvas.bookmarkPage(self.bookmark_name)
        canvas.restoreState()


def draw_body_page(canvas, doc) -> None:
    page_number = canvas.getPageNumber()
    canvas.saveState()
    canvas.setFillColor(tpl.NAVY)
    canvas.setFont(tpl.FONT_BOLD, 10)
    canvas.drawString(tpl.LEFT_MARGIN, tpl.PAGE_HEIGHT - 24, str(page_number))
    canvas.setFont(tpl.FONT_REGULAR, 8.5)
    canvas.drawString(tpl.LEFT_MARGIN + 20, tpl.PAGE_HEIGHT - 24, "Boletim de Diagnóstico do Endividamento de Mato Grosso do Sul e Campo Grande")
    canvas.drawRightString(tpl.PAGE_WIDTH - tpl.RIGHT_MARGIN, tpl.PAGE_HEIGHT - 24, "Projeto Mini Ministério da Fazenda")
    canvas.setStrokeColor(tpl.LIGHT_LINE)
    canvas.setLineWidth(0.8)
    canvas.line(tpl.LEFT_MARGIN, tpl.PAGE_HEIGHT - 30, tpl.PAGE_WIDTH - tpl.RIGHT_MARGIN, tpl.PAGE_HEIGHT - 30)
    canvas.line(tpl.LEFT_MARGIN, 24, tpl.PAGE_WIDTH - tpl.RIGHT_MARGIN, 24)
    canvas.setFillColor(tpl.SLATE)
    canvas.setFont(tpl.FONT_REGULAR, 7.8)
    canvas.drawString(tpl.LEFT_MARGIN, 12, "Elaboração própria a partir de dados do SICONFI/Tesouro Nacional.")
    canvas.drawRightString(tpl.PAGE_WIDTH - tpl.RIGHT_MARGIN, 12, "Ano-base 2024")
    canvas.restoreState()


class DebtDocTemplate(tpl.BaseDocTemplate):
    def __init__(self, filename: str) -> None:
        super().__init__(
            filename,
            pagesize=(tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT),
            leftMargin=tpl.LEFT_MARGIN,
            rightMargin=tpl.RIGHT_MARGIN,
            topMargin=tpl.TOP_MARGIN,
            bottomMargin=tpl.BOTTOM_MARGIN,
            title="Diagnóstico do Endividamento do Mato Grosso do Sul e de Campo Grande",
            author="PUC Minas | Curso de Ciências Econômicas | Eixo 4",
        )

        body_frame = tpl.Frame(
            tpl.LEFT_MARGIN,
            tpl.BOTTOM_MARGIN,
            tpl.BODY_WIDTH,
            tpl.BODY_HEIGHT,
            id="body",
            leftPadding=0,
            rightPadding=0,
            topPadding=0,
            bottomPadding=0,
        )
        full_frame = tpl.Frame(0, 0, tpl.PAGE_WIDTH, tpl.PAGE_HEIGHT, id="full", leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)

        self.addPageTemplates(
            [
                tpl.PageTemplate(id="FullPage", frames=[full_frame]),
                tpl.PageTemplate(id="Body", frames=[body_frame], onPage=draw_body_page),
            ]
        )

    def afterFlowable(self, flowable) -> None:
        if not hasattr(flowable, "toc_level"):
            return
        title = getattr(flowable, "toc_title", None)
        if title is None and hasattr(flowable, "getPlainText"):
            title = flowable.getPlainText()
        bookmark = getattr(flowable, "bookmark_name", None)
        if not title or not bookmark:
            return
        self.canv.bookmarkPage(bookmark)
        self.notify("TOCEntry", (flowable.toc_level, title, self.page, bookmark))
        try:
            self.canv.addOutlineEntry(title, bookmark, level=flowable.toc_level, closed=False)
        except ValueError:
            pass


def variation_table_from_structure(sheet: MatrixSheet, structure: Iterable[tuple[str, str | None, bool]]) -> tpl.Table:
    header = ["Discriminação", *[str(year) for year in sheet.years], "Variação\nNominal", "Variação\n(%)"]
    body = [header]
    emphasis_rows = []
    separator_rows = []

    for row_index, (label, _, is_total) in enumerate(structure, start=1):
        if "PASSIVOS FORA DA DC" in label:
            body.append([label, "", "", "", "", "", "", "", ""])
            separator_rows.append(row_index)
            continue

        values = sheet.get(label)
        nominal = nominal_change_opt(values)
        pct_var = pct_change_opt(values)
        body.append(
            [
                sheet.label(label),
                *[fmt_million(value, 2) for value in values],
                fmt_million(nominal, 2),
                "–" if pct_var is None else tpl.br_percent(pct_var, 1),
            ]
        )
        if is_total:
            emphasis_rows.append(row_index)

    table = tpl.Table(body, colWidths=[258, 58, 58, 58, 58, 58, 58, 78, 66], repeatRows=1, splitByRow=0)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), tpl.NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), tpl.WHITE),
        ("FONTNAME", (0, 0), (-1, 0), tpl.FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 7.1),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), tpl.FONT_REGULAR),
        ("FONTSIZE", (0, 1), (-1, -1), 6.8),
        ("LEADING", (0, 1), (-1, -1), 7.8),
        ("GRID", (0, 0), (-1, -1), 0.35, tpl.LIGHT_LINE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [tpl.WHITE, tpl.LIGHT_BG]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for row_number in emphasis_rows:
        commands.extend(
            [
                ("FONTNAME", (0, row_number), (-1, row_number), tpl.FONT_BOLD),
                ("BACKGROUND", (0, row_number), (-1, row_number), tpl.colors.HexColor("#E6EEF5")),
            ]
        )
    for row_number in separator_rows:
        commands.extend(
            [
                ("FONTNAME", (0, row_number), (-1, row_number), tpl.FONT_BOLD),
                ("BACKGROUND", (0, row_number), (-1, row_number), tpl.colors.HexColor("#EEF1F4")),
            ]
        )
    table.setStyle(tpl.TableStyle(commands))
    return table


def series_table(headers: list[str], rows: list[list[str]], col_widths: list[float]) -> tpl.Table:
    header_style = tpl.ParagraphStyle("DebtHeader", fontName=tpl.FONT_BOLD, fontSize=7.5, leading=8.6, textColor=tpl.WHITE)
    cell_style = tpl.ParagraphStyle("DebtCell", fontName=tpl.FONT_REGULAR, fontSize=7.2, leading=8.2, textColor=tpl.TEXT)
    data = [[tpl.Paragraph(text, header_style) for text in headers]]
    for row in rows:
        data.append([tpl.Paragraph(text, cell_style) for text in row])
    table = tpl.Table(data, colWidths=col_widths, repeatRows=1, splitByRow=0)
    table.setStyle(
        tpl.TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), tpl.NAVY),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [tpl.WHITE, tpl.LIGHT_BG]),
                ("GRID", (0, 0), (-1, -1), 0.35, tpl.LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def keep_table_block(caption_flowables: list, table: tpl.Table) -> KeepTogether:
    return KeepTogether([*caption_flowables, table])


def caption_block_unit(kind: str, number: int, title: str, unit: str, source: str, styles) -> list:
    return [
        tpl.Paragraph(f"{kind} {number}", styles["CaptionTitle"]),
        tpl.Paragraph(
            f"{title}<br/>Dados em: {unit}<br/>Elaboração própria<br/>Fonte: {source}",
            styles["CaptionMeta"],
        ),
    ]


def stacked_bar_chart_side_legend(
    path: Path,
    years: list[int],
    stacks: list[tuple[str, list[float], str]],
    title: str,
    ylabel: str,
    divisor: float = 1.0,
    line_item: tuple[str, list[float], str] | None = None,
) -> None:
    fig, ax = tpl.plt.subplots(figsize=(10.4, 3.2), dpi=180)
    bottom = [0.0] * len(years)
    for label, values, color in stacks:
        ax.bar(years, values, bottom=bottom, width=0.58, color=color, label=label)
        bottom = [bottom[idx] + values[idx] for idx in range(len(values))]
    if line_item is not None:
        label, values, color = line_item
        ax.plot(years, values, color=color, linewidth=2.4, marker="o", label=label)
    ax.set_title(title, loc="left", color="#0C2742", pad=10)
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(tpl.axis_formatter(divisor))
    ax.set_ylabel(ylabel)
    ax.grid(axis="y")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(
        frameon=False,
        loc="center left",
        bbox_to_anchor=(1.01, 0.5),
        ncol=1,
        borderaxespad=0.0,
    )
    fig.subplots_adjust(right=0.78)
    fig.savefig(path, bbox_inches="tight")
    tpl.plt.close(fig)


def generate_charts(estado: MatrixSheet, capital: MatrixSheet, indicadores: IndicatorSheet) -> dict[str, Path]:
    tpl.setup_plot()
    charts: dict[str, Path] = {}

    anos = estado.years
    ms_dc = [value or 0 for value in estado.get("DÍVIDA CONSOLIDADA")]
    ms_dcl = [value or 0 for value in estado.get("DCL")]
    ms_rcl = [value or 0 for value in estado.get("RCL AJUSTADA")]
    ms_ded = [value or 0 for value in estado.get("DEDUÇÕES")]
    ms_emprest = [value or 0 for value in estado.get("Empréstimos")]
    ms_reestr = [value or 0 for value in estado.get("Reestruturação")]
    ms_parcel = [value or 0 for value in estado.get("Parcelamento e Renegociação")]
    ms_prec = [value or 0 for value in estado.get("Precatórios")]
    ms_passivo = [value or 0 for value in estado.get("Passivo Atuarial")]
    ms_dcl_pct = [value or 0 for value in indicadores.get("DCL / RCL Ajustada", "MS")]
    ms_passivo_pct = [value or 0 for value in indicadores.get("Passivo Atuarial", "MS")]
    ms_liq_bruta_pct = [value or 0 for value in indicadores.get("Disp. de Caixa Bruta", "MS")]
    ms_liq_antes_pct = [value or 0 for value in indicadores.get("Disp. de Caixa Líquida (antes RP)", "MS")]
    ms_liq_apos_pct = [value or 0 for value in indicadores.get("Disp. de Caixa Líquida (após RP)", "MS")]

    cg_dc = [value or 0 for value in capital.get("DÍVIDA CONSOLIDADA")]
    cg_dcl = [value or 0 for value in capital.get("DCL")]
    cg_fin = [value or 0 for value in capital.get("Financiamentos")]
    cg_parcel = [value or 0 for value in capital.get("Parcelamento e Renegociação")]
    cg_rcl = [value or 0 for value in capital.get("RCL AJUSTADA")]
    cg_dcl_pct = [value or 0 for value in indicadores.get("DCL / RCL Ajustada", "Campo Grande")]

    charts["ms_dcl_ratio"] = CHART_DIR / "grafico_ms_dcl_ratio.png"
    tpl.combo_bar_line_chart(
        charts["ms_dcl_ratio"],
        anos,
        ms_dcl,
        ms_dcl_pct,
        "DCL do Estado e relação DCL / RCL ajustada",
        "DCL (R$ milhões)",
        "DCL / RCL ajustada (%)",
        bar_divisor=1.0,
    )

    charts["cg_dcl_ratio"] = CHART_DIR / "grafico_cg_dcl_ratio.png"
    tpl.combo_bar_line_chart(
        charts["cg_dcl_ratio"],
        anos,
        cg_dcl,
        cg_dcl_pct,
        "DCL de Campo Grande e relação DCL / RCL ajustada",
        "DCL (R$ milhões)",
        "DCL / RCL ajustada (%)",
        bar_divisor=1.0,
    )

    charts["ms_dc_composicao"] = CHART_DIR / "grafico_ms_dc_composicao.png"
    stacked_bar_chart_side_legend(
        charts["ms_dc_composicao"],
        anos,
        [
            ("Empréstimos", ms_emprest, "#2E7B83"),
            ("Reestruturação", ms_reestr, "#0C2742"),
            ("Parcelamento", ms_parcel, "#C49A45"),
            ("Precatórios", ms_prec, "#B85042"),
        ],
        "Composição da dívida consolidada do Estado",
        "R$ bilhões",
        divisor=1000.0,
        line_item=("Dívida consolidada", ms_dc, "#6EB63D"),
    )

    charts["ms_estoque_ajuste"] = CHART_DIR / "grafico_ms_estoque_ajuste.png"
    tpl.line_chart(
        charts["ms_estoque_ajuste"],
        anos,
        [
            ("DCL", ms_dcl, "#0C2742"),
            ("Deduções", ms_ded, "#2E7B83"),
            ("RCL ajustada", ms_rcl, "#C49A45"),
        ],
        "DCL, deduções e RCL ajustada do Estado",
        "R$ bilhões",
        divisor=1000.0,
    )

    charts["ms_passivo"] = CHART_DIR / "grafico_ms_passivo.png"
    tpl.combo_bar_line_chart(
        charts["ms_passivo"],
        anos,
        ms_passivo,
        ms_passivo_pct,
        "Passivo atuarial do Estado e proporção sobre a RCL ajustada",
        "Passivo atuarial (R$ milhões)",
        "Passivo / RCL ajustada (%)",
        bar_divisor=1.0,
    )

    charts["ms_liquidez"] = CHART_DIR / "grafico_ms_liquidez.png"
    tpl.line_chart(
        charts["ms_liquidez"],
        anos,
        [
            ("Disp. de caixa bruta / RCL", ms_liq_bruta_pct, "#0C2742"),
            ("Disp. líquida antes RP / RCL", ms_liq_antes_pct, "#2E7B83"),
            ("Disp. líquida após RP / RCL", ms_liq_apos_pct, "#6EB63D"),
        ],
        "Indicadores de liquidez do Estado",
        "% da RCL ajustada",
        divisor=1.0,
    )

    charts["cg_dc_composicao"] = CHART_DIR / "grafico_cg_dc_composicao.png"
    tpl.stacked_bar_chart(
        charts["cg_dc_composicao"],
        anos,
        [
            ("Financiamentos", cg_fin, "#0C2742"),
            ("Parcelamento", cg_parcel, "#C49A45"),
        ],
        "Composição da dívida consolidada de Campo Grande",
        "R$ milhões",
        divisor=1.0,
        line_item=("Dívida consolidada", cg_dc, "#6EB63D"),
    )

    charts["cg_estoque_ajuste"] = CHART_DIR / "grafico_cg_estoque_ajuste.png"
    tpl.line_chart(
        charts["cg_estoque_ajuste"],
        anos,
        [
            ("DCL", cg_dcl, "#0C2742"),
            ("Dívida consolidada", cg_dc, "#2E7B83"),
            ("RCL ajustada", cg_rcl, "#C49A45"),
        ],
        "DCL, dívida consolidada e RCL ajustada de Campo Grande",
        "R$ milhões",
        divisor=1.0,
    )

    return charts


def build_story(estado: MatrixSheet, capital: MatrixSheet, indicadores: IndicatorSheet, charts: dict[str, Path], styles) -> list:
    story: list = []

    anos = estado.years
    ms_dc = estado.get("DÍVIDA CONSOLIDADA")
    ms_dcl = estado.get("DCL")
    ms_rcl_bruta = estado.get("RCL")
    ms_rcl = estado.get("RCL AJUSTADA")
    ms_ded = estado.get("DEDUÇÕES")
    ms_passivo = estado.get("Passivo Atuarial")
    ms_rpnp = estado.get("RP Não-Processados")
    ms_reestr = estado.get("Reestruturação")
    ms_prec = estado.get("Precatórios")

    cg_dc = capital.get("DÍVIDA CONSOLIDADA")
    cg_dcl = capital.get("DCL")
    cg_rcl_bruta = capital.get("RCL")
    cg_rcl = capital.get("RCL AJUSTADA")
    cg_fin = capital.get("Financiamentos")
    cg_parcel = capital.get("Parcelamento e Renegociação")

    ms_dcl_pct = indicadores.get("DCL / RCL Ajustada", "MS")
    cg_dcl_pct = indicadores.get("DCL / RCL Ajustada", "Campo Grande")
    ms_passivo_pct = indicadores.get("Passivo Atuarial", "MS")
    ms_rpnp_pct = indicadores.get("RP Não-Processados", "MS")
    ms_depositos_judiciais_pct = indicadores.get("Apropr. Depósitos Judiciais", "MS")
    cg_passivo_pct = indicadores.get("Passivo Atuarial", "Campo Grande")
    cg_rpnp_pct = indicadores.get("RP Não-Processados", "Campo Grande")
    cg_depositos_judiciais_pct = indicadores.get("Apropr. Depósitos Judiciais", "Campo Grande")
    ms_liq_apos_pct = indicadores.get("Disp. de Caixa Líquida (após RP)", "MS")
    ms_liq_bruta_pct = indicadores.get("Disp. de Caixa Bruta", "MS")
    ms_liq_antes_pct = indicadores.get("Disp. de Caixa Líquida (antes RP)", "MS")
    servico_divida = extrair_servico_divida_etapa1()
    ms_servico = [servico_divida["MS (Estado)"][ano]["servico_divida"] / 1_000_000 for ano in anos]
    ms_juros = [servico_divida["MS (Estado)"][ano]["juros_encargos"] / 1_000_000 for ano in anos]
    ms_amortizacao = [servico_divida["MS (Estado)"][ano]["amortizacao_divida"] / 1_000_000 for ano in anos]
    cg_servico = [servico_divida["Campo Grande"][ano]["servico_divida"] / 1_000_000 for ano in anos]
    cg_juros = [servico_divida["Campo Grande"][ano]["juros_encargos"] / 1_000_000 for ano in anos]
    cg_amortizacao = [servico_divida["Campo Grande"][ano]["amortizacao_divida"] / 1_000_000 for ano in anos]
    ms_pessoal = [servico_divida["MS (Estado)"][ano]["pessoal_encargos"] / 1_000_000 for ano in anos]
    cg_pessoal = [servico_divida["Campo Grande"][ano]["pessoal_encargos"] / 1_000_000 for ano in anos]
    servico_divida_ms_pct = [ratio(valor, rcl) for valor, rcl in zip(ms_servico, ms_rcl)]
    servico_divida_cg_pct = [ratio(valor, rcl) for valor, rcl in zip(cg_servico, cg_rcl)]
    ms_pessoal_pct = [ratio(valor, rcl) for valor, rcl in zip(ms_pessoal, ms_rcl_bruta)]
    cg_pessoal_pct = [ratio(valor, rcl) for valor, rcl in zip(cg_pessoal, cg_rcl_bruta)]

    ms_dcl_min_year, ms_dcl_min_value = series_min(anos, ms_dcl)
    ms_passivo_max_year, ms_passivo_max_value = series_max(anos, ms_passivo)
    cg_dcl_max_year, cg_dcl_max_value = series_max(anos, cg_dcl)

    cover_title = "Diagnóstico do Endividamento\nMato Grosso do Sul\ne Campo Grande"
    cover_subtitle = (
        "Ano-base 2024 | Série histórica 2019-2024\n"
        "Etapa 2 do Projeto Mini Ministério da Fazenda\n"
        "Curso de Ciências Econômicas | Eixo 4 | PUC Minas"
    )

    story.extend([tpl.NextPageTemplate("FullPage"), DebtCoverPage(cover_title, cover_subtitle, IMG_MS, TEAM_MEMBERS), tpl.NextPageTemplate("Body"), tpl.PageBreak()])
    story.append(tpl.Paragraph("Expediente", styles["H1"]))
    story.append(
        tpl.Paragraph(
            "Este relatório refere-se à Etapa 2 do Projeto Mini Ministério da Fazenda, desenvolvido no Curso de Ciências Econômicas da PUC Minas, no âmbito do Eixo 4. "
            "A análise examina a trajetória do endividamento do Estado do Mato Grosso do Sul e do Município de Campo Grande entre 2019 e 2024, com base nos dados do SICONFI/Tesouro Nacional e nos demonstrativos do Relatório de Gestão Fiscal.",
            styles["Lead"],
        )
    )
    story.append(
        tpl.Paragraph(
            "No contexto do projeto, o diagnóstico foi estruturado para integrar frentes complementares de consolidação dos dados, leitura dos limites normativos, interpretação econômica e padronização editorial, de modo a apoiar a construção coletiva de um produto analítico único.",
            styles["Body"],
        )
    )
    expediente = tpl.Table(
        [
            [
                tpl.Paragraph(
                    "<b>Instituição e equipe</b><br/>Pontifícia Universidade Católica de Minas Gerais (PUC Minas)<br/>"
                    "Curso de Ciências Econômicas<br/>Eixo 4 - Projeto Mini Ministério da Fazenda<br/>"
                    "Etapa 2 - Diagnóstico do Endividamento do Mato Grosso do Sul e de sua Capital<br/><br/>"
                    + "<br/>".join(TEAM_MEMBERS),
                    styles["Body"],
                ),
                tpl.Paragraph(
                    "<b>Fontes, método e recorte</b><br/>Base quantitativa: SICONFI/Tesouro Nacional, com série de 2019 a 2024, em valores nominais de R$ milhões. "
                    "Os dados foram organizados em quadro estatístico consolidado pela equipe a partir dessas fontes.<br/><br/>"
                    "Estrutura analítica: decomposição da dívida consolidada, deduções, dívida consolidada líquida, RCL ajustada, serviço da dívida, despesa com pessoal, passivos fora da dívida consolidada e indicadores de liquidez.<br/><br/>"
                    "Referencial editorial: Boletim de Finanças dos Entes Subnacionais 2025, da Secretaria do Tesouro Nacional.<br/><br/>Data de fechamento: abril de 2026.",
                    styles["Body"],
                ),
            ]
        ],
        colWidths=[tpl.BODY_WIDTH / 2 - 10, tpl.BODY_WIDTH / 2 - 10],
        hAlign="LEFT",
    )
    expediente.setStyle(
        tpl.TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, tpl.LIGHT_LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.8, tpl.LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend(
        [
            expediente,
            tpl.Spacer(1, 10),
            tpl.bullet_box(
                "Chaves de leitura",
                [
                    f"No Estado, a relação DCL / RCL ajustada caiu de {tpl.br_percent(ms_dcl_pct[0], 2)} em 2019 para {tpl.br_percent(ms_dcl_pct[-1], 2)} em 2024, apesar da alta de {tpl.br_percent(pct_change_opt(ms_dc) or 0, 1)} da dívida consolidada.",
                    f"Em Campo Grande, a dívida consolidada líquida avançou para {tpl.br_money_text(cg_dcl[-1])} em 2024, mas o indicador DCL / RCL ajustada permaneceu muito baixo, em {tpl.br_percent(cg_dcl_pct[-1], 2)}.",
                    f"No Estado, a liquidez após restos a pagar passou de {tpl.br_percent(ms_liq_apos_pct[0], 2)} da RCL ajustada em 2019 para {tpl.br_percent(ms_liq_apos_pct[-1], 2)} em 2024, mantendo-se positiva ao fim da série.",
                ],
                styles,
            ),
            tpl.PageBreak(),
        ]
    )

    story.append(tpl.Paragraph("Sumário", styles["TOCTitle"]))
    toc = tpl.TableOfContents()
    toc.levelStyles = [
        tpl.ParagraphStyle(name="TOCLevel0", fontName=tpl.FONT_BOLD, fontSize=10, leading=13, textColor=tpl.NAVY),
        tpl.ParagraphStyle(name="TOCLevel1", fontName=tpl.FONT_REGULAR, fontSize=8.9, leading=11.4, leftIndent=18, textColor=tpl.TEXT),
        tpl.ParagraphStyle(name="TOCLevel2", fontName=tpl.FONT_REGULAR, fontSize=8.2, leading=10.8, leftIndent=34, textColor=tpl.SLATE),
    ]
    story.append(toc)

    story.extend(
        [
            tpl.NextPageTemplate("FullPage"),
            tpl.PageBreak(),
            DebtSectionPage("1", "1.0 - Introdução", "Objetivos, conceitos de referência e principais resultados do diagnóstico\nsobre endividamento, passivos e liquidez do estado e da capital.", IMG_MS, "sec_intro"),
            tpl.NextPageTemplate("Body"),
            tpl.PageBreak(),
        ]
    )

    story.append(tpl.make_heading("1.1 - Escopo analítico", styles["H2"], 1, "h2_11"))
    story.append(
        tpl.Paragraph(
            "Este relatório examina a trajetória da dívida consolidada, das deduções, da dívida consolidada líquida e da receita corrente líquida ajustada do Estado do Mato Grosso do Sul e do Município de Campo Grande no período de 2019 a 2024. "
            "A leitura destaca a composição das obrigações, o peso relativo do endividamento frente à base corrente de receitas e, no caso do Estado, a articulação entre passivos adicionais e liquidez.",
            styles["Lead"],
        )
    )
    story.append(
        tpl.Paragraph(
            "A estrutura do documento foi organizada para aproximar o tema do endividamento da leitura fiscal mais ampla desenvolvida na etapa anterior. "
            "Por isso, o foco não recai apenas sobre o estoque bruto da dívida, mas também sobre o papel das deduções financeiras, da RCL ajustada, do passivo atuarial e da disponibilidade de caixa na avaliação da posição fiscal dos entes.",
            styles["Body"],
        )
    )

    cards = tpl.Table(
        [
            [
                tpl.metric_card("Mato Grosso do Sul: DCL / RCL ajustada em 2024", tpl.br_percent(ms_dcl_pct[-1], 2), "Indicador muito abaixo do limite de referência de 200%.", styles),
                tpl.metric_card("Campo Grande: DCL / RCL ajustada em 2024", tpl.br_percent(cg_dcl_pct[-1], 2), "Indicador bastante inferior ao limite de referência de 120%.", styles),
                tpl.metric_card("Passivo atuarial do Estado em 2024", tpl.br_money_text(ms_passivo[-1]), f"Correspondeu a {tpl.br_percent(ms_passivo_pct[-1], 2)} da RCL ajustada.", styles),
                tpl.metric_card("Liquidez após RP do Estado em 2024", tpl.br_percent(ms_liq_apos_pct[-1], 2), "A disponibilidade líquida seguiu positiva no encerramento do período.", styles),
            ]
        ],
        colWidths=[tpl.BODY_WIDTH / 4 - 8] * 4,
        hAlign="LEFT",
    )
    cards.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 6)]))
    story.extend([cards, tpl.Spacer(1, 12)])
    story.append(
        tpl.overview_table(
            [
                ("Dívida consolidada 2024", tpl.br_money_text(ms_dc[-1]), tpl.br_money_text(cg_dc[-1]), "O estado encerrou 2024 com estoque muito superior em termos absolutos, mas a composição e a folga frente ao limite são distintas."),
                ("Dívida consolidada líquida 2024", tpl.br_money_text(ms_dcl[-1]), tpl.br_money_text(cg_dcl[-1]), "As deduções reduziram de forma relevante o endividamento líquido dos dois entes, sobretudo no Estado."),
                ("DCL / RCL ajustada 2024", tpl.br_percent(ms_dcl_pct[-1], 2), tpl.br_percent(cg_dcl_pct[-1], 2), "Os dois entes permaneceram abaixo dos limites de referência, com situação especialmente folgada em Campo Grande."),
                ("Variação da DCL 2019-2024", tpl.br_percent(pct_change_opt(ms_dcl) or 0, 1), tpl.br_percent(pct_change_opt(cg_dcl) or 0, 1), "O Estado reduziu a DCL no horizonte completo; a capital elevou o indicador a partir de uma base historicamente baixa."),
            ]
        )
    )
    story.extend(
        [
            tpl.Spacer(1, 10),
            tpl.bullet_box(
                "Principais conclusões preliminares",
                [
                    f"No Estado, a DCL caiu {tpl.br_percent(pct_change_opt(ms_dcl) or 0, 1)} entre 2019 e 2024, enquanto a RCL ajustada cresceu {tpl.br_percent(pct_change_opt(ms_rcl) or 0, 1)}.",
                    f"Em Campo Grande, a dívida consolidada líquida aumentou {tpl.br_percent(pct_change_opt(cg_dcl) or 0, 1)}, mas a relação DCL / RCL ajustada ficou em apenas {tpl.br_percent(cg_dcl_pct[-1], 2)} em 2024.",
                    f"No Estado, o passivo atuarial atingiu pico de {tpl.br_money_text(ms_passivo_max_value)} em {ms_passivo_max_year}, permanecendo como passivo relevante mesmo após recuo recente.",
                ],
                styles,
            ),
            tpl.PageBreak(),
        ]
    )

    story.append(tpl.make_heading("1.2 - Conceitos e limites de referência", styles["H2"], 1, "h2_12"))
    story.append(
        tpl.Paragraph(
            "A dívida consolidada representa o estoque bruto das obrigações financeiras de prazo mais longo, enquanto a dívida consolidada líquida desconta deduções como disponibilidade de caixa e demais haveres financeiros. "
            "A comparação entre DCL e RCL ajustada é o indicador central da leitura prudencial do endividamento, pois relaciona o passivo líquido à capacidade corrente de geração de receitas.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            "Nos termos da Resolução do Senado Federal nº 40, de 20 de dezembro de 2001, os limites de referência para a relação entre dívida consolidada líquida e receita corrente líquida são de 200% para Estados e de 120% para Municípios. "
            "Os dados consolidados a partir do SICONFI mostram que Mato Grosso do Sul e Campo Grande permaneceram abaixo desses parâmetros em toda a série observada.",
            styles["Body"],
        )
    )
    concept_panels = tpl.Table(
        [
            [
                tpl.info_panel(
                    "Como ler o indicador",
                    "Uma relação DCL / RCL ajustada menor sinaliza menor comprometimento do passivo líquido frente à receita corrente. O movimento do indicador pode resultar tanto de variações na dívida quanto de mudanças na base de receita e nas deduções.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
                tpl.info_panel(
                    "Leitura complementar",
                    "O diagnóstico fica mais robusto quando a DCL é observada em conjunto com a composição da dívida, o passivo atuarial e, no caso do Estado, com os indicadores de liquidez após restos a pagar.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    concept_panels.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([concept_panels, tpl.Spacer(1, 10)])
    intro_chart_grid = tpl.Table(
        [[tpl.scaled_image(charts["ms_dcl_ratio"], tpl.BODY_WIDTH / 2 - 8, 185), tpl.scaled_image(charts["cg_dcl_ratio"], tpl.BODY_WIDTH / 2 - 8, 185)]],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    intro_chart_grid.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend(
        tpl.caption_block_with_source(
            "Gráfico",
            1,
            "Dívida consolidada líquida e relação DCL / RCL ajustada do Estado e de Campo Grande",
            "SICONFI. Relatório de Gestão Fiscal - Demonstrativo da Dívida Consolidada Líquida.",
            styles,
        )
    )
    story.append(intro_chart_grid)
    story.append(
        keep_table_block(
            caption_block_unit(
                "Tabela",
                1,
                "Indicadores comparados de endividamento e liquidez",
                "% da RCL ou da RCL ajustada, conforme indicador",
                "SICONFI. Relatório de Gestão Fiscal - Demonstrativos da Dívida Consolidada Líquida e da Disponibilidade de Caixa e dos Restos a Pagar.",
                styles,
            ),
            series_table(
                ["Indicador", "2019", "2020", "2021", "2022", "2023", "2024"],
                [
                    ["DCL / RCL ajustada - MS (%)", *[tpl.br_percent(value, 2) for value in ms_dcl_pct]],
                    ["DCL / RCL ajustada - Campo Grande (%)", *[tpl.br_percent(value, 2) for value in cg_dcl_pct]],
                    ["Serviço da dívida / RCL ajustada - MS (%)", *[tpl.br_percent(value, 2) for value in servico_divida_ms_pct]],
                    ["Serviço da dívida / RCL ajustada - Campo Grande (%)", *[tpl.br_percent(value, 2) for value in servico_divida_cg_pct]],
                    ["Despesa com pessoal / RCL - MS (%)", *[tpl.br_percent(value, 2) for value in ms_pessoal_pct]],
                    ["Despesa com pessoal / RCL - Campo Grande (%)", *[tpl.br_percent(value, 2) for value in cg_pessoal_pct]],
                    ["Passivo atuarial / RCL ajustada - MS (%)", *[tpl.br_percent(value, 2) for value in ms_passivo_pct]],
                    ["Disp. líquida após RP / RCL ajustada - MS (%)", *[tpl.br_percent(value, 2) for value in ms_liq_apos_pct]],
                ],
                [250, 72, 72, 72, 72, 72, 72],
            ),
        )
    )
    story.append(tpl.Spacer(1, 10))
    story.append(tpl.make_heading("1.3 - Cobertura dos indicadores obrigatórios", styles["H2"], 1, "h2_13"))
    story.append(
        tpl.Paragraph(
            "Os dados do SICONFI permitem calcular a relação DCL / RCL ajustada, o serviço da dívida, a despesa com pessoal, a liquidez e os passivos fora da dívida consolidada disponíveis nos demonstrativos. "
            "Para o indicador Serviço da Dívida / RCL Ajustada, foram usadas as linhas de juros, encargos e amortizações dos demonstrativos fiscais consolidados, preservando a série anual de 2019 a 2024.",
            styles["Body"],
        )
    )
    story.append(
        keep_table_block(
            caption_block_unit(
                "Tabela",
                2,
                "Cobertura dos indicadores solicitados nas orientações",
                "% da RCL ou da RCL ajustada, conforme indicador",
                "SICONFI. Sistema de Informações Contábeis e Fiscais do Setor Público Brasileiro. Tesouro Nacional.",
                styles,
            ),
            series_table(
                ["Indicador", "Ente", "2019", "2020", "2021", "2022", "2023", "2024"],
                [
                    ["DCL / RCL ajustada (%)", "MS", *[fmt_percent(value, 2) for value in ms_dcl_pct]],
                    ["DCL / RCL ajustada (%)", "Campo Grande", *[fmt_percent(value, 2) for value in cg_dcl_pct]],
                    ["Serviço da dívida / RCL ajustada (%)", "MS", *[fmt_percent(value, 2) for value in servico_divida_ms_pct]],
                    ["Serviço da dívida / RCL ajustada (%)", "Campo Grande", *[fmt_percent(value, 2) for value in servico_divida_cg_pct]],
                    ["Despesa com pessoal / RCL (%)", "MS", *[fmt_percent(value, 2) for value in ms_pessoal_pct]],
                    ["Despesa com pessoal / RCL (%)", "Campo Grande", *[fmt_percent(value, 2) for value in cg_pessoal_pct]],
                    ["Passivo atuarial / RCL ajustada (%)", "MS", *[fmt_percent(value, 2) for value in ms_passivo_pct]],
                    ["Passivo atuarial / RCL ajustada (%)", "Campo Grande", *[fmt_percent(value, 2) for value in cg_passivo_pct]],
                    ["RP não processados / RCL ajustada (%)", "MS", *[fmt_percent(value, 2) for value in ms_rpnp_pct]],
                    ["RP não processados / RCL ajustada (%)", "Campo Grande", *[fmt_percent(value, 2) for value in cg_rpnp_pct]],
                    ["Apropriação de depósitos judiciais / RCL ajustada (%)", "MS", *[fmt_percent(value, 2) for value in ms_depositos_judiciais_pct]],
                    ["Apropriação de depósitos judiciais / RCL ajustada (%)", "Campo Grande", *[fmt_percent(value, 2) for value in cg_depositos_judiciais_pct]],
                ],
                [210, 86, 62, 62, 62, 62, 62, 62],
            ),
        )
    )
    story.append(
        tpl.Paragraph(
            "Nota: o serviço da dívida foi calculado como (juros e encargos da dívida + amortização da dívida) / RCL ajustada x 100. A despesa com pessoal foi comparada à RCL bruta. "
            "n.d. indica que o valor não está disponível nas bases consolidadas usadas nesta etapa.",
            styles["CaptionMeta"],
        )
    )

    story.extend(
        [
            tpl.NextPageTemplate("FullPage"),
            tpl.PageBreak(),
            DebtSectionPage("2", "2.0 - Estado do Mato Grosso do Sul", "Leitura da composição da dívida, do comportamento da DCL, dos passivos adicionais\ne da liquidez do estado no período de 2019 a 2024.", IMG_MS, "sec_estado"),
            tpl.NextPageTemplate("Body"),
            tpl.PageBreak(),
        ]
    )

    story.append(tpl.make_heading("2.1 - Estoque e composição da dívida consolidada", styles["H2"], 1, "h2_21"))
    story.append(
        tpl.Paragraph(
            f"A dívida consolidada do Mato Grosso do Sul passou de {tpl.br_money_text(ms_dc[0])} em 2019 para {tpl.br_money_text(ms_dc[-1])} em 2024, crescimento nominal de {tpl.br_percent(pct_change_opt(ms_dc) or 0, 1)}. "
            "Esse avanço, contudo, não se traduziu em aumento proporcional do endividamento líquido, porque a expansão da RCL ajustada e o nível elevado de deduções comprimiram a DCL ao longo da maior parte do período.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            f"A composição da dívida permaneceu fortemente concentrada em dívida contratual, que respondeu por {tpl.br_percent(ratio(estado.get('Dívida Contratual')[-1], ms_dc[-1]) or 0, 2)} do total em 2024. "
            f"Dentro desse bloco, a reestruturação da dívida foi a principal rubrica, com participação de {tpl.br_percent(ratio(ms_reestr[-1], ms_dc[-1]) or 0, 2)}, enquanto os empréstimos representaram {tpl.br_percent(ratio(estado.get('Empréstimos')[-1], ms_dc[-1]) or 0, 2)}. "
            f"Os precatórios perderam peso relativo na série, saindo de {tpl.br_percent(ratio(ms_prec[0], ms_dc[0]) or 0, 2)} para {tpl.br_percent(ratio(ms_prec[-1], ms_dc[-1]) or 0, 2)} da dívida consolidada.",
            styles["Body"],
        )
    )
    story.extend(tpl.caption_block("Gráfico", 2, "Composição da dívida consolidada do Estado", styles))
    story.append(tpl.scaled_image(charts["ms_dc_composicao"], tpl.BODY_WIDTH, 205))
    story.append(
        keep_table_block(
            tpl.caption_block("Tabela", 3, "Decomposição da dívida consolidada e dos passivos complementares do Estado", styles),
            variation_table_from_structure(estado, LINHAS_ESTADO),
        )
    )
    story.append(tpl.PageBreak())

    story.append(tpl.make_heading("2.2 - DCL, deduções e capacidade de absorção", styles["H2"], 1, "h2_22"))
    story.append(
        tpl.Paragraph(
            f"A dívida consolidada líquida do Estado caiu de {tpl.br_money_text(ms_dcl[0])} em 2019 para {tpl.br_money_text(ms_dcl[-1])} em 2024, recuo de {tpl.br_percent(pct_change_opt(ms_dcl) or 0, 1)}. "
            f"O menor valor da série foi registrado em {ms_dcl_min_year}, quando a DCL atingiu {tpl.br_money_text(ms_dcl_min_value)}. Em 2024 houve recomposição do indicador, associada à retração das deduções e ao crescimento relativamente mais lento da RCL ajustada frente ao aumento do estoque bruto da dívida.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            f"As deduções somavam {tpl.br_money_text(ms_ded[0])} em 2019 e alcançaram {tpl.br_money_text(ms_ded[-1])} em 2024, alta de {tpl.br_percent(pct_change_opt(ms_ded) or 0, 1)}. "
            f"Esse movimento foi sustentado sobretudo pela disponibilidade de caixa bruta e pelos demais haveres financeiros. No mesmo período, a RCL ajustada avançou {tpl.br_percent(pct_change_opt(ms_rcl) or 0, 1)}, o que reduziu a relação DCL / RCL ajustada de {tpl.br_percent(ms_dcl_pct[0], 2)} para {tpl.br_percent(ms_dcl_pct[-1], 2)}. "
            f"O serviço anual da dívida oscilou de {tpl.br_money_text(ms_servico[0])} para {tpl.br_money_text(ms_servico[-1])}, encerrando 2024 em {tpl.br_percent(servico_divida_ms_pct[-1], 2)} da RCL ajustada.",
            styles["Body"],
        )
    )
    story.extend(tpl.caption_block("Gráfico", 3, "DCL, deduções e RCL ajustada do Estado", styles))
    story.append(tpl.scaled_image(charts["ms_estoque_ajuste"], tpl.BODY_WIDTH, 205))
    story.append(
        keep_table_block(
            caption_block_unit(
                "Tabela",
                4,
                "Síntese dos principais indicadores do Estado",
                "R$ milhões e %",
                "SICONFI. Sistema de Informações Contábeis e Fiscais do Setor Público Brasileiro. Tesouro Nacional.",
                styles,
            ),
            series_table(
                ["Indicador", "2019", "2020", "2021", "2022", "2023", "2024"],
                [
                    ["Dívida consolidada (R$ mi)", *[fmt_million(value, 2) for value in ms_dc]],
                    ["Deduções (R$ mi)", *[fmt_million(value, 2) for value in ms_ded]],
                    ["DCL (R$ mi)", *[fmt_million(value, 2) for value in ms_dcl]],
                    ["RCL ajustada (R$ mi)", *[fmt_million(value, 2) for value in ms_rcl]],
                    ["DCL / RCL ajustada (%)", *[tpl.br_percent(value, 2) for value in ms_dcl_pct]],
                    ["Serviço da dívida (R$ mi)", *[fmt_million(value, 2) for value in ms_servico]],
                    ["Serviço da dívida / RCL ajustada (%)", *[tpl.br_percent(value, 2) for value in servico_divida_ms_pct]],
                    ["Despesa com pessoal / RCL (%)", *[tpl.br_percent(value, 2) for value in ms_pessoal_pct]],
                ],
                [210, 78, 78, 78, 78, 78, 78],
            ),
        )
    )
    story.append(tpl.PageBreak())

    story.append(tpl.make_heading("2.3 - Passivos adicionais e liquidez", styles["H2"], 1, "h2_23"))
    story.append(
        tpl.Paragraph(
            f"O passivo atuarial do Estado permaneceu elevado em toda a série. Após {tpl.br_money_text(ms_passivo[0])} em 2019, o indicador atingiu pico de {tpl.br_money_text(ms_passivo_max_value)} em {ms_passivo_max_year} e recuou para {tpl.br_money_text(ms_passivo[-1])} em 2024. "
            f"Mesmo com a queda recente, o passivo ainda correspondeu a {tpl.br_percent(ms_passivo_pct[-1], 2)} da RCL ajustada, magnitude superior à própria DCL do exercício.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            f"Os restos a pagar não processados representaram {tpl.br_money_text(ms_rpnp[-1])} em 2024, o equivalente a {tpl.br_percent(ratio(ms_rpnp[-1], ms_rcl[-1]) or 0, 2)} da RCL ajustada. "
            f"No campo da liquidez, a disponibilidade de caixa líquida após restos a pagar passou de {tpl.br_percent(ms_liq_apos_pct[0], 2)} da RCL ajustada em 2019 para {tpl.br_percent(ms_liq_apos_pct[-1], 2)} em 2024, mantendo trajetória positiva desde 2020, embora em patamar inferior ao pico observado em 2021.",
            styles["Body"],
        )
    )
    ms_state_grid = tpl.Table(
        [[tpl.scaled_image(charts["ms_passivo"], tpl.BODY_WIDTH / 2 - 8, 185), tpl.scaled_image(charts["ms_liquidez"], tpl.BODY_WIDTH / 2 - 8, 185)]],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    ms_state_grid.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend(
        tpl.caption_block_with_source(
            "Gráfico",
            4,
            "Passivo atuarial e indicadores de liquidez do Estado",
            "SICONFI. Relatório de Gestão Fiscal - Demonstrativos da Dívida Consolidada Líquida e da Disponibilidade de Caixa e dos Restos a Pagar.",
            styles,
        )
    )
    story.append(ms_state_grid)
    story.append(
        tpl.bullet_box(
            "Leitura sintética do Estado",
            [
                "A redução da DCL decorreu mais da combinação entre deduções elevadas e expansão da RCL ajustada do que de uma queda do estoque bruto da dívida.",
                "A dívida estadual permanece concentrada em obrigações contratuais reestruturadas, o que reduz a volatilidade da composição, mas mantém elevado o peso de passivos antigos.",
                "O passivo atuarial segue como principal fragilidade patrimonial fora da dívida consolidada, mesmo com recuo em 2024.",
                "A liquidez do Estado melhorou sensivelmente desde 2019, mas perdeu intensidade após o pico observado em 2021.",
            ],
            styles,
        )
    )

    story.extend(
        [
            tpl.NextPageTemplate("FullPage"),
            tpl.PageBreak(),
            DebtSectionPage("3", "3.0 - Análise das contas da capital - Campo Grande", "Diagnóstico do endividamento da capital, com foco na composição da dívida,\nna DCL e na leitura prudencial do indicador frente à RCL ajustada.", IMG_CG, "sec_cg"),
            tpl.NextPageTemplate("Body"),
            tpl.PageBreak(),
        ]
    )

    story.append(tpl.make_heading("3.1 - Estoque e composição da dívida consolidada", styles["H2"], 1, "h2_31"))
    story.append(
        tpl.Paragraph(
            f"A dívida consolidada de Campo Grande passou de {tpl.br_money_text(cg_dc[0])} em 2019 para {tpl.br_money_text(cg_dc[-1])} em 2024, crescimento nominal de {tpl.br_percent(pct_change_opt(cg_dc) or 0, 1)}. "
            "A trajetória foi ascendente em quase todo o período, com estabilização apenas entre 2022 e 2023 antes da nova alta em 2024.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            f"A composição do endividamento municipal tornou-se ainda mais concentrada em financiamentos, cuja participação subiu de {tpl.br_percent(ratio(cg_fin[0], cg_dc[0]) or 0, 2)} para {tpl.br_percent(ratio(cg_fin[-1], cg_dc[-1]) or 0, 2)} entre 2019 e 2024. "
            f"No sentido oposto, o parcelamento e renegociação perdeu peso relativo, caindo de {tpl.br_percent(ratio(cg_parcel[0], cg_dc[0]) or 0, 2)} para {tpl.br_percent(ratio(cg_parcel[-1], cg_dc[-1]) or 0, 2)} da dívida consolidada.",
            styles["Body"],
        )
    )
    story.extend(tpl.caption_block("Gráfico", 5, "Composição da dívida consolidada de Campo Grande", styles))
    story.append(tpl.scaled_image(charts["cg_dc_composicao"], tpl.BODY_WIDTH, 205))
    story.append(
        keep_table_block(
            tpl.caption_block("Tabela", 5, "Decomposição da dívida consolidada de Campo Grande", styles),
            variation_table_from_structure(capital, LINHAS_CAPITAL),
        )
    )
    story.append(tpl.PageBreak())

    story.append(tpl.make_heading("3.2 - DCL, RCL ajustada e leitura dos limites", styles["H2"], 1, "h2_32"))
    story.append(
        tpl.Paragraph(
            f"A dívida consolidada líquida da capital avançou de {tpl.br_money_text(cg_dcl[0])} em 2019 para {tpl.br_money_text(cg_dcl[-1])} em 2024, aumento de {tpl.br_percent(pct_change_opt(cg_dcl) or 0, 1)}. "
            f"O maior valor da série foi observado em {cg_dcl_max_year}, com {tpl.br_money_text(cg_dcl_max_value)}. Ainda assim, a base de comparação da receita corrente líquida ajustada permaneceu suficientemente ampla para manter o indicador DCL / RCL ajustada em patamar bastante baixo.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            f"A RCL ajustada de Campo Grande cresceu {tpl.br_percent(pct_change_opt(cg_rcl) or 0, 1)} no período, saindo de {tpl.br_money_text(cg_rcl[0])} para {tpl.br_money_text(cg_rcl[-1])}. "
            f"Com isso, a relação DCL / RCL ajustada passou de {tpl.br_percent(cg_dcl_pct[0], 2)} em 2019 para {tpl.br_percent(cg_dcl_pct[-1], 2)} em 2024, nível muito inferior ao limite de referência de 120%. "
            "A leitura prudencial, portanto, é de baixa alavancagem relativa, embora com tendência de elevação da dívida líquida desde 2021.",
            styles["Body"],
        )
    )
    cg_grid = tpl.Table(
        [[tpl.scaled_image(charts["cg_dcl_ratio"], tpl.BODY_WIDTH / 2 - 8, 185), tpl.scaled_image(charts["cg_estoque_ajuste"], tpl.BODY_WIDTH / 2 - 8, 185)]],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    cg_grid.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend(
        tpl.caption_block_with_source(
            "Gráfico",
            6,
            "Indicadores de endividamento e base de receita de Campo Grande",
            "SICONFI. Relatório de Gestão Fiscal - Demonstrativo da Dívida Consolidada Líquida.",
            styles,
        )
    )
    story.append(cg_grid)
    story.append(
        keep_table_block(
            caption_block_unit(
                "Tabela",
                6,
                "Síntese dos principais indicadores de Campo Grande",
                "R$ milhões e %",
                "SICONFI. Sistema de Informações Contábeis e Fiscais do Setor Público Brasileiro. Tesouro Nacional.",
                styles,
            ),
            series_table(
                ["Indicador", "2019", "2020", "2021", "2022", "2023", "2024"],
                [
                    ["Dívida consolidada (R$ mi)", *[fmt_million(value, 2) for value in cg_dc]],
                    ["Dívida consolidada líquida (R$ mi)", *[fmt_million(value, 2) for value in cg_dcl]],
                    ["RCL ajustada (R$ mi)", *[fmt_million(value, 2) for value in cg_rcl]],
                    ["DCL / RCL ajustada (%)", *[tpl.br_percent(value, 2) for value in cg_dcl_pct]],
                    ["Financiamentos (R$ mi)", *[fmt_million(value, 2) for value in cg_fin]],
                    ["Parcelamento e renegociação (R$ mi)", *[fmt_million(value, 2) for value in cg_parcel]],
                    ["Serviço da dívida (R$ mi)", *[fmt_million(value, 2) for value in cg_servico]],
                    ["Serviço da dívida / RCL ajustada (%)", *[fmt_percent(value, 2) for value in servico_divida_cg_pct]],
                    ["Despesa com pessoal / RCL (%)", *[fmt_percent(value, 2) for value in cg_pessoal_pct]],
                    ["Passivo atuarial / RCL ajustada (%)", *[fmt_percent(value, 2) for value in cg_passivo_pct]],
                    ["RP não processados / RCL ajustada (%)", *[fmt_percent(value, 2) for value in cg_rpnp_pct]],
                    ["Depósitos judiciais / RCL ajustada (%)", *[fmt_percent(value, 2) for value in cg_depositos_judiciais_pct]],
                ],
                [210, 78, 78, 78, 78, 78, 78],
            ),
        )
    )
    story.append(
        tpl.Paragraph(
            f"O serviço da dívida de Campo Grande passou de {tpl.br_money_text(cg_servico[0])} em 2019 para {tpl.br_money_text(cg_servico[-1])} em 2024, elevando a razão Serviço da Dívida / RCL ajustada de {tpl.br_percent(servico_divida_cg_pct[0], 2)} para {tpl.br_percent(servico_divida_cg_pct[-1], 2)}. "
            "A ausência de valores para passivos complementares da capital nos demonstrativos consolidados não deve ser lida como inexistência econômica desses riscos. A leitura adequada é que, na base SICONFI utilizada, a materialidade analítica da capital ficou concentrada na dívida consolidada, nas deduções, na DCL, no serviço anual da dívida e na composição contratual do passivo.",
            styles["Body"],
        )
    )
    story.extend(
        [
            tpl.Spacer(1, 8),
            tpl.bullet_box(
                "Leitura sintética de Campo Grande",
                [
                    "A capital mantém baixo endividamento relativo, mesmo com crescimento do estoque líquido desde 2021.",
                    "O perfil da dívida tornou-se mais concentrado em financiamentos, o que sugere menor peso de passivos renegociados antigos e maior relevância de contratos de crédito recentes.",
                    "A distância em relação ao limite de 120% da RCL ajustada permaneceu ampla em toda a série, preservando folga prudencial do ponto de vista estritamente quantitativo.",
                    "Os demonstrativos consolidados não evidenciam valores para passivo atuarial, restos a pagar não processados ou depósitos judiciais da capital; por isso, o diagnóstico registra a lacuna. e recomenda monitoramento sem substituir dado ausente por estimativa.",
                ],
                styles,
            )
        ]
    )

    story.extend(
        [
            tpl.NextPageTemplate("FullPage"),
            tpl.PageBreak(),
            DebtSectionPage("4", "4.0 - Conclusão", "Síntese interpretativa da trajetória do endividamento, das deduções,\ndos passivos adicionais e dos principais pontos de atenção para o estado e a capital.", IMG_MS, "sec_conclusao"),
            tpl.NextPageTemplate("Body"),
            tpl.PageBreak(),
        ]
    )
    story.append(
        tpl.Paragraph(
            f"O diagnóstico consolidado mostra que o Mato Grosso do Sul encerrou 2024 com dívida consolidada de {tpl.br_money_text(ms_dc[-1])}, mas com dívida consolidada líquida de {tpl.br_money_text(ms_dcl[-1])} e relação DCL / RCL ajustada de {tpl.br_percent(ms_dcl_pct[-1], 2)}. "
            "O Estado combina, portanto, endividamento bruto relevante com folga confortável frente ao limite prudencial, sustentada por deduções elevadas e pela expansão da base corrente de receitas. "
            "Ainda assim, a alta da DCL em 2024 e o peso persistente do passivo atuarial mostram que a posição fiscal não pode ser avaliada apenas pelo indicador legal de endividamento.",
            styles["Lead"],
        )
    )
    story.append(
        tpl.Paragraph(
            f"Em Campo Grande, a situação é distinta: a dívida consolidada e a DCL cresceram no período, alcançando {tpl.br_money_text(cg_dc[-1])} e {tpl.br_money_text(cg_dcl[-1])} em 2024, respectivamente. "
            f"Mesmo assim, a relação DCL / RCL ajustada ficou em apenas {tpl.br_percent(cg_dcl_pct[-1], 2)}, o que indica baixa alavancagem relativa. O principal traço estrutural da capital é a concentração crescente do endividamento em financiamentos, mais do que uma pressão quantitativa imediata sobre os limites de referência. "
            f"O serviço da dívida representou {tpl.br_percent(servico_divida_cg_pct[-1], 2)} da RCL ajustada em 2024; já os passivos complementares da capital permanecem como itens de monitoramento por ausência de valores destacados na base consolidada.",
            styles["Lead"],
        )
    )
    story.append(tpl.make_heading("4.1 - Desafios e oportunidades para a política fiscal e financeira subnacional", styles["H2"], 1, "h2_41"))
    story.append(
        tpl.Paragraph(
            "A leitura consolidada do endividamento mostra que o cumprimento dos limites legais não esgota a análise de sustentabilidade fiscal. Em ambos os entes, a principal contribuição do diagnóstico é revelar que a folga prudencial convive com riscos estruturais distintos: no Estado, a combinação entre passivo atuarial elevado, perda recente de liquidez relativa e recomposição da DCL; na capital, a elevação da DCL a partir de uma base baixa e a concentração crescente do passivo em financiamentos.",
            styles["Body"],
        )
    )
    diagnostic_panels = tpl.Table(
        [
            [
                tpl.info_panel(
                    "Desafio central",
                    "Transformar a folga diante do limite de endividamento em sustentabilidade efetiva. Isso exige olhar para a qualidade das deduções, para a persistência do passivo atuarial, para a trajetória do caixa e para a composição do passivo, e não apenas para o indicador legal observado no encerramento do exercício.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
                tpl.info_panel(
                    "Oportunidade estratégica",
                    "Usar o espaço prudencial disponível para qualificar a gestão financeira e a seleção de investimentos, evitando ampliar despesas permanentes sem lastro. O desafio não é simplesmente contrair menos dívida, mas utilizar melhor a capacidade fiscal e financeira para apoiar desenvolvimento com menor vulnerabilidade futura.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    diagnostic_panels.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([diagnostic_panels, tpl.Spacer(1, 10)])
    story.append(
        tpl.bullet_box(
            "Síntese crítica dos principais desafios e oportunidades",
            [
                f"No Estado, a relação DCL / RCL ajustada de {tpl.br_percent(ms_dcl_pct[-1], 2)} indica folga legal, mas a presença de passivo atuarial equivalente a {tpl.br_percent(ms_passivo_pct[-1], 2)} da RCL ajustada reduz o conforto da leitura estritamente normativa.",
                f"A composição da dívida estadual permanece concentrada em reestruturação, responsável por {tpl.br_percent(ratio(ms_reestr[-1], ms_dc[-1]) or 0, 2)} do estoque de 2024, o que sinaliza baixa volatilidade da composição, mas também persistência de passivos de longo curso.",
                f"Em Campo Grande, a relação DCL / RCL ajustada de {tpl.br_percent(cg_dcl_pct[-1], 2)} preserva ampla margem prudencial, porém a concentração de {tpl.br_percent(ratio(cg_fin[-1], cg_dc[-1]) or 0, 2)} da dívida em financiamentos pede governança mais rigorosa sobre novas operações e contrapartidas em investimento.",
                "A principal oportunidade comum aos dois entes é converter a informação fiscal em planejamento: monitorar a dívida como instrumento de política pública, e não apenas como restrição legal a ser cumprida ao final do exercício.",
            ],
            styles,
        )
    )
    story.append(tpl.make_heading("4.2 - Cenários prospectivos e projeção analítica", styles["H2"], 1, "h2_42"))
    story.append(
        tpl.Paragraph(
            "Para elevar o valor analítico do relatório, esta subseção apresenta cenários qualitativos para o triênio 2025-2027. Não se trata de projeção econométrica, mas de uma leitura prospectiva baseada na trajetória recente da DCL, da RCL ajustada, das deduções, do passivo atuarial e da liquidez, combinada com diferentes hipóteses de gestão e ambiente macroeconômico.",
            styles["Body"],
        )
    )
    scenario_panels = tpl.Table(
        [
            [
                tpl.info_panel(
                    "Como ler os cenários",
                    "Os cenários combinam hipóteses sobre crescimento da receita, comportamento do caixa, ritmo de novos financiamentos, evolução dos passivos e qualidade da gestão. O objetivo é antecipar riscos e orientar escolhas de política antes que o problema apareça apenas no fechamento contábil.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
                tpl.info_panel(
                    "Utilidade prática",
                    "A leitura prospectiva ajuda a diferenciar o que é folga estrutural do que é folga temporária. Ela também permite escolher entre respostas mais defensivas ou mais ativas, dependendo da intensidade do risco e da capacidade de coordenação institucional dos entes.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    scenario_panels.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([scenario_panels, tpl.Spacer(1, 10)])
    story.append(
        tpl.narrative_table(
            ["Cenário", "Hipóteses para Mato Grosso do Sul", "Hipóteses para Campo Grande", "Implicações fiscais esperadas"],
            [
                [
                    "Reequilíbrio gradual",
                    "Estabilização da DCL em patamar próximo ao de 2024, preservação das deduções, recomposição moderada da liquidez e gestão mais ativa do passivo atuarial.",
                    "Crescimento mais lento da DCL, manutenção da RCL ajustada em expansão e freio na abertura de novas pressões permanentes sobre custeio e serviço da dívida.",
                    "O indicador legal seguiria confortável nos dois entes, com melhora gradual da qualidade da posição fiscal e menor risco de deterioração súbita entre 2025 e 2027.",
                ],
                [
                    "Gestão ativa com investimento seletivo",
                    "Uso da folga prudencial para priorizar projetos de maior retorno social e econômico, combinado com governança de caixa, seleção rigorosa de operações e disciplina sobre passivos complementares.",
                    "Financiamentos concentrados em carteira restrita de investimentos urbanos prioritários, com exigência de contrapartida financeira e monitoramento de impacto sobre a RCL ajustada.",
                    "É o cenário mais equilibrado: preserva capacidade de investimento, mas exige coordenação forte para evitar que a expansão do passivo reduza a sustentabilidade futura.",
                ],
                [
                    "Inércia ou estresse adverso",
                    "Queda da liquidez, nova alta da DCL, manutenção de passivos elevados e menor capacidade de absorver choques de receita ou mudanças no custo do financiamento.",
                    "Continuidade do crescimento da DCL e dos financiamentos sem filtro suficiente de prioridade, com maior exposição a restrições orçamentárias e postergação de investimentos essenciais.",
                    "A folga legal permaneceria relevante no curto prazo, mas o risco de fragilização estrutural aumentaria, reduzindo a qualidade do ajuste e o espaço para políticas públicas nos exercícios seguintes.",
                ],
            ],
            [112, 184, 184, 250],
        )
    )
    story.extend(
        [
            tpl.Spacer(1, 8),
            tpl.bullet_box(
                "Projeção analítica 2025-2027",
                [
                    "Sem intervenção gerencial adicional, a tendência mais provável é de manutenção da folga legal com piora relativa da qualidade do endividamento, sobretudo se a DCL continuar subindo mais rápido que a capacidade de geração de caixa.",
                    "No Estado, o principal ponto de atenção é a diferença entre folga formal e vulnerabilidade patrimonial, dada a persistência do passivo atuarial e a perda de intensidade da liquidez após 2021.",
                    "Em Campo Grande, a pressão mais relevante não está no limite legal atual, mas na trajetória de crescimento do endividamento líquido e na necessidade de vincular novos financiamentos a investimentos claramente prioritários.",
                ],
                styles,
            )
        ]
    )
    story.append(tpl.make_heading("4.3 - Estratégias, propostas alternativas e plano de ação", styles["H2"], 1, "h2_43"))
    story.append(
        tpl.Paragraph(
            "À luz dos cenários projetados, o relatório recomenda tratar o endividamento como variável estratégica de gestão e não apenas como obrigação normativa. Para isso, foram organizadas duas rotas de política fiscal e financeira, além de uma rota inercial não recomendada, a fim de explicitar alternativas de ação para diferentes stakeholders.",
            styles["Body"],
        )
    )
    story.append(
        tpl.bullet_box(
            "Alternativas de política fiscal e financeira",
            [
                "Rota A - Consolidação prudencial acelerada: prioriza recomposição de caixa, contenção mais forte de compromissos e desaceleração do uso de crédito. É mais defensiva e adequada a contexto de estresse ou frustração relevante de receita.",
                "Rota B - Reequilíbrio com investimento seletivo: combina disciplina fiscal, governança da dívida e preservação de investimentos de maior retorno social e econômico. É a alternativa mais equilibrada para os dois entes e a mais aderente ao diagnóstico atual.",
                "Rota C - Inércia administrativa: mantém decisões fragmentadas e baixa integração entre endividamento, caixa e investimento. Não é recomendada porque preserva folga legal no curto prazo, mas amplia vulnerabilidades de médio prazo.",
            ],
            styles,
        )
    )
    story.append(
        tpl.narrative_table(
            ["Frente estratégica", "Mato Grosso do Sul", "Campo Grande", "Stakeholders, horizonte e evidência de sucesso"],
            [
                [
                    "1. Gestão da dívida e do caixa",
                    "Aprimorar o acompanhamento da DCL, das deduções e da liquidez como painel único de decisão, com gatilhos de contingência quando houver deterioração do caixa ou recomposição excessiva do endividamento líquido.",
                    "Vincular novos financiamentos a cronograma físico-financeiro crível, fonte de pagamento e filtro de prioridade, evitando que a baixa DCL / RCL ajustada seja interpretada como licença para expansão indiscriminada do crédito.",
                    "Executivo e tesouraria. Horizonte imediato. Sucesso medido por maior previsibilidade do caixa e menor descolamento entre dívida contratada, execução e capacidade de pagamento.",
                ],
                [
                    "2. Passivos estruturais",
                    "Tratar passivo atuarial e restos a pagar como núcleo da gestão patrimonial, com monitoramento periódico e integração ao planejamento fiscal de médio prazo.",
                    "Monitorar passivos financeiros e contratuais associados aos financiamentos, com leitura de custo futuro de manutenção, operação e contrapartidas urbanas.",
                    "Executivo, controle e áreas setoriais. Horizonte de 6 a 18 meses. Sucesso medido por redução de riscos ocultos e maior transparência patrimonial.",
                ],
                [
                    "3. Investimento e endividamento",
                    "Priorizar logística, infraestrutura produtiva e projetos com alto retorno econômico e social, preservando apenas operações coerentes com sustentabilidade financeira e impacto público demonstrável.",
                    "Hierarquizar drenagem, mobilidade, escolas, saúde, manutenção urbana e infraestrutura básica segundo criticidade social, custo de ciclo de vida e capacidade real de execução.",
                    "Executivo, Legislativo e sociedade. Horizonte de 12 a 24 meses. Sucesso medido por carteira mais concentrada em prioridades e menor dispersão dos recursos financiados.",
                ],
                [
                    "4. Transparência e monitoramento",
                    "Instituir painel trimestral com DCL, deduções, passivo atuarial, liquidez e operações de crédito, articulado ao planejamento orçamentário e à comunicação pública.",
                    "Criar rotina de monitoramento mensal da trajetória da DCL, do serviço da dívida e da execução dos projetos financiados, com linguagem acessível para Câmara, controle e sociedade.",
                    "Executivo, Legislativo e órgãos de controle. Horizonte imediato e contínuo. Sucesso medido por decisão mais tempestiva e menor assimetria de informação entre stakeholders.",
                ],
            ],
            [112, 184, 184, 250],
        )
    )
    story.extend(
        [
            tpl.Spacer(1, 8),
            tpl.bullet_box(
                "Responsabilidades dos principais stakeholders",
                [
                    "Poder Executivo estadual e municipal: liderar a estratégia escolhida, revisar prioridades, controlar o ritmo de expansão do passivo e integrar dívida, caixa e investimento no mesmo ciclo decisório.",
                    "Poder Legislativo: avaliar os efeitos intertemporais de novas despesas permanentes e de novas operações de crédito, reforçando a coerência entre autorização política e sustentabilidade futura.",
                    "Órgãos de controle: combinar fiscalização de legalidade com verificação de qualidade do gasto financiado, transparência patrimonial e consistência das justificativas para expansão do passivo.",
                    "Sociedade e usuários dos serviços públicos: acompanhar metas, cobrar clareza distributiva das escolhas e participar do debate sobre quais investimentos e serviços devem ser protegidos em cenários mais restritivos.",
                ],
                styles,
            )
        ]
    )
    story.append(tpl.make_heading("4.4 - Trade-offs, implicações sociais e critérios éticos", styles["H2"], 1, "h2_44"))
    story.append(
        tpl.Paragraph(
            "A gestão do endividamento envolve dilemas que não podem ser resolvidos apenas com base em limites formais. Em ambos os entes, as escolhas fiscais afetam capacidade de investimento, qualidade dos serviços públicos, distribuição territorial dos recursos e proteção de grupos mais vulneráveis. Por isso, o ajuste recomendado deve ser seletivo, transparente e eticamente justificável.",
            styles["Body"],
        )
    )
    tradeoff_panels = tpl.Table(
        [
            [
                tpl.info_panel(
                    "Folga legal versus risco estrutural",
                    "Cumprir o limite de DCL / RCL ajustada é condição necessária, mas insuficiente. A leitura responsável precisa considerar passivos patrimoniais, liquidez, qualidade das deduções e custo futuro das decisões presentes, evitando a falsa sensação de conforto fiscal.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
                tpl.info_panel(
                    "Investimento financiado versus prudência intertemporal",
                    "Reduzir investimento preserva caixa no curtíssimo prazo, mas pode aumentar custos sociais e econômicos futuros. Por outro lado, financiar sem priorização eleva o passivo e transfere riscos para exercícios posteriores. O equilíbrio está em selecionar melhor, e não apenas cortar ou expandir.",
                    styles,
                    tpl.BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[tpl.BODY_WIDTH / 2 - 6, tpl.BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    tradeoff_panels.setStyle(tpl.TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([tradeoff_panels, tpl.Spacer(1, 10)])
    story.append(
        tpl.Paragraph(
            "Há ainda uma dimensão distributiva importante. Se o ajuste recair de forma linear sobre manutenção urbana, saúde, educação ou políticas de maior capilaridade social, o custo do reequilíbrio tende a ser regressivo. Em sentido oposto, se a expansão do endividamento não estiver vinculada a projetos com retorno coletivo claro, o ganho presente também se torna socialmente questionável. O critério ético central, portanto, é explicitar quem suporta os custos do ajuste e quais benefícios públicos se pretende proteger.",
            styles["Body"],
        )
    )
    story.append(
        tpl.bullet_box(
            "Princípios orientadores para uma resposta fiscal equilibrada",
            [
                "Substituir cortes lineares por revisão de gasto baseada em evidências, produtividade, criticidade social e custo de longo prazo.",
                "Preservar serviços e investimentos com maior retorno social, econômico e territorial, especialmente aqueles que evitam deterioração futura mais cara.",
                "Dar transparência aos critérios distributivos do ajuste e à seleção dos investimentos financiados, para que a sociedade compreenda o custo e o benefício intertemporal das escolhas.",
                "Evitar que a folga legal diante dos limites seja usada para expandir despesas permanentes sem reforço equivalente da capacidade de financiamento e de gestão.",
            ],
            styles,
        )
    )
    story.append(tpl.make_heading("4.5 - Endividamento, macroeconomia e desenvolvimento econômico e social", styles["H2"], 1, "h2_45"))
    story.append(
        tpl.Paragraph(
            "O diagnóstico do endividamento precisa ser lido em interação com as condições macroeconômicas e com os objetivos de desenvolvimento. No Mato Grosso do Sul, a capacidade de financiamento do setor público é influenciada por uma economia sensível ao ciclo do agronegócio, da logística e das cadeias exportadoras. Em Campo Grande, a sustentabilidade do passivo se conecta à dinâmica urbana, à capacidade de ampliar a base tributária e à qualidade dos investimentos em infraestrutura, mobilidade e serviços públicos.",
            styles["Body"],
        )
    )
    story.append(
        tpl.Paragraph(
            "Essa interação mostra que política fiscal e política de desenvolvimento não devem ser tratadas como agendas opostas. Uma estratégia equilibrada busca reduzir vulnerabilidades financeiras sem sacrificar investimentos estruturantes, enquanto uma estratégia inovadora usa informação fiscal para escolher melhor onde investir, em que ritmo financiar e como proteger a solvência futura. Em outras palavras, o melhor uso do endividamento é aquele que amplia capacidade produtiva, qualidade urbana e retorno social sem comprometer a governança intertemporal das contas públicas.",
            styles["Body"],
        )
    )
    story.append(
        tpl.bullet_box(
            "Soluções inovadoras e equilibradas sugeridas pelo relatório",
            [
                "Painel fiscal e patrimonial com atualização periódica de DCL, deduções, liquidez, passivos complementares e execução dos projetos financiados.",
                "Testes de estresse sobre dívida e caixa para simular efeitos de desaceleração da receita, elevação de custos financeiros e choques setoriais relevantes.",
                "Carteira priorizada de investimentos, com ranking por retorno social, econômico, territorial e fiscal, para orientar o uso prudente do crédito.",
                "Integração entre planejamento, orçamento, tesouraria e avaliação de políticas públicas, reduzindo a distância entre autorização de crédito, execução financeira e entrega efetiva à população.",
            ],
            styles,
        )
    )
    story.append(
        tpl.bullet_box(
            "Síntese final",
            [
                "O relatório mostra que a situação de endividamento dos dois entes é mais favorável quando lida pelo limite formal do que quando observada pela qualidade estrutural da posição fiscal.",
                "A superação desse desafio depende de projeção de cenários, seleção entre alternativas de política e explicitação dos trade-offs envolvidos na gestão do passivo.",
                "A alternativa mais consistente para Mato Grosso do Sul e Campo Grande é o reequilíbrio com investimento seletivo, combinando prudência financeira, transparência e foco em projetos de maior retorno público.",
                "Com isso, o endividamento deixa de ser apenas um problema contábil e passa a ser tratado como instrumento de política fiscal e financeira articulado ao desenvolvimento econômico e social.",
            ],
            styles,
        )
    )

    story.extend(
        [
            tpl.NextPageTemplate("FullPage"),
            tpl.PageBreak(),
            DebtSectionPage("5", "5.0 - Referências", "Fontes utilizadas na elaboração do relatório e nas análises apresentadas\nsobre o endividamento do estado e da capital.", IMG_CG, "sec_refs"),
            tpl.NextPageTemplate("Body"),
            tpl.PageBreak(),
        ]
    )
    story.append(
        tpl.Paragraph(
            "SICONFI. Sistema de Informações Contábeis e Fiscais do Setor Público Brasileiro. Brasília: Tesouro Nacional, [s.d.]. Disponível em: https://siconfi.tesouro.gov.br/siconfi/index.jsf. Acesso em: 23 abr. 2026.",
            styles["Ref"],
        )
    )
    story.append(
        tpl.Paragraph(
            "SENADO FEDERAL. Resolução do Senado Federal nº 40, de 20 de dezembro de 2001. Dispõe sobre os limites globais para o montante da dívida pública consolidada e da dívida pública mobiliária dos Estados, do Distrito Federal e dos Municípios. Disponível em: https://legis.senado.leg.br/norma/562458. Acesso em: 23 abr. 2026.",
            styles["Ref"],
        )
    )
    story.append(
        tpl.Paragraph(
            "SECRETARIA DO TESOURO NACIONAL. Boletim de Finanças dos Entes Subnacionais 2025. Brasília: Ministério da Fazenda, 2026.",
            styles["Ref"],
        )
    )
    return story


def main() -> None:
    tpl.register_fonts()
    ensure_dirs()
    workbook = load_workbook(INPUT_XLSX, data_only=True)
    estado = MatrixSheet(workbook, workbook.sheetnames[0])
    capital = MatrixSheet(workbook, workbook.sheetnames[1])
    indicadores = IndicatorSheet(workbook, workbook.sheetnames[2])
    charts = generate_charts(estado, capital, indicadores)
    styles = tpl.build_styles()
    story = build_story(estado, capital, indicadores, charts, styles)
    doc = DebtDocTemplate(str(OUTPUT_PDF))
    doc.multiBuild(story)
    print(OUTPUT_PDF.name)


if __name__ == "__main__":
    main()
