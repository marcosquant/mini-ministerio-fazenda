from __future__ import annotations

import hashlib
import unicodedata
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Flowable,
    Frame,
    Image,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus.doctemplate import NextPageTemplate
from reportlab.platypus.tableofcontents import TableOfContents

try:
    hashlib.md5(b"", usedforsecurity=False)
except TypeError:
    import reportlab.pdfbase.pdfdoc as pdfdoc
    import reportlab.lib.utils as rl_utils

    def _safe_md5(*args, **kwargs):
        kwargs.pop("usedforsecurity", None)
        return hashlib.md5(*args, **kwargs)

    pdfdoc.md5 = _safe_md5
    rl_utils.md5 = _safe_md5


ROOT = Path(__file__).resolve().parent
INPUT_XLSX = ROOT / "demonstrativos_fiscais_MS_2019_2024.xlsx"
OUTPUT_DIR = ROOT / "build_relatorio"
CHART_DIR = OUTPUT_DIR / "charts"
OUTPUT_PDF = ROOT / "Relatorio_Diagnostico_Fiscal_MS_Campo_Grande_2026.pdf"

IMG_MS = ROOT / "img_ms_natureza.jpg"
IMG_CG = ROOT / "img_campo_grande.jpg"

PIB_MS_YEARS = [2019, 2020, 2021, 2022, 2023]
PIB_BRASIL = [7389131.00, 7609597.73, 9012142.00, 10079676.38, 10943345.44]
PIB_CENTRO_OESTE = [731351.48, 791250.73, 932165.62, 1069310.34, 1159827.00]
PIB_MS = [106943.25, 122627.72, 142203.77, 166407.33, 184402.00]
PIB_GROWTH_BRASIL = [1.22, -3.28, 4.76, 3.02, 3.24]
PIB_GROWTH_MS = [-0.53, 0.25, 0.85, 4.76, 13.44]
PIB_MS_POP_2023 = 2757013
PIB_MS_PER_CAPITA_2023 = 66884.75
PIB_MS_PART_CENTRO_OESTE_2023 = 15.90
PIB_MS_PART_BRASIL_2023 = 1.69
PIB_VAB_2023 = {
    "Agropecuária": 41832.0,
    "Indústria": 36059.0,
    "Serviços": 83478.0,
}
PIB_VAB_SHARE_2023 = {
    "Agropecuária": 25.92,
    "Indústria": 22.35,
    "Serviços": 51.73,
}

PAGE_WIDTH = 836.22
PAGE_HEIGHT = 595.276
LEFT_MARGIN = 42
RIGHT_MARGIN = 40
TOP_MARGIN = 56
BOTTOM_MARGIN = 36
BODY_WIDTH = PAGE_WIDTH - LEFT_MARGIN - RIGHT_MARGIN
BODY_HEIGHT = PAGE_HEIGHT - TOP_MARGIN - BOTTOM_MARGIN

NAVY = colors.HexColor("#0C2742")
NAVY_LIGHT = colors.HexColor("#1E4C73")
TEAL = colors.HexColor("#2E7B83")
LIME = colors.HexColor("#6EB63D")
GOLD = colors.HexColor("#C49A45")
RED = colors.HexColor("#B85042")
GREEN = colors.HexColor("#31784B")
SLATE = colors.HexColor("#516171")
TEXT = colors.HexColor("#243240")
LIGHT_BG = colors.HexColor("#F4F7FA")
LIGHT_LINE = colors.HexColor("#D6DEE6")
CARD_BG = colors.HexColor("#EEF3F7")
WHITE = colors.white

FONT_REGULAR = "Helvetica"
FONT_BOLD = "Helvetica-Bold"
FONT_ITALIC = "Helvetica-Oblique"

TEAM_MEMBERS = [
    "Daniel Cassimiro Oliveira dos Santos",
    "David Germano dos Santos",
    "Eduardo de Freitas Nunes Maio",
    "Lucas Soares da Cruz",
    "Marcos Roberto Souza",
    "Philipe Gomes dos Santos",
]


def register_fonts() -> None:
    global FONT_REGULAR, FONT_BOLD, FONT_ITALIC
    font_dir = Path(r"C:\Windows\Fonts")
    regular = font_dir / "segoeui.ttf"
    bold = font_dir / "segoeuib.ttf"
    italic = font_dir / "segoeuii.ttf"
    if regular.exists() and bold.exists() and italic.exists():
        pdfmetrics.registerFont(TTFont("SegoeUI", str(regular)))
        pdfmetrics.registerFont(TTFont("SegoeUI-Bold", str(bold)))
        pdfmetrics.registerFont(TTFont("SegoeUI-Italic", str(italic)))
        FONT_REGULAR = "SegoeUI"
        FONT_BOLD = "SegoeUI-Bold"
        FONT_ITALIC = "SegoeUI-Italic"


def normalize(text: str) -> str:
    cleaned = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    return " ".join(cleaned.upper().split())


def br_number(value: float, decimals: int = 2) -> str:
    pattern = f"{{:,.{decimals}f}}"
    return pattern.format(value).replace(",", "X").replace(".", ",").replace("X", ".")


def br_percent(value: float, decimals: int = 1, signed: bool = False) -> str:
    sign = "+" if signed and value > 0 else ""
    return f"{sign}{br_number(value, decimals)}%"


def br_money_millions(value: float, decimals: int = 2) -> str:
    return br_number(value, decimals)


def br_money_text(value_millions: float, decimals: int = 1) -> str:
    prefix = "-R$ " if value_millions < 0 else "R$ "
    absolute = abs(value_millions)
    if absolute >= 1000:
        return f"{prefix}{br_number(absolute / 1000, decimals)} bilhões"
    return f"{prefix}{br_number(absolute, decimals)} milhões"


def pct_change(series: list[float]) -> float:
    if not series[0]:
        return 0.0
    return (series[-1] - series[0]) / series[0] * 100


def nominal_change(series: list[float]) -> float:
    return series[-1] - series[0]


def share(part: float, total: float) -> float:
    if not total:
        return 0.0
    return part / total * 100


def extrema_years(years: list[int], values: list[float]) -> tuple[int, float, int, float]:
    max_value = max(values)
    min_value = min(values)
    return years[values.index(max_value)], max_value, years[values.index(min_value)], min_value


def ensure_dirs() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    CHART_DIR.mkdir(exist_ok=True)


class SheetData:
    def __init__(self, sheet_name: str) -> None:
        wb = load_workbook(INPUT_XLSX, data_only=True)
        ws = wb[sheet_name]
        self.sheet_name = sheet_name
        self.years = [int(cell.value) for cell in ws[4][1:]]
        self.rows: dict[str, list[float]] = {}
        self.labels: dict[str, str] = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            label = row[0]
            if not label:
                continue
            key = normalize(label)
            self.rows[key] = [float(value or 0) for value in row[1:]]
            self.labels[key] = str(label)

    def get(self, label: str) -> list[float]:
        wanted = normalize(label)
        if wanted in self.rows:
            return self.rows[wanted]
        for key, values in self.rows.items():
            if wanted in key or key in wanted:
                return values
        raise KeyError(f"Linha não encontrada em {self.sheet_name}: {label}")

    def label(self, label: str) -> str:
        wanted = normalize(label)
        if wanted in self.labels:
            return self.labels[wanted]
        for key, value in self.labels.items():
            if wanted in key or key in wanted:
                return value
        raise KeyError(f"Rótulo não encontrado em {self.sheet_name}: {label}")


def axis_formatter(divisor: float) -> FuncFormatter:
    def _formatter(value: float, _: float) -> str:
        scaled = value / divisor
        if abs(scaled) >= 10:
            return br_number(scaled, 0)
        return br_number(scaled, 1)

    return FuncFormatter(_formatter)


def setup_plot() -> None:
    plt.rcParams.update(
        {
            "font.family": "Segoe UI",
            "axes.titlesize": 12,
            "axes.titleweight": "bold",
            "axes.edgecolor": "#C6D0DA",
            "axes.labelcolor": "#374957",
            "xtick.color": "#4C5B68",
            "ytick.color": "#4C5B68",
            "figure.facecolor": "white",
            "axes.facecolor": "#F7F9FC",
            "grid.color": "#D8E2EB",
            "grid.linestyle": "-",
            "grid.linewidth": 0.8,
        }
    )


def line_chart(
    path: Path,
    years: list[int],
    series_items: list[tuple[str, list[float], str]],
    title: str,
    ylabel: str,
    divisor: float = 1.0,
) -> None:
    fig, ax = plt.subplots(figsize=(9.2, 3.2), dpi=180)
    for label, values, color in series_items:
        ax.plot(years, values, marker="o", markersize=5.5, linewidth=2.5, color=color, label=label)
    ax.set_title(title, loc="left", color="#0C2742", pad=10)
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(axis_formatter(divisor))
    ax.set_ylabel(ylabel)
    ax.grid(axis="y")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(frameon=False, loc="upper left", ncol=len(series_items))
    plt.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def stacked_bar_chart(
    path: Path,
    years: list[int],
    stacks: list[tuple[str, list[float], str]],
    title: str,
    ylabel: str,
    divisor: float = 1.0,
    line_item: tuple[str, list[float], str] | None = None,
) -> None:
    fig, ax = plt.subplots(figsize=(9.2, 3.2), dpi=180)
    bottom = [0.0] * len(years)
    for label, values, color in stacks:
        ax.bar(years, values, bottom=bottom, width=0.6, color=color, label=label)
        bottom = [bottom[idx] + values[idx] for idx in range(len(values))]
    if line_item is not None:
        label, values, color = line_item
        ax.plot(years, values, color=color, linewidth=2.4, marker="o", label=label)
    ax.set_title(title, loc="left", color="#0C2742", pad=10)
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(axis_formatter(divisor))
    ax.set_ylabel(ylabel)
    ax.grid(axis="y")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(frameon=False, loc="upper left", ncol=min(3, len(stacks) + (1 if line_item else 0)))
    plt.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def result_bar_chart(
    path: Path,
    years: list[int],
    values: list[float],
    title: str,
    ylabel: str,
    divisor: float = 1.0,
) -> None:
    fig, ax = plt.subplots(figsize=(9.2, 3.0), dpi=180)
    colors_list = ["#31784B" if value >= 0 else "#B85042" for value in values]
    ax.bar(years, values, width=0.6, color=colors_list)
    ax.axhline(0, color="#7E8A97", linewidth=1)
    ax.set_title(title, loc="left", color="#0C2742", pad=10)
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(axis_formatter(divisor))
    ax.set_ylabel(ylabel)
    ax.grid(axis="y")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def combo_bar_line_chart(
    path: Path,
    years: list[int],
    bars: list[float],
    line: list[float],
    title: str,
    bar_label: str,
    line_label: str,
    bar_divisor: float = 1.0,
) -> None:
    fig, ax1 = plt.subplots(figsize=(9.2, 3.2), dpi=180)
    ax1.bar(years, bars, width=0.62, color="#2E7B83", label=bar_label)
    ax1.set_title(title, loc="left", color="#0C2742", pad=10)
    ax1.set_xticks(years)
    ax1.set_ylabel(bar_label)
    ax1.yaxis.set_major_formatter(axis_formatter(bar_divisor))
    ax1.grid(axis="y")
    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)

    ax2 = ax1.twinx()
    ax2.plot(years, line, color="#C49A45", marker="o", linewidth=2.4, label=line_label)
    ax2.set_ylabel(line_label)
    ax2.set_ylim(0, max(line) * 1.35)
    ax2.spines["top"].set_visible(False)

    lines_1, labels_1 = ax1.get_legend_handles_labels()
    lines_2, labels_2 = ax2.get_legend_handles_labels()
    ax1.legend(lines_1 + lines_2, labels_1 + labels_2, frameon=False, loc="upper left", ncol=2)
    plt.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def horizontal_bar_chart(
    path: Path,
    labels: list[str],
    values: list[float],
    title: str,
    xlabel: str,
) -> None:
    fig, ax = plt.subplots(figsize=(9.2, 3.0), dpi=180)
    color_map = ["#6EB63D", "#0C2742", "#2E7B83", "#C49A45", "#B85042"]
    colors_list = color_map[: len(labels)]
    ax.barh(labels, values, color=colors_list)
    ax.set_title(title, loc="left", color="#0C2742", pad=10)
    ax.set_xlabel(xlabel)
    ax.grid(axis="x")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.invert_yaxis()
    for idx, value in enumerate(values):
        ax.text(value + 0.6, idx, f"{br_number(value, 2)}%", va="center", fontsize=9, color="#243240")
    plt.tight_layout()
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def scaled_image(path: Path, max_width: float, max_height: float) -> Image:
    image = ImageReader(str(path))
    img_w, img_h = image.getSize()
    scale = min(max_width / img_w, max_height / img_h)
    return Image(str(path), width=img_w * scale, height=img_h * scale)


def draw_cover_image(canvas, image_path: Path, x: float, y: float, width: float, height: float) -> None:
    reader = ImageReader(str(image_path))
    img_w, img_h = reader.getSize()
    scale = max(width / img_w, height / img_h)
    draw_w = img_w * scale
    draw_h = img_h * scale
    draw_x = x + (width - draw_w) / 2
    draw_y = y + (height - draw_h) / 2
    canvas.drawImage(reader, draw_x, draw_y, draw_w, draw_h, preserveAspectRatio=True, mask="auto")


class CoverPage(Flowable):
    def __init__(self, title: str, subtitle: str, image_path: Path, members: list[str]) -> None:
        super().__init__()
        self.title = title
        self.subtitle = subtitle
        self.image_path = image_path
        self.members = members

    def wrap(self, availWidth: float, availHeight: float) -> tuple[float, float]:
        return PAGE_WIDTH, PAGE_HEIGHT

    def draw(self) -> None:
        canvas = self.canv
        draw_cover_image(canvas, self.image_path, 0, 0, PAGE_WIDTH, PAGE_HEIGHT)
        canvas.saveState()
        canvas.setFillColor(colors.Color(0.05, 0.15, 0.24, alpha=0.86))
        canvas.rect(0, 0, PAGE_WIDTH * 0.42, PAGE_HEIGHT, fill=1, stroke=0)
        canvas.setFillColor(WHITE)
        canvas.setFont(FONT_BOLD, 11)
        canvas.drawString(32, PAGE_HEIGHT - 42, str(canvas.getPageNumber()))
        canvas.setFont(FONT_REGULAR, 9)
        canvas.drawString(56, PAGE_HEIGHT - 42, "Boletim de Diagnóstico Fiscal 2026")
        canvas.drawRightString(PAGE_WIDTH - 28, PAGE_HEIGHT - 42, "Projeto Mini Ministério da Fazenda")
        canvas.setFillColor(LIME)
        canvas.rect(32, PAGE_HEIGHT - 156, 92, 8, fill=1, stroke=0)
        text = canvas.beginText(32, PAGE_HEIGHT - 188)
        text.setFont(FONT_BOLD, 24)
        text.setFillColor(WHITE)
        for line in self.title.split("\n"):
            text.textLine(line)
        canvas.drawText(text)
        text = canvas.beginText(32, PAGE_HEIGHT - 300)
        text.setFont(FONT_REGULAR, 12)
        text.setFillColor(colors.HexColor("#DCE7EF"))
        for line in self.subtitle.split("\n"):
            text.textLine(line)
        canvas.drawText(text)
        member_header = canvas.beginText(32, PAGE_HEIGHT - 372)
        member_header.setFont(FONT_BOLD, 10.5)
        member_header.setFillColor(WHITE)
        member_header.textLine("Equipe")
        canvas.drawText(member_header)
        member_text = canvas.beginText(32, PAGE_HEIGHT - 390)
        member_text.setFont(FONT_REGULAR, 9.8)
        member_text.setFillColor(colors.HexColor("#DCE7EF"))
        for member in self.members:
            member_text.textLine(member)
        canvas.drawText(member_text)
        canvas.setFillColor(WHITE)
        canvas.rect(0, 0, PAGE_WIDTH, 56, fill=1, stroke=0)
        canvas.setFillColor(NAVY)
        canvas.setFont(FONT_BOLD, 11)
        canvas.drawString(32, 34, "Pontifícia Universidade Católica de Minas Gerais")
        canvas.setFont(FONT_REGULAR, 9.5)
        canvas.drawString(32, 19, "Graduação em Ciências Econômicas | Eixo 4 - Projeto Mini Ministério da Fazenda")
        canvas.drawRightString(PAGE_WIDTH - 32, 27, "Etapa 1 | Março de 2026")
        canvas.restoreState()


class SectionPage(Flowable):
    def __init__(self, number: str, title: str, subtitle: str, image_path: Path, bookmark_name: str) -> None:
        super().__init__()
        self.number = number
        self.title = title
        self.subtitle = subtitle
        self.image_path = image_path
        self.toc_level = 0
        self.toc_title = title
        self.bookmark_name = bookmark_name

    def wrap(self, availWidth: float, availHeight: float) -> tuple[float, float]:
        return PAGE_WIDTH, PAGE_HEIGHT

    def draw(self) -> None:
        canvas = self.canv
        draw_cover_image(canvas, self.image_path, 0, 0, PAGE_WIDTH, PAGE_HEIGHT)
        canvas.saveState()
        canvas.setFillColor(colors.Color(0.04, 0.12, 0.20, alpha=0.68))
        canvas.rect(0, 0, PAGE_WIDTH, PAGE_HEIGHT, fill=1, stroke=0)
        canvas.setFillColor(colors.Color(0.43, 0.71, 0.24, alpha=0.92))
        canvas.rect(0, PAGE_HEIGHT - 86, PAGE_WIDTH, 10, fill=1, stroke=0)
        canvas.setFillColor(WHITE)
        canvas.setFont(FONT_BOLD, 11)
        canvas.drawString(32, PAGE_HEIGHT - 42, str(canvas.getPageNumber()))
        canvas.setFont(FONT_REGULAR, 9)
        canvas.drawString(56, PAGE_HEIGHT - 42, "Boletim de Diagnóstico Fiscal 2026")
        canvas.drawRightString(PAGE_WIDTH - 28, PAGE_HEIGHT - 42, "Relatório temático")
        canvas.setFont(FONT_BOLD, 56)
        canvas.setFillColor(LIME)
        canvas.drawString(40, PAGE_HEIGHT - 190, self.number)
        canvas.setFont(FONT_BOLD, 28)
        canvas.setFillColor(WHITE)
        canvas.drawString(40, PAGE_HEIGHT - 240, self.title)
        canvas.setFont(FONT_REGULAR, 12.5)
        canvas.setFillColor(colors.HexColor("#DDE7EE"))
        text = canvas.beginText(42, PAGE_HEIGHT - 274)
        text.textLines(self.subtitle)
        canvas.drawText(text)
        canvas.restoreState()


class FiscalDocTemplate(BaseDocTemplate):
    def __init__(self, filename: str) -> None:
        super().__init__(
            filename,
            pagesize=(PAGE_WIDTH, PAGE_HEIGHT),
            leftMargin=LEFT_MARGIN,
            rightMargin=RIGHT_MARGIN,
            topMargin=TOP_MARGIN,
            bottomMargin=BOTTOM_MARGIN,
            title="Diagnóstico Fiscal do Mato Grosso do Sul e de Campo Grande",
            author="PUC Minas | Curso de Ciências Econômicas | Eixo 4",
        )

        body_frame = Frame(
            LEFT_MARGIN,
            BOTTOM_MARGIN,
            BODY_WIDTH,
            BODY_HEIGHT,
            id="body",
            leftPadding=0,
            rightPadding=0,
            topPadding=0,
            bottomPadding=0,
        )
        full_frame = Frame(
            0,
            0,
            PAGE_WIDTH,
            PAGE_HEIGHT,
            id="full",
            leftPadding=0,
            rightPadding=0,
            topPadding=0,
            bottomPadding=0,
        )

        self.addPageTemplates(
            [
                PageTemplate(id="FullPage", frames=[full_frame]),
                PageTemplate(id="Body", frames=[body_frame], onPage=draw_body_page),
            ]
        )

    def afterFlowable(self, flowable) -> None:
        if not hasattr(flowable, "toc_level"):
            return
        title = getattr(flowable, "toc_title", None)
        if title is None and hasattr(flowable, "getPlainText"):
            title = flowable.getPlainText()
        bookmark = getattr(flowable, "bookmark_name", None)
        if not title or not bookmark:
            return
        self.canv.bookmarkPage(bookmark)
        self.notify("TOCEntry", (flowable.toc_level, title, self.page, bookmark))
        try:
            self.canv.addOutlineEntry(title, bookmark, level=flowable.toc_level, closed=False)
        except ValueError:
            pass


def draw_body_page(canvas, doc) -> None:
    page_number = canvas.getPageNumber()
    canvas.saveState()
    canvas.setFillColor(NAVY)
    canvas.setFont(FONT_BOLD, 10)
    canvas.drawString(LEFT_MARGIN, PAGE_HEIGHT - 24, str(page_number))
    canvas.setFont(FONT_REGULAR, 8.5)
    canvas.drawString(LEFT_MARGIN + 20, PAGE_HEIGHT - 24, "Boletim de Diagnóstico Fiscal do Mato Grosso do Sul e de Campo Grande")
    canvas.drawRightString(PAGE_WIDTH - RIGHT_MARGIN, PAGE_HEIGHT - 24, "Projeto Mini Ministério da Fazenda")
    canvas.setStrokeColor(LIGHT_LINE)
    canvas.setLineWidth(0.8)
    canvas.line(LEFT_MARGIN, PAGE_HEIGHT - 30, PAGE_WIDTH - RIGHT_MARGIN, PAGE_HEIGHT - 30)
    canvas.line(LEFT_MARGIN, 24, PAGE_WIDTH - RIGHT_MARGIN, 24)
    canvas.setFillColor(SLATE)
    canvas.setFont(FONT_REGULAR, 7.8)
    canvas.drawString(LEFT_MARGIN, 12, "Elaboração própria a partir de dados do SICONFI/Tesouro Nacional.")
    canvas.drawRightString(PAGE_WIDTH - RIGHT_MARGIN, 12, "Ano-base 2024")
    canvas.restoreState()


def build_styles():
    base = getSampleStyleSheet()
    return {
        "Body": ParagraphStyle(
            "Body",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=9.5,
            leading=13.2,
            alignment=TA_JUSTIFY,
            textColor=TEXT,
            spaceAfter=7,
        ),
        "Lead": ParagraphStyle(
            "Lead",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=10.6,
            leading=14.6,
            alignment=TA_JUSTIFY,
            textColor=TEXT,
            spaceAfter=9,
        ),
        "H1": ParagraphStyle("H1", parent=base["Heading1"], fontName=FONT_BOLD, fontSize=20, leading=24, textColor=NAVY, spaceAfter=8),
        "H2": ParagraphStyle("H2", parent=base["Heading2"], fontName=FONT_BOLD, fontSize=14.6, leading=18, textColor=NAVY, spaceAfter=6, spaceBefore=8),
        "H3": ParagraphStyle("H3", parent=base["Heading3"], fontName=FONT_BOLD, fontSize=11.6, leading=14, textColor=NAVY_LIGHT, spaceAfter=6, spaceBefore=8),
        "TOCTitle": ParagraphStyle("TOCTitle", parent=base["Heading1"], fontName=FONT_BOLD, fontSize=21, leading=24, textColor=NAVY, spaceAfter=10),
        "CardLabel": ParagraphStyle("CardLabel", parent=base["BodyText"], fontName=FONT_BOLD, fontSize=8.3, leading=10.2, textColor=SLATE),
        "CardValue": ParagraphStyle("CardValue", parent=base["BodyText"], fontName=FONT_BOLD, fontSize=15, leading=17, textColor=NAVY),
        "CardNote": ParagraphStyle("CardNote", parent=base["BodyText"], fontName=FONT_REGULAR, fontSize=8, leading=10, textColor=TEXT),
        "BoxTitle": ParagraphStyle("BoxTitle", parent=base["BodyText"], fontName=FONT_BOLD, fontSize=10.2, leading=12, textColor=NAVY, spaceAfter=4),
        "BoxBullet": ParagraphStyle("BoxBullet", parent=base["BodyText"], fontName=FONT_REGULAR, fontSize=8.8, leading=11.8, textColor=TEXT, leftIndent=12, bulletIndent=0, spaceAfter=2),
        "CaptionTitle": ParagraphStyle("CaptionTitle", parent=base["BodyText"], fontName=FONT_BOLD, fontSize=9, leading=10.5, textColor=NAVY, spaceAfter=1),
        "CaptionMeta": ParagraphStyle("CaptionMeta", parent=base["BodyText"], fontName=FONT_REGULAR, fontSize=7.4, leading=9.2, textColor=SLATE, spaceAfter=4),
        "PanelTitle": ParagraphStyle("PanelTitle", parent=base["BodyText"], fontName=FONT_BOLD, fontSize=10.4, leading=12.2, textColor=NAVY, spaceAfter=4),
        "PanelText": ParagraphStyle("PanelText", parent=base["BodyText"], fontName=FONT_REGULAR, fontSize=8.7, leading=11.4, textColor=TEXT),
        "Ref": ParagraphStyle("Ref", parent=base["BodyText"], fontName=FONT_REGULAR, fontSize=9.2, leading=12.8, textColor=TEXT, spaceAfter=7),
    }


def make_heading(text: str, style: ParagraphStyle, level: int, bookmark: str) -> Paragraph:
    paragraph = Paragraph(text, style)
    paragraph.toc_level = level
    paragraph.toc_title = text
    paragraph.bookmark_name = bookmark
    return paragraph


def caption_block_with_source(kind: str, number: int, title: str, source: str, styles) -> list:
    return [
        Paragraph(f"{kind} {number}", styles["CaptionTitle"]),
        Paragraph(
            f"{title}<br/>Dados em: R$ milhões<br/>Elaboração própria<br/>Fonte: {source}",
            styles["CaptionMeta"],
        ),
    ]


def caption_block(kind: str, number: int, title: str, styles) -> list:
    return caption_block_with_source(
        kind,
        number,
        title,
        "SICONFI. Sistema de Informações Contábeis e Fiscais do Setor Público Brasileiro. Tesouro Nacional.",
        styles,
    )


def make_variation_table(sheet: SheetData, rows: Iterable[str]) -> Table:
    header = ["Discriminação", *[str(year) for year in sheet.years], "Variação\nNominal", "Variação\n(%)"]
    body = [header]
    emphasis_rows = []
    for row_index, label in enumerate(rows, start=1):
        values = sheet.get(label)
        body.append(
            [
                sheet.label(label),
                *[br_money_millions(value, 2) for value in values],
                br_money_millions(nominal_change(values), 2),
                br_percent(pct_change(values), 1),
            ]
        )
        label_norm = normalize(label)
        if "SALDO" in label_norm or "RESULTADO" in label_norm or label_norm == label_norm.upper():
            emphasis_rows.append(row_index)

    table = Table(body, colWidths=[258, 58, 58, 58, 58, 58, 58, 78, 66], repeatRows=1)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 1), (-1, -1), 7.2),
        ("LEADING", (0, 1), (-1, -1), 8.6),
        ("GRID", (0, 0), (-1, -1), 0.35, LIGHT_LINE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_number in emphasis_rows:
        commands.extend(
            [
                ("FONTNAME", (0, row_number), (-1, row_number), FONT_BOLD),
                ("BACKGROUND", (0, row_number), (-1, row_number), colors.HexColor("#E6EEF5")),
            ]
        )
    table.setStyle(TableStyle(commands))
    return table


def metric_card(label: str, value: str, note: str, styles) -> Table:
    card = Table(
        [[Paragraph(label, styles["CardLabel"])], [Paragraph(value, styles["CardValue"])], [Paragraph(note, styles["CardNote"])]],
        colWidths=[170],
    )
    card.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), CARD_BG),
                ("BOX", (0, 0), (-1, -1), 0.8, LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return card


def bullet_box(title: str, bullets: list[str], styles) -> Table:
    rows = [[Paragraph(title, styles["BoxTitle"])]]
    rows.extend([[Paragraph(item, styles["BoxBullet"], bulletText="•")] for item in bullets])
    box = Table(rows, colWidths=[BODY_WIDTH])
    box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF4E8")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#C8D7B1")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    return box


def info_panel(title: str, text: str, styles, width: float) -> Table:
    panel = Table([[Paragraph(title, styles["PanelTitle"])], [Paragraph(text, styles["PanelText"])]], colWidths=[width])
    panel.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F7FA")),
                ("BOX", (0, 0), (-1, -1), 0.8, LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return panel


def overview_table(metrics: list[tuple[str, str, str, str]]) -> Table:
    header_style = ParagraphStyle(
        "OverviewHeader",
        fontName=FONT_BOLD,
        fontSize=8.2,
        leading=9.8,
        textColor=WHITE,
        alignment=TA_LEFT,
    )
    cell_style = ParagraphStyle(
        "OverviewCell",
        fontName=FONT_REGULAR,
        fontSize=8.0,
        leading=10.2,
        textColor=TEXT,
        alignment=TA_LEFT,
    )

    data = [
        [
            Paragraph("Indicador", header_style),
            Paragraph("Mato Grosso do Sul", header_style),
            Paragraph("Campo Grande", header_style),
            Paragraph("Leitura sintética", header_style),
        ]
    ]
    for indicador, ms, cg, leitura in metrics:
        data.append(
            [
                Paragraph(indicador, cell_style),
                Paragraph(ms, cell_style),
                Paragraph(cg, cell_style),
                Paragraph(leitura, cell_style),
            ]
        )

    table = Table(data, colWidths=[152, 112, 105, 363], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
                ("GRID", (0, 0), (-1, -1), 0.35, LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def narrative_table(headers: list[str], rows: list[list[str]], col_widths: list[float]) -> Table:
    header_style = ParagraphStyle(
        "NarrativeHeader",
        fontName=FONT_BOLD,
        fontSize=7.9,
        leading=9.2,
        textColor=WHITE,
        alignment=TA_LEFT,
    )
    cell_style = ParagraphStyle(
        "NarrativeCell",
        fontName=FONT_REGULAR,
        fontSize=7.6,
        leading=9.4,
        textColor=TEXT,
        alignment=TA_LEFT,
    )

    data = [[Paragraph(text, header_style) for text in headers]]
    for row in rows:
        data.append([Paragraph(text, cell_style) for text in row])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
                ("GRID", (0, 0), (-1, -1), 0.35, LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def generate_charts(ms1: SheetData, ms2: SheetData, cg1: SheetData, cg2: SheetData) -> dict[str, Path]:
    setup_plot()
    charts: dict[str, Path] = {}

    charts["ms_receitas_correntes"] = CHART_DIR / "grafico_ms_receitas_correntes.png"
    line_chart(
        charts["ms_receitas_correntes"],
        ms1.years,
        [
            ("Receita corrente", ms1.get("RECEITA CORRENTE"), "#0C2742"),
            ("Receita primária corrente", ms1.get("SALDO (A) - Receitas Primárias Correntes"), "#2E7B83"),
            ("ICMS", ms1.get("ICMS"), "#C49A45"),
        ],
        "Evolução das receitas correntes do estado",
        "R$ bilhões",
        divisor=1000,
    )

    charts["ms_receitas_capital"] = CHART_DIR / "grafico_ms_receitas_capital.png"
    stacked_bar_chart(
        charts["ms_receitas_capital"],
        ms1.years,
        [
            ("Receitas financeiras de capital", ms1.get("Receitas Financeiras de Capital"), "#8DB1D1"),
            ("Transferências de capital", ms1.get("Transferências de Capital"), "#2E7B83"),
        ],
        "Receitas de capital e componente primário do estado",
        "R$ milhões",
        line_item=("Receitas primárias de capital", ms1.get("SALDO (B) - Receitas Primárias de Capital"), "#C49A45"),
    )

    charts["ms_despesas_correntes"] = CHART_DIR / "grafico_ms_despesas_correntes.png"
    line_chart(
        charts["ms_despesas_correntes"],
        ms1.years,
        [
            ("Despesa corrente", ms1.get("DESPESA CORRENTE"), "#0C2742"),
            ("Despesa primária corrente", ms1.get("SALDO (D) - Despesas Primárias Correntes"), "#2E7B83"),
            ("Outras despesas correntes", ms1.get("Outras Despesas Correntes"), "#B85042"),
        ],
        "Evolução das despesas correntes do estado",
        "R$ bilhões",
        divisor=1000,
    )

    charts["ms_despesas_capital"] = CHART_DIR / "grafico_ms_despesas_capital.png"
    stacked_bar_chart(
        charts["ms_despesas_capital"],
        ms1.years,
        [
            ("Investimentos", ms1.get("Investimentos"), "#2E7B83"),
            ("Demais inversões", ms1.get("Demais Inversões"), "#A7C6DF"),
            ("Despesas financeiras de capital", ms1.get("Despesas Financeiras de Capital"), "#B85042"),
        ],
        "Despesa de capital do estado por componente",
        "R$ bilhões",
        divisor=1000,
        line_item=("Despesa primária de capital", ms1.get("SALDO (E) - Despesas Primárias de Capital"), "#C49A45"),
    )

    charts["ms_resultado_primario"] = CHART_DIR / "grafico_ms_resultado_primario.png"
    result_bar_chart(charts["ms_resultado_primario"], ms1.years, ms1.get("RESULTADO PRIMÁRIO (G = C - F)"), "Resultado primário do estado", "R$ bilhões", divisor=1000)

    charts["ms_resultado_orcamentario"] = CHART_DIR / "grafico_ms_resultado_orcamentario.png"
    result_bar_chart(charts["ms_resultado_orcamentario"], ms2.years, ms2.get("RESULTADO ORÇAMENTÁRIO"), "Resultado orçamentário do estado", "R$ bilhões", divisor=1000)

    charts["cg_receitas_correntes"] = CHART_DIR / "grafico_cg_receitas_correntes.png"
    line_chart(
        charts["cg_receitas_correntes"],
        cg1.years,
        [
            ("Receita corrente", cg1.get("RECEITA CORRENTE"), "#0C2742"),
            ("Receita primária corrente", cg1.get("SALDO (A) - Receitas Primárias Correntes"), "#2E7B83"),
            ("ISS", cg1.get("ISS"), "#C49A45"),
        ],
        "Evolução das receitas correntes de Campo Grande",
        "R$ bilhões",
        divisor=1000,
    )

    charts["cg_receitas_capital"] = CHART_DIR / "grafico_cg_receitas_capital.png"
    stacked_bar_chart(
        charts["cg_receitas_capital"],
        cg1.years,
        [
            ("Receitas financeiras de capital", cg1.get("Receitas Financeiras de Capital"), "#8DB1D1"),
            ("Transferências de capital", cg1.get("Transferências de Capital"), "#2E7B83"),
        ],
        "Receitas de capital de Campo Grande",
        "R$ milhões",
        line_item=("Receitas primárias de capital", cg1.get("SALDO (B) - Receitas Primárias de Capital"), "#C49A45"),
    )

    charts["cg_despesas_correntes"] = CHART_DIR / "grafico_cg_despesas_correntes.png"
    line_chart(
        charts["cg_despesas_correntes"],
        cg1.years,
        [
            ("Despesa corrente", cg1.get("DESPESA CORRENTE"), "#0C2742"),
            ("Despesa primária corrente", cg1.get("SALDO (D) - Despesas Primárias Correntes"), "#2E7B83"),
            ("Pessoal e encargos", cg1.get("Pessoal e Encargos Sociais"), "#B85042"),
        ],
        "Evolução das despesas correntes de Campo Grande",
        "R$ bilhões",
        divisor=1000,
    )

    charts["cg_despesas_capital"] = CHART_DIR / "grafico_cg_despesas_capital.png"
    stacked_bar_chart(
        charts["cg_despesas_capital"],
        cg1.years,
        [
            ("Investimentos", cg1.get("Investimentos"), "#2E7B83"),
            ("Demais inversões", cg1.get("Demais Inversões"), "#A7C6DF"),
            ("Despesas financeiras de capital", cg1.get("Despesas Financeiras de Capital"), "#B85042"),
        ],
        "Despesa de capital de Campo Grande por componente",
        "R$ milhões",
        line_item=("Despesa primária de capital", cg1.get("SALDO (E) - Despesas Primárias de Capital"), "#C49A45"),
    )

    charts["cg_resultado_primario"] = CHART_DIR / "grafico_cg_resultado_primario.png"
    result_bar_chart(charts["cg_resultado_primario"], cg1.years, cg1.get("RESULTADO PRIMÁRIO (G = C - F)"), "Resultado primário de Campo Grande", "R$ milhões")

    charts["cg_resultado_orcamentario"] = CHART_DIR / "grafico_cg_resultado_orcamentario.png"
    result_bar_chart(charts["cg_resultado_orcamentario"], cg2.years, cg2.get("RESULTADO ORÇAMENTÁRIO"), "Resultado orçamentário de Campo Grande", "R$ milhões")

    charts["pib_ms_evolucao"] = CHART_DIR / "grafico_pib_ms_evolucao.png"
    combo_bar_line_chart(
        charts["pib_ms_evolucao"],
        PIB_MS_YEARS,
        PIB_MS,
        [value / total * 100 for value, total in zip(PIB_MS, PIB_BRASIL)],
        "Evolução do PIB nominal de Mato Grosso do Sul e participação no PIB do Brasil",
        "PIB estadual (R$ bilhões)",
        "Participação no Brasil (%)",
        bar_divisor=1000,
    )

    charts["pib_ms_setores"] = CHART_DIR / "grafico_pib_ms_setores.png"
    horizontal_bar_chart(
        charts["pib_ms_setores"],
        list(PIB_VAB_SHARE_2023.keys()),
        list(PIB_VAB_SHARE_2023.values()),
        "Participação dos grandes setores no valor adicionado bruto de Mato Grosso do Sul em 2023",
        "Participação (%)",
    )

    charts["ms_resultado_primario_pib"] = CHART_DIR / "grafico_ms_resultado_primario_pib.png"
    result_bar_chart(
        charts["ms_resultado_primario_pib"],
        PIB_MS_YEARS,
        [valor / pib * 100 for valor, pib in zip(ms1.get("RESULTADO PRIMÁRIO (G = C - F)")[: len(PIB_MS_YEARS)], PIB_MS)],
        "Resultado primário do estado como proporção do PIB",
        "% do PIB",
        divisor=1,
    )

    return charts


def build_story(ms1: SheetData, ms2: SheetData, cg1: SheetData, cg2: SheetData, charts: dict[str, Path], styles) -> list:
    story: list = []

    ms_receita_corrente = ms1.get("RECEITA CORRENTE")
    ms_receita_primaria_total = ms1.get("RECEITA PRIMÁRIA TOTAL (C = A + B)")
    ms_despesa_primaria_total = ms1.get("DESPESA PRIMÁRIA TOTAL (F = D + E)")
    ms_resultado_primario = ms1.get("RESULTADO PRIMÁRIO (G = C - F)")
    ms_resultado_orc = ms2.get("RESULTADO ORÇAMENTÁRIO")
    ms_transfer_corr = ms1.get("Transferências Correntes")
    ms_receita_fin_corr = ms1.get("Receitas Financeiras Correntes")
    ms_pessoal = ms1.get("Pessoal e Encargos Sociais")
    ms_outras_corr = ms1.get("Outras Despesas Correntes")
    ms_receita_capital = ms1.get("RECEITA DE CAPITAL")
    ms_receita_prim_capital = ms1.get("SALDO (B) - Receitas Primárias de Capital")
    ms_desp_capital = ms1.get("DESPESA DE CAPITAL")
    ms_desp_prim_capital = ms1.get("SALDO (E) - Despesas Primárias de Capital")
    ms_resultado_primario_pct_pib = [valor / pib * 100 for valor, pib in zip(ms_resultado_primario[: len(PIB_MS_YEARS)], PIB_MS)]

    cg_receita_corrente = cg1.get("RECEITA CORRENTE")
    cg_receita_primaria_total = cg1.get("RECEITA PRIMÁRIA TOTAL (C = A + B)")
    cg_despesa_primaria_total = cg1.get("DESPESA PRIMÁRIA TOTAL (F = D + E)")
    cg_resultado_primario = cg1.get("RESULTADO PRIMÁRIO (G = C - F)")
    cg_resultado_orc = cg2.get("RESULTADO ORÇAMENTÁRIO")
    cg_pessoal = cg1.get("Pessoal e Encargos Sociais")
    cg_receita_capital = cg1.get("RECEITA DE CAPITAL")
    cg_receita_prim_capital = cg1.get("SALDO (B) - Receitas Primárias de Capital")
    cg_desp_capital = cg1.get("DESPESA DE CAPITAL")
    cg_desp_prim_capital = cg1.get("SALDO (E) - Despesas Primárias de Capital")

    ms_peak_fin_year, ms_peak_fin_value, _, _ = extrema_years(ms1.years, ms_receita_fin_corr)
    _, _, ms_best_primary_year, ms_best_primary_value = extrema_years(ms1.years, ms_resultado_primario)
    ms_peak_primary_def_year = ms1.years[ms_resultado_primario.index(min(ms_resultado_primario))]
    ms_worst_primary_def = min(ms_resultado_primario)
    pib_ms_part_brasil = [valor / total * 100 for valor, total in zip(PIB_MS, PIB_BRASIL)]

    cover_title = "Diagnóstico Fiscal\nMato Grosso do Sul\ne Campo Grande"
    cover_subtitle = (
        "Ano-base 2024 | Série histórica 2019-2024\n"
        "Etapa 1 do Projeto Mini Ministério da Fazenda\n"
        "Curso de Ciências Econômicas | Eixo 4 | PUC Minas"
    )

    story.extend([NextPageTemplate("FullPage"), CoverPage(cover_title, cover_subtitle, IMG_MS, TEAM_MEMBERS), NextPageTemplate("Body"), PageBreak()])
    story.append(Paragraph("Expediente", styles["H1"]))
    story.append(
        Paragraph(
            "Este relatório refere-se à Etapa 1 do Projeto Mini Ministério da Fazenda, desenvolvido no Curso de Ciências Econômicas da PUC Minas, no âmbito do Eixo 4. "
            "A análise utiliza dados fiscais do Estado do Mato Grosso do Sul e do Município de Campo Grande extraídos do SICONFI e está organizada em linguagem técnica voltada ao diagnóstico fiscal comparado.",
            styles["Lead"],
        )
    )
    expediente = Table(
        [
            [
                Paragraph(
                    "<b>Instituição e equipe</b><br/>Pontifícia Universidade Católica de Minas Gerais (PUC Minas)<br/>"
                    "Curso de Ciências Econômicas<br/>Eixo 4 - Projeto Mini Ministério da Fazenda<br/>"
                    "Etapa 1 - Diagnóstico da Situação Fiscal do Mato Grosso do Sul e de sua Capital<br/><br/>"
                    + "<br/>".join(TEAM_MEMBERS),
                    styles["Body"],
                ),
                Paragraph(
                    "<b>Fontes, método e recorte</b><br/>Base quantitativa: dados fiscais do Estado do Mato Grosso do Sul e de Campo Grande extraídos do SICONFI, "
                    "com série de 2019 a 2024, em valores nominais de R$ milhões.<br/><br/>Estrutura analítica: primeiro demonstrativo voltado ao resultado primário "
                    "e segundo demonstrativo voltado ao resultado orçamentário.<br/><br/>Referencial editorial: Boletim de Finanças dos Entes "
                    "Subnacionais 2025, da Secretaria do Tesouro Nacional.<br/><br/>Data de fechamento: março de 2026.",
                    styles["Body"],
                ),
            ]
        ],
        colWidths=[BODY_WIDTH / 2 - 10, BODY_WIDTH / 2 - 10],
        hAlign="LEFT",
    )
    expediente.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, LIGHT_LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.8, LIGHT_LINE),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend(
        [
            expediente,
            Spacer(1, 10),
            bullet_box(
                "Chaves de leitura",
                [
                    "O Estado do Mato Grosso do Sul elevou a receita corrente em "
                    f"{br_percent(pct_change(ms_receita_corrente), 1)}, mas encerrou 2024 com déficit orçamentário de {br_money_text(abs(ms_resultado_orc[-1]))}.",
                    "Campo Grande apresentou expansão mais moderada da receita corrente "
                    f"({br_percent(pct_change(cg_receita_corrente), 1)}), mas manteve déficits primários em todos os exercícios da série.",
                    "Nos dois entes, o avanço das despesas primárias e o peso das despesas financeiras impediram a conversão da expansão de receitas em equilíbrio orçamentário duradouro.",
                ],
                styles,
            ),
            PageBreak(),
        ]
    )

    story.append(Paragraph("Sumário", styles["TOCTitle"]))
    toc = TableOfContents()
    toc.levelStyles = [
        ParagraphStyle(name="TOCLevel0", fontName=FONT_BOLD, fontSize=10, leading=13, textColor=NAVY),
        ParagraphStyle(name="TOCLevel1", fontName=FONT_REGULAR, fontSize=8.9, leading=11.4, leftIndent=18, textColor=TEXT),
        ParagraphStyle(name="TOCLevel2", fontName=FONT_REGULAR, fontSize=8.2, leading=10.8, leftIndent=34, textColor=SLATE),
    ]
    story.append(toc)

    story.extend(
        [
            NextPageTemplate("FullPage"),
            PageBreak(),
            SectionPage("1", "1.0 - Introdução", "Objetivos, recorte temporal, fontes de dados e principais resultados\ndo diagnóstico fiscal do Estado do Mato Grosso do Sul e de Campo Grande.", IMG_MS, "sec_intro"),
            NextPageTemplate("Body"),
            PageBreak(),
        ]
    )
    story.append(make_heading("1.1 - Escopo analítico", styles["H2"], 1, "h2_11"))
    story.append(
        Paragraph(
            "Este relatório examina a trajetória das receitas, despesas, resultados primário e orçamentário do Estado do Mato Grosso do Sul e do Município de Campo Grande no período de 2019 a 2024. "
            "A leitura articula os dois demonstrativos fiscais constantes da base de dados: o primeiro, focado na decomposição do resultado primário; o segundo, na consolidação do resultado orçamentário.",
            styles["Lead"],
        )
    )
    story.append(
        Paragraph(
            "O documento está estruturado para destacar a evolução das contas públicas, a composição das principais rubricas fiscais e os pontos de inflexão observados ao longo da série. "
            "O objetivo é oferecer um diagnóstico coeso, comparável e tecnicamente consistente sobre a situação fiscal do estado e de sua capital.",
            styles["Body"],
        )
    )
    cards = Table(
        [
            [
                metric_card("Mato Grosso do Sul: receita corrente em 2024", br_money_text(ms_receita_corrente[-1]), f"Alta nominal de {br_percent(pct_change(ms_receita_corrente), 1)} entre 2019 e 2024.", styles),
                metric_card("Mato Grosso do Sul: resultado orçamentário em 2024", br_money_text(ms_resultado_orc[-1]), "Saldo deficitário no segundo demonstrativo, após superávit apenas em 2020.", styles),
                metric_card("Campo Grande: receita corrente em 2024", br_money_text(cg_receita_corrente[-1]), f"Crescimento nominal de {br_percent(pct_change(cg_receita_corrente), 1)} no período.", styles),
                metric_card("Campo Grande: resultado orçamentário em 2024", br_money_text(cg_resultado_orc[-1]), "Déficit mantido em toda a série, com piora relevante frente a 2023.", styles),
            ]
        ],
        colWidths=[BODY_WIDTH / 4 - 8] * 4,
        hAlign="LEFT",
    )
    cards.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 6)]))
    story.extend(
        [
            cards,
            Spacer(1, 12),
            overview_table(
                [
                    ("Receita corrente 2024", br_money_text(ms_receita_corrente[-1]), br_money_text(cg_receita_corrente[-1]), "Em ambos os casos houve expansão relevante da arrecadação, mas em ritmos distintos."),
                    ("Despesa primária total 2024", br_money_text(ms_despesa_primaria_total[-1]), br_money_text(cg_despesa_primaria_total[-1]), "A pressão de despesa permaneceu elevada e absorveu parcela expressiva do ganho de receita."),
                    ("Resultado primário 2024", br_money_text(ms_resultado_primario[-1]), br_money_text(cg_resultado_primario[-1]), "O estado voltou a registrar déficit primário robusto; a capital manteve déficit em toda a série."),
                    ("Resultado orçamentário 2024", br_money_text(ms_resultado_orc[-1]), br_money_text(cg_resultado_orc[-1]), "As receitas financeiras não foram suficientes para neutralizar as pressões de gasto."),
                ]
            ),
            Spacer(1, 10),
            bullet_box(
                "Principais conclusões preliminares",
                [
                    f"No Estado, a receita primária total cresceu {br_percent(pct_change(ms_receita_primaria_total), 1)}, enquanto a despesa primária total aumentou {br_percent(pct_change(ms_despesa_primaria_total), 1)}.",
                    f"Em Campo Grande, a receita primária total avançou {br_percent(pct_change(cg_receita_primaria_total), 1)}, ao passo que a despesa primária total subiu {br_percent(pct_change(cg_despesa_primaria_total), 1)}.",
                    "A diferença entre a dinâmica das receitas e das despesas foi mais crítica a partir de 2022, quando a série passou a registrar déficits mais profundos.",
                ],
                styles,
            ),
            NextPageTemplate("FullPage"),
            PageBreak(),
        ]
    )

    story.extend(
        [
            SectionPage("1.2", "1.2 - Panorama Econômico", "Contextualização do Produto Interno Bruto de Mato Grosso do Sul,\ncom foco na evolução recente, estrutura setorial e relação com o esforço fiscal.", IMG_MS, "sec_pib"),
            NextPageTemplate("Body"),
            PageBreak(),
        ]
    )
    story.append(make_heading("1.2.1 - Produto Interno Bruto do Mato Grosso do Sul", styles["H2"], 1, "h2_121"))
    story.append(
        Paragraph(
            "A contextualização do PIB nesta seção foi organizada em cinco eixos analíticos: evolução do PIB estadual, participação de Mato Grosso do Sul no PIB do Brasil, resultado primário como proporção do PIB, valor adicionado por atividades econômicas e participação percentual dos grandes setores. Esses tópicos foram confrontados com a publicação oficial da SEMADESC/IBGE referente ao PIB estadual de 2023.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            "A verificação confirma o núcleo das informações econômicas: o dado oficial mais recente do PIB das unidades da federação é o de 2023, divulgado em 14 de novembro de 2025 pelo IBGE. Nesse exercício, Mato Grosso do Sul alcançou PIB nominal de R$ 184,4 bilhões, crescimento real de 13,44%, participação de 1,69% no PIB brasileiro e posição de 15ª maior economia estadual. Como a série fiscal deste relatório vai até 2024, os indicadores relacionados ao PIB são apresentados até 2023 para preservar comparabilidade metodológica.",
            styles["Lead"],
        )
    )
    pib_cards = Table(
        [
            [
                metric_card("PIB de Mato Grosso do Sul em 2023", br_money_text(PIB_MS[-1]), "Último valor oficial disponível para o PIB estadual.", styles),
                metric_card("Crescimento real em 2023", br_percent(PIB_GROWTH_MS[-1], 2), "Segunda maior taxa entre as UFs, atrás apenas do Acre.", styles),
                metric_card("PIB per capita em 2023", f"R$ {br_number(PIB_MS_PER_CAPITA_2023, 2)}", "6º maior valor entre as unidades da federação.", styles),
                metric_card("Participação no PIB do Brasil", br_percent(PIB_MS_PART_BRASIL_2023, 2), "Mato Grosso do Sul respondeu por 15,9% do PIB do Centro-Oeste.", styles),
            ]
        ],
        colWidths=[BODY_WIDTH / 4 - 8] * 4,
        hAlign="LEFT",
    )
    pib_cards.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 6)]))
    story.extend([pib_cards, Spacer(1, 10)])
    pib_panels = Table(
        [
            [
                info_panel(
                    "Análise da Informação",
                    "Entre 2019 e 2023, o PIB nominal de Mato Grosso do Sul passou de R$ 106,9 bilhões para R$ 184,4 bilhões. No mesmo intervalo, a participação do estado no produto nacional saiu de cerca de 1,45% para 1,69%, refletindo aceleração mais intensa que a observada na média brasileira, sobretudo em 2022 e 2023.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
                info_panel(
                    "Entenda a Informação",
                    "O PIB e os resultados fiscais têm calendários de divulgação diferentes. Por isso, o relatório mantém a análise fiscal até 2024, mas limita a relação resultado/PIB até 2023, último ano com estimativa oficial do PIB estadual validada pelo IBGE e pela SEMADESC.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[BODY_WIDTH / 2 - 6, BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    pib_panels.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([pib_panels, Spacer(1, 10)])
    story.extend(caption_block_with_source("Gráfico PIB", 1, "Evolução do PIB nominal de Mato Grosso do Sul e participação no PIB do Brasil", "IBGE/SEMADESC. Produto Interno Bruto de Mato Grosso do Sul - 2023.", styles))
    story.append(scaled_image(charts["pib_ms_evolucao"], BODY_WIDTH, 210))
    pib_series_table = Table(
        [
            ["Ano", "Brasil", "Centro-Oeste", "Mato Grosso do Sul", "Part. de MS no Brasil (%)"],
            *[
                [
                    str(year),
                    br_money_millions(br, 2),
                    br_money_millions(co, 2),
                    br_money_millions(ms, 2),
                    br_percent(part, 2),
                ]
                for year, br, co, ms, part in zip(PIB_MS_YEARS, PIB_BRASIL, PIB_CENTRO_OESTE, PIB_MS, pib_ms_part_brasil)
            ],
        ],
        colWidths=[58, 150, 140, 150, 110],
        repeatRows=1,
    )
    pib_series_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 9.8),
                ("GRID", (0, 0), (-1, -1), 0.35, LIGHT_LINE),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend(caption_block_with_source("Tabela PIB", 1, "Comparação do PIB nominal do Brasil, do Centro-Oeste e de Mato Grosso do Sul (2019-2023)", "IBGE/SEMADESC. Produto Interno Bruto de Mato Grosso do Sul - 2023.", styles))
    story.append(pib_series_table)
    story.append(PageBreak())
    story.append(make_heading("1.2.2 - PIB, composição setorial e esforço fiscal", styles["H2"], 1, "h2_122"))
    story.append(
        Paragraph(
            "A estrutura produtiva estadual em 2023 reforça o papel da agropecuária na aceleração recente da economia. Segundo a SEMADESC/IBGE, a agropecuária respondeu por 25,92% do valor adicionado bruto estadual, a indústria por 22,35% e os serviços por 51,73%. Dentro de serviços, o maior peso individual foi da administração pública, defesa, educação e saúde públicas e seguridade social, com 16,63% do VAB, seguida por comércio e reparação de veículos automotores e motocicletas e por atividades imobiliárias.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"A incorporação do PIB ao diagnóstico fiscal ajuda a calibrar a magnitude do desequilíbrio primário do estado. Em 2019, o resultado primário correspondia a {br_percent(ms_resultado_primario_pct_pib[0], 2)} do PIB; em 2022, o indicador atingiu {br_percent(ms_resultado_primario_pct_pib[3], 2)}; e em 2023 permaneceu negativo em {br_percent(ms_resultado_primario_pct_pib[4], 2)}. Em outras palavras, mesmo com forte expansão do PIB nominal, o esforço fiscal não se converteu em superávit estrutural no fim da série disponível.",
            styles["Body"],
        )
    )
    pib_chart_grid = Table(
        [[scaled_image(charts["pib_ms_setores"], BODY_WIDTH / 2 - 8, 180), scaled_image(charts["ms_resultado_primario_pib"], BODY_WIDTH / 2 - 8, 180)]],
        colWidths=[BODY_WIDTH / 2 - 6, BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    pib_chart_grid.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend(caption_block_with_source("Gráfico PIB", 2, "Composição setorial do VAB em 2023 e resultado primário como % do PIB", "IBGE/SEMADESC. Produto Interno Bruto de Mato Grosso do Sul - 2023; SICONFI/Tesouro Nacional.", styles))
    story.append(pib_chart_grid)
    pib_fiscal_table = Table(
        [
            ["Ano", "PIB MS (R$ milhões)", "Crescimento real do PIB (%)", "Resultado primário (R$ milhões)", "Resultado primário / PIB (%)"],
            *[
                [
                    str(year),
                    br_money_millions(pib, 2),
                    br_percent(growth, 2),
                    br_money_millions(primary, 2),
                    br_percent(primary_pct, 2),
                ]
                for year, pib, growth, primary, primary_pct in zip(PIB_MS_YEARS, PIB_MS, PIB_GROWTH_MS, ms_resultado_primario[: len(PIB_MS_YEARS)], ms_resultado_primario_pct_pib)
            ],
        ],
        colWidths=[52, 128, 128, 150, 120],
        repeatRows=1,
    )
    pib_fiscal_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 9.8),
                ("GRID", (0, 0), (-1, -1), 0.35, LIGHT_LINE),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_BG]),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend(caption_block_with_source("Tabela PIB", 2, "PIB de Mato Grosso do Sul, crescimento real e resultado primário relativo ao PIB (2019-2023)", "IBGE/SEMADESC. Produto Interno Bruto de Mato Grosso do Sul - 2023; SICONFI/Tesouro Nacional.", styles))
    story.append(pib_fiscal_table)
    story.extend([NextPageTemplate("FullPage"), PageBreak()])

    story.extend(
        [
            SectionPage("2", "2.0 - Estado do Mato Grosso do Sul", "Leitura do primeiro e do segundo demonstrativos, com foco na composição das receitas,\ndas despesas e nos resultados fiscal e orçamentário do estado.", IMG_MS, "sec_estado"),
            NextPageTemplate("Body"),
            PageBreak(),
        ]
    )
    story.append(make_heading("2.1 - Primeiro demonstrativo", styles["H2"], 1, "h2_21"))
    story.append(
        Paragraph(
            "O primeiro demonstrativo permite decompor o resultado primário em grandes blocos de receita e despesa, tornando visíveis tanto a base de arrecadação do estado quanto as rubricas que mais pressionaram o gasto ao longo do período.",
            styles["Body"],
        )
    )
    story.append(make_heading("2.1.1 - Receita corrente e receitas primárias correntes", styles["H3"], 2, "h3_211"))
    story.append(
        Paragraph(
            f"A receita corrente do Mato Grosso do Sul passou de {br_money_text(ms_receita_corrente[0])} em 2019 para {br_money_text(ms_receita_corrente[-1])} em 2024, avanço nominal de {br_percent(pct_change(ms_receita_corrente), 1)}. "
            f"O principal vetor dessa expansão foi a arrecadação tributária própria, especialmente o ICMS, que alcançou {br_money_text(ms1.get('ICMS')[-1])} em 2024 e elevou sua participação na receita corrente de "
            f"{br_percent(share(ms1.get('ICMS')[0], ms_receita_corrente[0]), 1)} para {br_percent(share(ms1.get('ICMS')[-1], ms_receita_corrente[-1]), 1)}.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"As transferências correntes também cresceram no horizonte completo da série, mas perderam peso relativo: saíram de {br_percent(share(ms_transfer_corr[0], ms_receita_corrente[0]), 1)} da receita corrente em 2019 para "
            f"{br_percent(share(ms_transfer_corr[-1], ms_receita_corrente[-1]), 1)} em 2024. Já as receitas financeiras correntes, ainda que pouco representativas na composição total, atingiram pico de {br_money_text(ms_peak_fin_value)} em {ms_peak_fin_year} e encerraram 2024 em patamar bastante superior ao observado no início da série.",
            styles["Body"],
        )
    )
    story.extend(caption_block("Gráfico", 1, "Evolução das receitas correntes e do ICMS", styles))
    story.append(scaled_image(charts["ms_receitas_correntes"], BODY_WIDTH, 210))
    story.extend(caption_block("Tabela", 1, "Receitas correntes e saldo primário corrente do estado", styles))
    story.append(
        make_variation_table(
            ms1,
            ["RECEITA CORRENTE", "Impostos, Taxas e Contribuições de Melhoria", "ICMS", "IPVA", "Transferências Correntes", "Cota-Parte do FPE", "Receitas Financeiras Correntes", "Demais Receitas Correntes", "SALDO (A) - Receitas Primárias Correntes"],
        )
    )
    story.append(PageBreak())
    story.append(make_heading("2.1.2 - Receitas de capital, receitas primárias de capital e receita primária total", styles["H3"], 2, "h3_212"))
    story.append(
        Paragraph(
            f"As receitas de capital do estado permaneceram reduzidas quando comparadas às receitas correntes. Em 2024, somaram {br_money_text(ms_receita_capital[-1])}, equivalentes a pouco mais de "
            f"{br_percent(share(ms_receita_capital[-1], ms_receita_primaria_total[-1]), 1)} da receita primária total. O comportamento foi volátil ao longo da série, com maior nível em 2021 e posterior acomodação.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"A parcela primária dessas receitas continuou ancorada sobretudo em transferências de capital. O saldo primário de capital passou de {br_money_text(ms_receita_prim_capital[0])} para "
            f"{br_money_text(ms_receita_prim_capital[-1])}, enquanto a receita primária total atingiu {br_money_text(ms_receita_primaria_total[-1])} em 2024. Isso confirma que o ajuste estrutural da arrecadação estadual depende muito mais da base corrente do que de ingressos de capital.",
            styles["Body"],
        )
    )
    story.extend(caption_block("Gráfico", 2, "Receitas de capital do estado por componente", styles))
    story.append(scaled_image(charts["ms_receitas_capital"], BODY_WIDTH, 210))
    story.extend(caption_block("Tabela", 2, "Receitas de capital e receita primária total do estado", styles))
    story.append(
        make_variation_table(
            ms1,
            ["RECEITA DE CAPITAL", "Receitas Financeiras de Capital", "Operações de Crédito", "Alienação de Bens", "Amortização de Empréstimos", "Transferências de Capital", "SALDO (B) - Receitas Primárias de Capital", "RECEITA PRIMÁRIA TOTAL (C = A + B)"],
        )
    )
    story.append(PageBreak())
    story.append(make_heading("2.1.3 - Despesas correntes e despesas primárias correntes", styles["H3"], 2, "h3_213"))
    story.append(
        Paragraph(
            f"A despesa corrente estadual cresceu {br_percent(pct_change(ms1.get('DESPESA CORRENTE')), 1)} entre 2019 e 2024, chegando a {br_money_text(ms1.get('DESPESA CORRENTE')[-1])}. "
            f"O maior bloco continuou sendo pessoal e encargos sociais, que somaram {br_money_text(ms_pessoal[-1])} em 2024. Contudo, o item mais dinâmico da série foi o de outras despesas correntes, cuja expansão nominal atingiu {br_percent(pct_change(ms_outras_corr), 1)}, muito acima da observada no grupo de pessoal.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"Essa mudança alterou a composição do gasto corrente. A participação do grupo de pessoal recuou de {br_percent(share(ms_pessoal[0], ms1.get('DESPESA CORRENTE')[0]), 1)} para {br_percent(share(ms_pessoal[-1], ms1.get('DESPESA CORRENTE')[-1]), 1)}, "
            f"enquanto as demais despesas correntes passaram de {br_percent(share(ms_outras_corr[0], ms1.get('DESPESA CORRENTE')[0]), 1)} para {br_percent(share(ms_outras_corr[-1], ms1.get('DESPESA CORRENTE')[-1]), 1)}. "
            "Em 2024, a despesa primária corrente praticamente igualou o saldo das receitas primárias correntes, o que estreitou a margem de acomodação do resultado fiscal.",
            styles["Body"],
        )
    )
    story.extend(caption_block("Gráfico", 3, "Despesa corrente e despesa primária corrente do estado", styles))
    story.append(scaled_image(charts["ms_despesas_correntes"], BODY_WIDTH, 210))
    story.extend(caption_block("Tabela", 3, "Despesas correntes do estado", styles))
    story.append(make_variation_table(ms1, ["DESPESA CORRENTE", "Pessoal e Encargos Sociais", "Juros e Encargos da Dívida", "Outras Despesas Correntes", "SALDO (D) - Despesas Primárias Correntes"]))
    story.append(PageBreak())
    story.append(make_heading("2.1.4 - Despesas de capital, despesas primárias de capital e despesa primária total", styles["H3"], 2, "h3_214"))
    story.append(
        Paragraph(
            f"As despesas de capital foram o componente mais volátil do gasto estadual. Após saírem de {br_money_text(ms_desp_capital[0])} em 2019, atingiram pico de {br_money_text(max(ms_desp_capital))} em 2022 e encerraram 2024 em {br_money_text(ms_desp_capital[-1])}. "
            "O movimento foi fortemente associado aos investimentos, que seguiram como principal rubrica do grupo.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"Mesmo com a desaceleração recente, a despesa primária de capital fechou 2024 em {br_money_text(ms_desp_prim_capital[-1])}, valor muito superior ao saldo primário de capital das receitas "
            f"({br_money_text(ms_receita_prim_capital[-1])}). Essa assimetria ajuda a explicar por que a despesa primária total chegou a {br_money_text(ms_despesa_primaria_total[-1])}, mantendo elevada a necessidade de financiamento do estado.",
            styles["Body"],
        )
    )
    story.extend(caption_block("Gráfico", 4, "Despesa de capital do estado por componente", styles))
    story.append(scaled_image(charts["ms_despesas_capital"], BODY_WIDTH, 205))
    story.extend(caption_block("Tabela", 4, "Despesas de capital e despesa primária total do estado", styles))
    story.append(make_variation_table(ms1, ["DESPESA DE CAPITAL", "Investimentos", "Demais Inversões", "Despesas Financeiras de Capital", "Amortização da Dívida", "SALDO (E) - Despesas Primárias de Capital", "DESPESA PRIMÁRIA TOTAL (F = D + E)"]))
    story.append(PageBreak())
    story.append(make_heading("2.1.5 - Resultado primário", styles["H3"], 2, "h3_215"))
    story.append(
        Paragraph(
            f"O estado registrou déficit primário em cinco dos seis anos analisados. O único superávit ocorreu em {ms_best_primary_year}, quando o resultado alcançou {br_money_text(ms_best_primary_value)}. "
            f"O pior resultado da série foi observado em {ms_peak_primary_def_year}, com saldo de {br_money_text(ms_worst_primary_def)}.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"Em 2024, o resultado primário permaneceu negativo em {br_money_text(ms_resultado_primario[-1])}, o equivalente a {br_percent(abs(ms_resultado_primario[-1]) / ms_receita_primaria_total[-1] * 100, 1)} da receita primária total do exercício. "
            "Isso revela que, embora a arrecadação tenha avançado, ela não foi suficiente para compensar a expansão das despesas primárias, sobretudo nas rubricas de capital e em outras despesas correntes.",
            styles["Body"],
        )
    )
    story.extend(caption_block("Gráfico", 5, "Resultado primário do estado", styles))
    story.append(scaled_image(charts["ms_resultado_primario"], BODY_WIDTH, 198))
    story.extend(caption_block("Tabela", 5, "Consolidação do resultado primário do estado", styles))
    story.append(make_variation_table(ms2, ["RECEITAS PRIMÁRIAS", "DESPESAS PRIMÁRIAS", "RESULTADO PRIMÁRIO"]))
    story.append(PageBreak())
    story.append(make_heading("2.2 - Segundo demonstrativo", styles["H2"], 1, "h2_22"))
    story.append(
        Paragraph(
            "O segundo demonstrativo consolida o resultado orçamentário ao incorporar receitas e despesas financeiras ao resultado primário. Sua leitura permite verificar se o saldo final do exercício foi capaz de absorver a necessidade de financiamento gerada pela dinâmica primária.",
            styles["Body"],
        )
    )
    story.append(
        Paragraph(
            f"No caso do Mato Grosso do Sul, apenas 2020 apresentou superávit orçamentário, de {br_money_text(ms_resultado_orc[1])}. Em 2024, as receitas financeiras somaram {br_money_text(ms2.get('RECEITAS FINANCEIRAS')[-1])}, "
            f"mas as despesas financeiras chegaram a {br_money_text(ms2.get('DESPESAS FINANCEIRAS')[-1])}. Como o componente financeiro não compensou o déficit primário, o resultado orçamentário encerrou o exercício em {br_money_text(ms_resultado_orc[-1])}.",
            styles["Body"],
        )
    )
    story.extend(caption_block("Gráfico", 6, "Resultado orçamentário do estado", styles))
    story.append(scaled_image(charts["ms_resultado_orcamentario"], BODY_WIDTH, 198))
    story.extend(caption_block("Tabela", 6, "Segundo demonstrativo do estado do Mato Grosso do Sul", styles))
    story.append(make_variation_table(ms2, ["RECEITAS PRIMÁRIAS", "DESPESAS PRIMÁRIAS", "RESULTADO PRIMÁRIO", "RECEITAS FINANCEIRAS", "DESPESAS FINANCEIRAS", "RESULTADO ORÇAMENTÁRIO"]))
    story.extend([NextPageTemplate("FullPage"), PageBreak()])
    story.extend(
        [
            SectionPage("3", "3.0 - Análise das contas da capital - Campo Grande", "Diagnóstico da composição das receitas e despesas municipais e da permanência\nde déficits primários e orçamentários ao longo de toda a série.", IMG_CG, "sec_cg"),
            NextPageTemplate("Body"),
            PageBreak(),
        ]
    )
    story.append(make_heading("3.1 - Primeiro demonstrativo", styles["H2"], 1, "h2_31"))
    story.append(Paragraph("No município de Campo Grande, o primeiro demonstrativo também expõe a formação do resultado primário. A série revela crescimento relevante da arrecadação, porém acompanhado por uma trajetória de gasto que manteve o saldo fiscal em terreno negativo em todos os exercícios.", styles["Body"]))
    story.append(make_heading("3.1.1 - Receita corrente e receitas primárias correntes", styles["H3"], 2, "h3_311"))
    story.append(Paragraph(f"A receita corrente municipal passou de {br_money_text(cg_receita_corrente[0])} em 2019 para {br_money_text(cg_receita_corrente[-1])} em 2024, o que representa crescimento nominal de {br_percent(pct_change(cg_receita_corrente), 1)}. As transferências correntes permaneceram como maior componente da receita, mas o ISS ganhou relevância e alcançou {br_money_text(cg1.get('ISS')[-1])} ao final da série.", styles["Body"]))
    story.append(Paragraph(f"As receitas financeiras correntes seguiram trajetória ascendente até {cg1.years[cg1.get('Receitas Financeiras Correntes').index(max(cg1.get('Receitas Financeiras Correntes')))]}, quando atingiram {br_money_text(max(cg1.get('Receitas Financeiras Correntes')))}, e recuaram em 2024. Ainda assim, o saldo das receitas primárias correntes chegou a {br_money_text(cg1.get('SALDO (A) - Receitas Primárias Correntes')[-1])}, confirmando que a base corrente do município se expandiu de forma consistente, embora insuficiente para garantir equilíbrio fiscal.", styles["Body"]))
    story.extend(caption_block("Gráfico", 7, "Evolução das receitas correntes de Campo Grande", styles))
    story.append(scaled_image(charts["cg_receitas_correntes"], BODY_WIDTH, 210))
    story.extend(caption_block("Tabela", 7, "Receitas correntes e saldo primário corrente de Campo Grande", styles))
    story.append(make_variation_table(cg1, ["RECEITA CORRENTE", "Impostos, Taxas e Contribuições de Melhoria", "ISS", "IPTU", "Transferências Correntes", "Cota-Parte do FPM", "Receitas Financeiras Correntes", "Demais Receitas Correntes", "SALDO (A) - Receitas Primárias Correntes"]))
    story.append(PageBreak())
    story.append(make_heading("3.1.2 - Receitas de capital, receitas primárias de capital e receita primária total", styles["H3"], 2, "h3_312"))
    story.append(Paragraph(f"As receitas de capital de Campo Grande mantiveram participação reduzida no total arrecadado. Em 2024, elas alcançaram {br_money_text(cg_receita_capital[-1])}, valor próximo ao observado em 2023, mas superior ao nível de 2019. A série mostra forte volatilidade das receitas financeiras de capital, sobretudo nas operações de crédito, e avanço das transferências de capital ao fim do período.", styles["Body"]))
    story.append(Paragraph(f"Essa composição fez com que o saldo das receitas primárias de capital chegasse a {br_money_text(cg_receita_prim_capital[-1])} em 2024, levando a receita primária total do município a {br_money_text(cg_receita_primaria_total[-1])}. Ainda assim, a contribuição do bloco de capital permaneceu modesta quando comparada à relevância das receitas correntes na estrutura fiscal da capital.", styles["Body"]))
    story.extend(caption_block("Gráfico", 8, "Receitas de capital de Campo Grande", styles))
    story.append(scaled_image(charts["cg_receitas_capital"], BODY_WIDTH, 205))
    story.extend(caption_block("Tabela", 8, "Receitas de capital e receita primária total de Campo Grande", styles))
    story.append(make_variation_table(cg1, ["RECEITA DE CAPITAL", "Receitas Financeiras de Capital", "Operações de Crédito", "Alienação de Bens", "Amortização de Empréstimos", "Transferências de Capital", "SALDO (B) - Receitas Primárias de Capital", "RECEITA PRIMÁRIA TOTAL (C = A + B)"]))
    story.append(PageBreak())
    story.append(make_heading("3.1.3 - Despesas correntes e despesas primárias correntes", styles["H3"], 2, "h3_313"))
    story.append(Paragraph(f"As despesas correntes do município cresceram {br_percent(pct_change(cg1.get('DESPESA CORRENTE')), 1)} entre 2019 e 2024, encerrando a série em {br_money_text(cg1.get('DESPESA CORRENTE')[-1])}. O principal grupo permaneceu sendo pessoal e encargos sociais, cuja despesa avançou para {br_money_text(cg_pessoal[-1])} no último ano do painel.", styles["Body"]))
    story.append(Paragraph(f"A despesa primária corrente chegou a {br_money_text(cg1.get('SALDO (D) - Despesas Primárias Correntes')[-1])} em 2024, superando o saldo das receitas primárias correntes em aproximadamente {br_money_text(cg1.get('SALDO (D) - Despesas Primárias Correntes')[-1] - cg1.get('SALDO (A) - Receitas Primárias Correntes')[-1])}. Esse descompasso mostra que o desequilíbrio fiscal da capital não decorre de um único item, mas de uma pressão de custeio relativamente disseminada.", styles["Body"]))
    story.extend(caption_block("Gráfico", 9, "Despesa corrente de Campo Grande", styles))
    story.append(scaled_image(charts["cg_despesas_correntes"], BODY_WIDTH, 205))
    story.extend(caption_block("Tabela", 9, "Despesas correntes de Campo Grande", styles))
    story.append(make_variation_table(cg1, ["DESPESA CORRENTE", "Pessoal e Encargos Sociais", "Juros e Encargos da Dívida", "Outras Despesas Correntes", "SALDO (D) - Despesas Primárias Correntes"]))
    story.append(PageBreak())
    story.append(make_heading("3.1.4 - Despesas de capital, despesas primárias de capital e despesa primária total", styles["H3"], 2, "h3_314"))
    story.append(Paragraph(f"A despesa de capital de Campo Grande saiu de {br_money_text(cg_desp_capital[0])} em 2019 para {br_money_text(cg_desp_capital[-1])} em 2024, crescimento de {br_percent(pct_change(cg_desp_capital), 1)}. Os investimentos responderam pela maior parcela desse grupo, alcançando {br_money_text(cg1.get('Investimentos')[-1])} em 2024.", styles["Body"]))
    story.append(Paragraph(f"A despesa primária de capital encerrou 2024 em {br_money_text(cg_desp_prim_capital[-1])}, novamente acima do ingresso primário de capital ({br_money_text(cg_receita_prim_capital[-1])}). Somada à pressão das despesas correntes, essa diferença levou a despesa primária total do município a {br_money_text(cg_despesa_primaria_total[-1])}.", styles["Body"]))
    story.extend(caption_block("Gráfico", 10, "Despesa de capital de Campo Grande por componente", styles))
    story.append(scaled_image(charts["cg_despesas_capital"], BODY_WIDTH, 205))
    story.extend(caption_block("Tabela", 10, "Despesas de capital e despesa primária total de Campo Grande", styles))
    story.append(make_variation_table(cg1, ["DESPESA DE CAPITAL", "Investimentos", "Demais Inversões", "Despesas Financeiras de Capital", "Amortização da Dívida", "SALDO (E) - Despesas Primárias de Capital", "DESPESA PRIMÁRIA TOTAL (F = D + E)"]))
    story.append(PageBreak())
    story.append(make_heading("3.1.5 - Resultado primário", styles["H3"], 2, "h3_315"))
    story.append(Paragraph(f"Campo Grande apresentou déficit primário em todos os anos da série. O menor desequilíbrio foi observado em {cg1.years[cg_resultado_primario.index(max(cg_resultado_primario))]}, com saldo de {br_money_text(max(cg_resultado_primario))}, enquanto o maior déficit ocorreu em {cg1.years[cg_resultado_primario.index(min(cg_resultado_primario))]}, quando o resultado atingiu {br_money_text(min(cg_resultado_primario))}.", styles["Body"]))
    story.append(Paragraph(f"Em 2024, o déficit primário municipal foi de {br_money_text(cg_resultado_primario[-1])}, o que corresponde a {br_percent(abs(cg_resultado_primario[-1]) / cg_receita_primaria_total[-1] * 100, 1)} da receita primária total. Mesmo com expansão da arrecadação própria e das transferências, a capital não conseguiu converter esse ganho em recomposição do saldo fiscal.", styles["Body"]))
    story.extend(caption_block("Gráfico", 11, "Resultado primário de Campo Grande", styles))
    story.append(scaled_image(charts["cg_resultado_primario"], BODY_WIDTH, 198))
    story.extend(caption_block("Tabela", 11, "Consolidação do resultado primário de Campo Grande", styles))
    story.append(make_variation_table(cg2, ["RECEITAS PRIMÁRIAS", "DESPESAS PRIMÁRIAS", "RESULTADO PRIMÁRIO"]))
    story.append(PageBreak())
    story.append(make_heading("3.2 - Segundo demonstrativo", styles["H2"], 1, "h2_32"))
    story.append(Paragraph("No segundo demonstrativo, a situação da capital torna-se ainda menos favorável porque o resultado orçamentário incorpora despesas financeiras superiores às receitas financeiras em boa parte da série.", styles["Body"]))
    story.append(Paragraph(f"Em 2024, Campo Grande arrecadou {br_money_text(cg2.get('RECEITAS FINANCEIRAS')[-1])} em receitas financeiras e registrou {br_money_text(cg2.get('DESPESAS FINANCEIRAS')[-1])} em despesas dessa natureza. Com isso, o déficit orçamentário atingiu {br_money_text(cg_resultado_orc[-1])}, bastante acima do saldo de 2023. A trajetória confirma que o município ainda não dispõe de margem financeira suficiente para compensar seu desequilíbrio primário.", styles["Body"]))
    story.extend(caption_block("Gráfico", 12, "Resultado orçamentário de Campo Grande", styles))
    story.append(scaled_image(charts["cg_resultado_orcamentario"], BODY_WIDTH, 198))
    story.extend(caption_block("Tabela", 12, "Segundo demonstrativo de Campo Grande", styles))
    story.append(make_variation_table(cg2, ["RECEITAS PRIMÁRIAS", "DESPESAS PRIMÁRIAS", "RESULTADO PRIMÁRIO", "RECEITAS FINANCEIRAS", "DESPESAS FINANCEIRAS", "RESULTADO ORÇAMENTÁRIO"]))
    story.extend([NextPageTemplate("FullPage"), PageBreak()])
    story.extend([SectionPage("4", "4.0 - Conclusão", "Síntese interpretativa dos resultados e dos principais vetores de pressão fiscal\nobservados no estado e na capital ao longo da série 2019-2024.", IMG_MS, "sec_conclusao"), NextPageTemplate("Body"), PageBreak()])
    story.append(Paragraph(f"O diagnóstico consolidado mostra que o Mato Grosso do Sul ampliou sua receita corrente em {br_percent(pct_change(ms_receita_corrente), 1)} entre 2019 e 2024, com destaque para o fortalecimento da arrecadação de ICMS e para a expansão do saldo das receitas primárias correntes. Apesar disso, a combinação entre aumento de outras despesas correntes, manutenção de elevado gasto com pessoal e forte oscilação das despesas de capital impediu a convergência para resultados primários positivos em caráter duradouro. O estado só apresentou superávit primário e orçamentário em 2020, encerrando 2024 com déficit orçamentário de {br_money_text(ms_resultado_orc[-1])}.", styles["Lead"]))
    story.append(Paragraph(f"Na capital, a expansão da arrecadação foi acompanhada por crescimento equivalente ou superior em segmentos relevantes da despesa, especialmente pessoal, encargos e investimentos. O município elevou sua receita corrente em {br_percent(pct_change(cg_receita_corrente), 1)}, mas manteve déficit primário em toda a série e agravou o déficit orçamentário no encerramento do período. Em 2024, o resultado orçamentário negativo de {br_money_text(cg_resultado_orc[-1])} evidencia que a melhora da base arrecadatória ainda não foi capaz de reequilibrar o fluxo de gastos e o componente financeiro.", styles["Lead"]))
    story.append(make_heading("4.1 - Fatores explicativos dos resultados fiscais", styles["H2"], 1, "h2_41"))
    story.append(Paragraph("As evidências externas consultadas permitem contextualizar os principais achados do relatório sem atribuir causalidade mecânica a um único evento. A interpretação abaixo deve ser lida como uma análise de fatores temporalmente compatíveis com a trajetória observada nas receitas, despesas e resultados fiscais do Estado e da capital.", styles["Body"]))
    explanatory_panels = Table(
        [
            [
                info_panel(
                    "Leitura metodológica",
                    "Os demonstrativos fiscais mostram o resultado contábil da série 2019-2024. Já os fatores externos reunidos nesta subseção funcionam como elementos de contexto institucional e econômico que ajudam a interpretar mudanças de trajetória, sobretudo em 2020, 2022 e 2023.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
                info_panel(
                    "Chave de interpretação",
                    "Quando o texto relaciona mudanças fiscais a medidas legais ou ao ciclo econômico, a leitura é inferencial. Em outras palavras, trata-se de uma associação sustentada por coincidência temporal, magnitude dos eventos e aderência às fontes oficiais consultadas.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[BODY_WIDTH / 2 - 6, BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    explanatory_panels.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([explanatory_panels, Spacer(1, 10)])
    story.append(Paragraph("No caso do Estado, o superávit isolado de 2020 coincide com um ambiente excepcional de suporte federativo. A Lei Complementar nº 173, de 27 de maio de 2020, instituiu o Programa Federativo de Enfrentamento ao Coronavírus, enquanto a Lei nº 14.041, de 18 de agosto de 2020, consolidou o apoio financeiro da União para compensar perdas nos fundos de participação. Essas medidas ajudam a interpretar por que o único resultado positivo da série ocorreu em um exercício fora do padrão histórico recente.", styles["Body"]))
    story.append(Paragraph("Já a forte expansão da receita estadual observada no relatório é compatível com o ciclo econômico favorável captado pelas Contas Regionais. Segundo a SEMADESC e o IBGE, o PIB de Mato Grosso do Sul cresceu 13,44% em 2023, com avanço de 12,6% da agropecuária, impulsionado por safra recorde e pelo desempenho de culturas e cadeias como soja, milho, cana, bovinos, suínos e aves. Ao mesmo tempo, a persistência dos déficits a partir de 2022 deve ser lida à luz de dois condicionantes adicionais: a limitação das alíquotas de ICMS sobre combustíveis, energia, comunicações e transporte coletivo promovida pela Lei Complementar nº 194, de 23 de junho de 2022, e a rigidez da despesa. Em 2024, a SEFAZ-MS registrou receita de ICMS de R$ 17,03 bilhões e despesa com pessoal equivalente a 53,90% da receita corrente líquida, quadro compatível com crescimento de arrecadação sem recomposição estrutural do resultado fiscal.", styles["Body"]))
    story.append(Paragraph("Em Campo Grande, o padrão de desequilíbrio também encontra respaldo em evidências institucionais externas. No Relatório de Gestão Fiscal consolidado de 2023, a despesa total com pessoal atingiu 56,86% da receita corrente líquida ajustada, acima do limite legal de 54%, reforçando a hipótese de elevada rigidez orçamentária. Em paralelo, o Relatório de Atividades de 2024 destaca ampliação de projetos em educação, saúde, mobilidade e infraestrutura viária. Em termos interpretativos, esse conjunto sugere que a capital ampliou sua agenda de entregas públicas em um ambiente fiscal já pressionado, o que ajuda a explicar a manutenção de déficits primários em toda a série e o agravamento do resultado orçamentário ao fim do período.", styles["Body"]))
    story.append(
        bullet_box(
            "Fatores de contexto associados aos achados",
            [
                "O superávit de 2020 deve ser lido como um ponto fora da curva, influenciado por medidas extraordinárias de suporte financeiro da União durante a pandemia.",
                "A aceleração da arrecadação estadual foi favorecida pelo ciclo econômico de 2023, especialmente pelo peso do agronegócio na estrutura produtiva sul-mato-grossense.",
                "A piora dos saldos a partir de 2022 é compatível com um ambiente de menor flexibilidade tributária sobre bases relevantes de ICMS e com persistência de despesas rígidas.",
                "Na capital, o quadro de despesa com pessoal acima do limite e a expansão simultânea de frentes de atuação pública ajudam a interpretar o caráter estrutural do desequilíbrio fiscal.",
            ],
            styles,
        )
    )
    story.append(make_heading("4.2 - Cenários prospectivos e projeção analítica", styles["H2"], 1, "h2_42"))
    story.append(Paragraph("Para ampliar o valor analítico do diagnóstico, esta subseção apresenta cenários qualitativos para o triênio 2025-2027. Não se trata de projeção econométrica, mas de uma leitura prospectiva baseada na direção recente das variáveis fiscais, nas restrições institucionais identificadas e no ambiente macroeconômico que condiciona a arrecadação e o custo do gasto.", styles["Body"]))
    scenario_panels = Table(
        [
            [
                info_panel(
                    "Como ler os cenários",
                    "Cada cenário combina hipóteses sobre atividade econômica, elasticidade da receita, rigidez da despesa e capacidade de gestão. O objetivo é mostrar como diferentes escolhas de política e diferentes choques externos podem alterar a trajetória fiscal dos dois entes.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
                info_panel(
                    "Utilidade prática",
                    "Os cenários funcionam como instrumento de planejamento. Eles ajudam a hierarquizar riscos, definir gatilhos de contingência e orientar a alocação de esforços entre ajuste de curto prazo, preservação de serviços essenciais e retomada de capacidade de investimento.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[BODY_WIDTH / 2 - 6, BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    scenario_panels.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([scenario_panels, Spacer(1, 10)])
    story.append(
        narrative_table(
            ["Cenário", "Condições e gatilhos", "Implicações fiscais esperadas"],
            [
                [
                    "Reequilíbrio gradual",
                    "Crescimento econômico moderado, manutenção de arrecadação corrente em linha com o PIB nominal, contenção seletiva de despesas correntes, revisão de contratos e maior priorização do investimento de maior retorno social e econômico.",
                    "No Mato Grosso do Sul, o déficit primário tenderia a diminuir progressivamente e o resultado orçamentário poderia se aproximar do equilíbrio. Em Campo Grande, o déficit seguiria presente no curto prazo, mas com redução gradual e recuperação parcial da capacidade de investimento prioritário.",
                ],
                [
                    "Inércia fiscal",
                    "Desaceleração da receita após o pico recente, ausência de mudanças estruturais na gestão do gasto, manutenção da pressão de pessoal e de custeio e execução de investimentos sem filtro mais rigoroso de prioridade.",
                    "Os dois entes preservariam déficits recorrentes. O Estado manteria dificuldade de converter crescimento de receita em poupança corrente, enquanto a capital seguiria com desequilíbrio estrutural, maior dependência de receitas extraordinárias e menor espaço para absorver choques.",
                ],
                [
                    "Estresse adverso",
                    "Choque negativo sobre agronegócio, ICMS ou transferências, juros elevados, restrições de crédito e eventos climáticos que pressionem receita e gasto ao mesmo tempo, sem resposta gerencial suficientemente rápida.",
                    "A deterioração fiscal se aprofundaria entre 2025 e 2027, com risco de contingenciamento linear, postergação de investimentos essenciais, aumento de restos a pagar e piora na qualidade de serviços públicos e de manutenção urbana, sobretudo no município.",
                ],
            ],
            [126, 254, 374],
        )
    )
    story.extend(
        [
            Spacer(1, 8),
            bullet_box(
                "Indicadores de monitoramento para atualização dos cenários",
                [
                    "Resultado primário e resultado orçamentário como proporção da receita corrente e, no caso do Estado, também como proporção do PIB.",
                    "Despesa com pessoal em relação à receita corrente líquida, com acompanhamento separado de pessoal, encargos e outras despesas correntes.",
                    "Poupança corrente, execução de investimentos prioritários e participação de receitas próprias na composição da arrecadação total.",
                    "Receitas mais sensíveis ao ciclo econômico, especialmente ICMS no Estado e ISS, IPTU e transferências na capital.",
                ],
                styles,
            ),
            Spacer(1, 8),
        ]
    )
    story.append(make_heading("4.3 - Estratégias e plano de ação para sustentabilidade fiscal", styles["H2"], 1, "h2_43"))
    story.append(Paragraph("À luz dos cenários projetados, a resposta recomendada não é um ajuste linear, mas um plano escalonado que combine disciplina fiscal, proteção de políticas essenciais e melhoria da qualidade do gasto. O foco deve recair sobre medidas capazes de reduzir rigidezes, qualificar o investimento e fortalecer a previsibilidade orçamentária.", styles["Body"]))
    story.append(
        narrative_table(
            ["Frente estratégica", "Mato Grosso do Sul", "Campo Grande", "Horizonte e evidência de sucesso"],
            [
                [
                    "1. Revisão de gasto corrente",
                    "Instituir revisão periódica de contratos, benefícios, custeio administrativo e crescimento vegetativo da folha, preservando saúde, educação e segurança como núcleos essenciais.",
                    "Adotar programa de ajuste de despesa corrente com foco em pessoal, horas extras, contratos de serviços e custeio de unidades, com metas gerenciais por secretaria.",
                    "Horizonte de 6 a 12 meses. Sucesso medido por desaceleração da despesa primária corrente e melhora da poupança corrente.",
                ],
                [
                    "2. Gestão de pessoal e encargos",
                    "Tratar a despesa com pessoal como variável estratégica, com planejamento de reposições, ganhos de produtividade e avaliação de carreiras, evitando expansão inercial da folha.",
                    "Priorizar trajetória de retorno sustentável ao limite legal, com gestão ativa da força de trabalho e reavaliação de estruturas administrativas intensivas em despesa permanente.",
                    "Horizonte de 12 a 24 meses. Sucesso medido por redução gradual da pressão da folha sobre a receita corrente líquida.",
                ],
                [
                    "3. Qualidade da receita",
                    "Fortalecer inteligência tributária e monitoramento setorial do ICMS, reduzindo dependência excessiva de bases mais voláteis e aprimorando a previsão de arrecadação.",
                    "Expandir ações sobre ISS, IPTU, ITBI e dívida ativa com uso de dados, cobrança mais eficiente e simplificação para elevar conformidade sem aumentar insegurança ao contribuinte.",
                    "Horizonte de 12 a 24 meses. Sucesso medido por maior previsibilidade da receita própria e menor dependência de ingressos extraordinários.",
                ],
                [
                    "4. Priorização do investimento",
                    "Organizar carteira de projetos por retorno econômico e social, protegendo manutenção, logística, infraestrutura produtiva e obras com maior efeito multiplicador.",
                    "Hierarquizar pavimentação, drenagem, escolas, saúde e manutenção urbana segundo criticidade social, custo de ciclo de vida e fonte estável de financiamento.",
                    "Horizonte de 12 meses em diante. Sucesso medido por maior participação de projetos prioritários executados e menor dispersão orçamentária.",
                ],
                [
                    "5. Gestão financeira e contingência",
                    "Aprimorar fluxo de caixa, cronograma de pagamentos, análise de risco de arrecadação e gatilhos de contingenciamento não linear em caso de frustração de receita.",
                    "Estruturar rotina de tesouraria e matriz de contingência para fornecedores, folha e investimento, reduzindo necessidade de respostas emergenciais descoordenadas.",
                    "Horizonte imediato. Sucesso medido por menor deterioração do resultado orçamentário e maior previsibilidade de execução financeira.",
                ],
                [
                    "6. Governança e transparência",
                    "Criar painel fiscal de acompanhamento com metas trimestrais de receita, despesa, investimento e resultado, articulado ao planejamento de médio prazo.",
                    "Implementar governança fiscal com monitoramento mensal e comunicação pública simples para reforçar credibilidade junto a Câmara, órgãos de controle e sociedade.",
                    "Horizonte imediato e contínuo. Sucesso medido por decisões tempestivas, transparência e redução de desvios entre orçamento e execução.",
                ],
            ],
            [104, 208, 208, 234],
        )
    )
    story.extend(
        [
            Spacer(1, 8),
            bullet_box(
                "Responsabilidades dos principais stakeholders",
                [
                    "Poder Executivo estadual e municipal: liderar revisão de gasto, priorização de investimentos, gestão de caixa e definição de metas fiscais realistas.",
                    "Poder Legislativo: avaliar impactos de novas despesas permanentes, escrutinar renúncias e apoiar mecanismos de transparência e monitoramento de resultados.",
                    "Órgãos de controle: combinar fiscalização de legalidade com avaliação de qualidade do gasto, execução de investimentos e consistência das justificativas de contingenciamento.",
                    "Sociedade, contribuintes e usuários de serviços públicos: acompanhar metas, cobrar transparência e participar do debate sobre prioridades distributivas do ajuste.",
                ],
                styles,
            ),
            Spacer(1, 8),
        ]
    )
    story.append(make_heading("4.4 - Trade-offs, implicações sociais e critérios éticos", styles["H2"], 1, "h2_44"))
    story.append(Paragraph("Os resultados fiscais não podem ser tratados como problema exclusivamente contábil. Em ambos os entes, qualquer estratégia de consolidação envolve dilemas entre estabilidade de curto prazo, proteção social, capacidade de investimento e qualidade dos serviços públicos. Por isso, o ajuste recomendado deve ser seletivo, transparente e socialmente defensável.", styles["Body"]))
    tradeoff_panels = Table(
        [
            [
                info_panel(
                    "Ajuste fiscal versus proteção de serviços essenciais",
                    "Cortes lineares tendem a ser administrativamente fáceis, mas socialmente regressivos. A solução mais consistente é preservar funções de alto retorno social e concentrar o ajuste em ineficiências, contratos pouco efetivos e despesas cujo benefício público seja menor ou mais difuso.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
                info_panel(
                    "Liquidez imediata versus desenvolvimento de médio prazo",
                    "Suspender indiscriminadamente investimentos melhora o caixa no curtíssimo prazo, mas pode agravar custos futuros, reduzir produtividade e comprometer infraestrutura urbana e logística. O desafio é proteger investimentos críticos e, ao mesmo tempo, evitar expansão de projetos sem sustentabilidade financeira.",
                    styles,
                    BODY_WIDTH / 2 - 8,
                ),
            ]
        ],
        colWidths=[BODY_WIDTH / 2 - 6, BODY_WIDTH / 2 - 6],
        hAlign="LEFT",
    )
    tradeoff_panels.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 8)]))
    story.extend([tradeoff_panels, Spacer(1, 10)])
    story.append(Paragraph("Há ainda um trade-off federativo relevante: bases tributárias ligadas ao ciclo econômico podem impulsionar a arrecadação em fases favoráveis, mas não substituem governança fiscal robusta. Em termos de desenvolvimento econômico e social, a solução equilibrada passa por elevar eficiência gerencial, melhorar a coordenação entre política fiscal e planejamento setorial e evitar que o ônus do ajuste recaia desproporcionalmente sobre grupos mais vulneráveis ou sobre investimentos com alto retorno público.", styles["Body"]))
    story.append(
        bullet_box(
            "Princípios orientadores para uma resposta fiscal equilibrada",
            [
                "Substituir cortes lineares por revisão de gasto baseada em evidências, produtividade e impacto social.",
                "Preservar políticas e investimentos com maior retorno social, urbano e econômico, especialmente em saúde, educação, manutenção e infraestrutura crítica.",
                "Dar transparência aos critérios distributivos do ajuste, explicitando quem suporta o custo da consolidação e quais benefícios coletivos se busca proteger.",
                "Combinar disciplina de curto prazo com soluções inovadoras de médio prazo, como painéis fiscais, inteligência de dados, carteira priorizada de projetos e monitoramento contínuo de riscos.",
            ],
            styles,
        )
    )
    story.append(
        bullet_box(
            "Síntese final",
            [
                "O Estado do Mato Grosso do Sul apresenta quadro de expansão de receita, mas ainda dependente de disciplina adicional sobre a trajetória das despesas primárias.",
                "Campo Grande exibe desequilíbrio estrutural mais persistente, com déficits em todos os exercícios e baixa capacidade de compensação via receitas financeiras.",
                "No cenário de reequilíbrio gradual, os dois entes podem reduzir vulnerabilidades sem sacrificar funções públicas essenciais; no cenário inercial ou adverso, o risco é de continuidade ou aprofundamento dos déficits.",
                "A leitura combinada dos dois demonstrativos sugere que, para ambos os entes, o desafio central não está apenas em arrecadar mais, mas em alinhar a dinâmica do gasto à capacidade de financiamento de médio prazo e aos objetivos de desenvolvimento econômico e social.",
            ],
            styles,
        )
    )
    story.extend([NextPageTemplate("FullPage"), PageBreak()])
    story.extend([SectionPage("5", "5.0 - Referências", "Fontes utilizadas na elaboração do relatório e nas análises apresentadas\nsobre a situação fiscal do estado e da capital.", IMG_CG, "sec_refs"), NextPageTemplate("Body"), PageBreak()])
    story.append(Paragraph("SICONFI. Sistema de Informações Contábeis e Fiscais do Setor Público Brasileiro. Brasília: Tesouro Nacional, [s.d.]. Disponível em: https://siconfi.tesouro.gov.br/siconfi/index.jsf. Acesso em: 27 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("BRASIL. Lei Complementar nº 173, de 27 de maio de 2020. Estabelece o Programa Federativo de Enfrentamento ao Coronavírus SARS-CoV-2 (Covid-19). Brasília: Presidência da República, 2020. Disponível em: https://www.planalto.gov.br/ccivil_03/Leis/LCP/Lcp173.htm. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("BRASIL. Lei nº 14.041, de 18 de agosto de 2020. Dispõe sobre a prestação de apoio financeiro pela União aos Estados, ao Distrito Federal e aos Municípios com o objetivo de mitigar dificuldades financeiras decorrentes da pandemia. Brasília: Presidência da República, 2020. Disponível em: https://www.planalto.gov.br/ccivil_03/_ato2019-2022/2020/lei/l14041.htm. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("BRASIL. Lei Complementar nº 194, de 23 de junho de 2022. Considera essenciais combustíveis, energia elétrica, comunicações e transporte coletivo para fins de incidência do ICMS. Brasília: Presidência da República, 2022. Disponível em: https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp194.htm. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("SEMADESC; IBGE. Produto Interno Bruto de Mato Grosso do Sul - 2023. Campo Grande: Secretaria de Estado de Meio Ambiente, Desenvolvimento, Ciência, Tecnologia e Inovação, 2025. Disponível em: https://www.semadesc.ms.gov.br/wp-content/uploads/2025/11/Produto-Interno-Bruto-de-Mato-Grosso-do-Sul-2023-Versao-Final.pdf. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("IBGE. GDP grows in all the 27 states in Brazil in 2023. Agência de Notícias, 14 nov. 2025. Disponível em: https://agenciadenoticias.ibge.gov.br/en/agencia-news/45154-gdp-grows-in-all-the-27-states-in-brazil-in-2023.html. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("MATO GROSSO DO SUL. Secretaria de Estado de Fazenda. Relatório da Gestão Orçamentária e Financeira 2024. Campo Grande: SEFAZ-MS, 2025. Disponível em: https://www.sefaz.ms.gov.br/wp-content/uploads/2025/12/MINUTA-RELATORIO-GESTAO-O.F-2024.pdf. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("PREFEITURA MUNICIPAL DE CAMPO GRANDE. Relatório de Gestão Fiscal Consolidado: demonstrativo da despesa com pessoal, janeiro a dezembro de 2023, 3º quadrimestre. Campo Grande: Prefeitura Municipal, 2024. Disponível em: https://cdn.campogrande.ms.gov.br/portal/prod/uploads/sites/30/2024/03/RELATORIO-QUADRIMESTRAL-3o-RDQA-DE-2023_r.pdf. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("PREFEITURA MUNICIPAL DE CAMPO GRANDE. Relatório de Atividades 2024. Campo Grande: Prefeitura Municipal, 2025. Disponível em: https://cdn.campogrande.ms.gov.br/portal/prod/uploads/sites/29/2025/02/Relatorio-de-Atividades-2024.pdf. Acesso em: 29 mar. 2026.", styles["Ref"]))
    story.append(Paragraph("SECRETARIA DO TESOURO NACIONAL. Boletim de Finanças dos Entes Subnacionais 2025. Brasília: Ministério da Fazenda, 2026.", styles["Ref"]))
    return story


def main() -> None:
    register_fonts()
    ensure_dirs()
    ms1 = SheetData("Dem1_Estado_MS")
    ms2 = SheetData("Dem2_Estado_MS")
    cg1 = SheetData("Dem1_Capital_CG")
    cg2 = SheetData("Dem2_Capital_CG")
    charts = generate_charts(ms1, ms2, cg1, cg2)
    styles = build_styles()
    story = build_story(ms1, ms2, cg1, cg2, charts, styles)
    doc = FiscalDocTemplate(str(OUTPUT_PDF))
    doc.multiBuild(story)
    print(OUTPUT_PDF.name)


if __name__ == "__main__":
    main()
