from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from generate_relatorio_pdf import (
    PIB_BRASIL,
    PIB_CENTRO_OESTE,
    PIB_GROWTH_MS,
    PIB_MS,
    PIB_MS_YEARS,
    SheetData,
    br_money_millions,
    br_money_text,
    br_percent,
    nominal_change,
    pct_change,
)


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "tabelas_relatorio"
WORKBOOK_PATH = OUTPUT_DIR / "tabelas_relatorio.xlsx"


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return normalized


def write_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerows(rows)


def write_workbook(tables: list[tuple[str, str, list[list[str]]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for title, sheet_name, rows in tables:
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append([title])
        ws["A1"].font = Font(bold=True)
        ws.append([])
        for row in rows:
            ws.append(row)
        for cell in ws[3]:
            cell.font = Font(bold=True)

        max_cols = max(len(row) for row in rows) if rows else 1
        for column_index in range(1, max_cols + 1):
            width = 14
            for row in rows:
                if column_index <= len(row):
                    width = max(width, min(len(str(row[column_index - 1])) + 2, 60))
            ws.column_dimensions[chr(64 + column_index)].width = width

    wb.save(WORKBOOK_PATH)


def variation_rows(sheet: SheetData, labels: list[str]) -> list[list[str]]:
    rows = [["Discriminacao", *[str(year) for year in sheet.years], "Variacao Nominal", "Variacao (%)"]]
    for label in labels:
        values = sheet.get(label)
        rows.append(
            [
                sheet.label(label),
                *[br_money_millions(value, 2) for value in values],
                br_money_millions(nominal_change(values), 2),
                br_percent(pct_change(values), 1),
            ]
        )
    return rows


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)

    ms1 = SheetData("Dem1_Estado_MS")
    ms2 = SheetData("Dem2_Estado_MS")
    cg1 = SheetData("Dem1_Capital_CG")
    cg2 = SheetData("Dem2_Capital_CG")

    ms_receita_corrente = ms1.get("RECEITA CORRENTE")
    ms_receita_primaria_total = ms1.get("RECEITA PRIMARIA TOTAL (C = A + B)")
    ms_despesa_primaria_total = ms1.get("DESPESA PRIMARIA TOTAL (F = D + E)")
    ms_resultado_primario = ms2.get("RESULTADO PRIMARIO")
    ms_resultado_orc = ms2.get("RESULTADO ORCAMENTARIO")

    cg_receita_corrente = cg1.get("RECEITA CORRENTE")
    cg_receita_primaria_total = cg1.get("RECEITA PRIMARIA TOTAL (C = A + B)")
    cg_despesa_primaria_total = cg1.get("DESPESA PRIMARIA TOTAL (F = D + E)")
    cg_resultado_primario = cg2.get("RESULTADO PRIMARIO")
    cg_resultado_orc = cg2.get("RESULTADO ORCAMENTARIO")

    pib_ms_part_brasil = [value / total * 100 for value, total in zip(PIB_MS, PIB_BRASIL)]
    ms_resultado_primario_pct_pib = [value / pib * 100 for value, pib in zip(ms_resultado_primario[: len(PIB_MS_YEARS)], PIB_MS)]

    tables: list[tuple[str, str, list[list[str]]]] = [
        (
            "Visao geral 2024",
            "visao_geral",
            [
                ["Indicador", "Mato Grosso do Sul", "Campo Grande", "Leitura sintetica"],
                [
                    "Receita corrente 2024",
                    br_money_text(ms_receita_corrente[-1]),
                    br_money_text(cg_receita_corrente[-1]),
                    "Em ambos os casos houve expansao relevante da arrecadacao, mas em ritmos distintos.",
                ],
                [
                    "Despesa primaria total 2024",
                    br_money_text(ms_despesa_primaria_total[-1]),
                    br_money_text(cg_despesa_primaria_total[-1]),
                    "A pressao de despesa permaneceu elevada e absorveu parcela expressiva do ganho de receita.",
                ],
                [
                    "Resultado primario 2024",
                    br_money_text(ms_resultado_primario[-1]),
                    br_money_text(cg_resultado_primario[-1]),
                    "O estado voltou a registrar deficit primario robusto; a capital manteve deficit em toda a serie.",
                ],
                [
                    "Resultado orcamentario 2024",
                    br_money_text(ms_resultado_orc[-1]),
                    br_money_text(cg_resultado_orc[-1]),
                    "As receitas financeiras nao foram suficientes para neutralizar as pressoes de gasto.",
                ],
            ],
        ),
        (
            "Tabela PIB 1 - Comparacao do PIB nominal do Brasil, do Centro-Oeste e de Mato Grosso do Sul (2019-2023)",
            "pib_tabela_1",
            [
                ["Ano", "Brasil", "Centro-Oeste", "Mato Grosso do Sul", "Part. de MS no Brasil (%)"],
                *[
                    [
                        str(year),
                        br_money_millions(br, 2),
                        br_money_millions(co, 2),
                        br_money_millions(ms, 2),
                        br_percent(part, 2),
                    ]
                    for year, br, co, ms, part in zip(PIB_MS_YEARS, PIB_BRASIL, PIB_CENTRO_OESTE, PIB_MS, pib_ms_part_brasil)
                ],
            ],
        ),
        (
            "Tabela PIB 2 - PIB de Mato Grosso do Sul, crescimento real e resultado primario relativo ao PIB (2019-2023)",
            "pib_tabela_2",
            [
                ["Ano", "PIB MS (R$ milhoes)", "Crescimento real do PIB (%)", "Resultado primario (R$ milhoes)", "Resultado primario / PIB (%)"],
                *[
                    [
                        str(year),
                        br_money_millions(pib, 2),
                        br_percent(growth, 2),
                        br_money_millions(primary, 2),
                        br_percent(primary_pct, 2),
                    ]
                    for year, pib, growth, primary, primary_pct in zip(
                        PIB_MS_YEARS,
                        PIB_MS,
                        PIB_GROWTH_MS,
                        ms_resultado_primario[: len(PIB_MS_YEARS)],
                        ms_resultado_primario_pct_pib,
                    )
                ],
            ],
        ),
        (
            "Tabela 1 - Receitas correntes e saldo primario corrente do estado",
            "estado_tabela_1",
            variation_rows(
                ms1,
                [
                    "RECEITA CORRENTE",
                    "Impostos, Taxas e Contribuicoes de Melhoria",
                    "ICMS",
                    "IPVA",
                    "Transferencias Correntes",
                    "Cota-Parte do FPE",
                    "Receitas Financeiras Correntes",
                    "Demais Receitas Correntes",
                    "SALDO (A) - Receitas Primarias Correntes",
                ],
            ),
        ),
        (
            "Tabela 2 - Receitas de capital e receita primaria total do estado",
            "estado_tabela_2",
            variation_rows(
                ms1,
                [
                    "RECEITA DE CAPITAL",
                    "Receitas Financeiras de Capital",
                    "Operacoes de Credito",
                    "Alienacao de Bens",
                    "Amortizacao de Emprestimos",
                    "Transferencias de Capital",
                    "SALDO (B) - Receitas Primarias de Capital",
                    "RECEITA PRIMARIA TOTAL (C = A + B)",
                ],
            ),
        ),
        (
            "Tabela 3 - Despesas correntes do estado",
            "estado_tabela_3",
            variation_rows(
                ms1,
                [
                    "DESPESA CORRENTE",
                    "Pessoal e Encargos Sociais",
                    "Juros e Encargos da Divida",
                    "Outras Despesas Correntes",
                    "SALDO (D) - Despesas Primarias Correntes",
                ],
            ),
        ),
        (
            "Tabela 4 - Despesas de capital e despesa primaria total do estado",
            "estado_tabela_4",
            variation_rows(
                ms1,
                [
                    "DESPESA DE CAPITAL",
                    "Investimentos",
                    "Demais Inversoes",
                    "Despesas Financeiras de Capital",
                    "Amortizacao da Divida",
                    "SALDO (E) - Despesas Primarias de Capital",
                    "DESPESA PRIMARIA TOTAL (F = D + E)",
                ],
            ),
        ),
        (
            "Tabela 5 - Consolidacao do resultado primario do estado",
            "estado_tabela_5",
            variation_rows(
                ms2,
                ["RECEITAS PRIMARIAS", "DESPESAS PRIMARIAS", "RESULTADO PRIMARIO"],
            ),
        ),
        (
            "Tabela 6 - Segundo demonstrativo do estado do Mato Grosso do Sul",
            "estado_tabela_6",
            variation_rows(
                ms2,
                [
                    "RECEITAS PRIMARIAS",
                    "DESPESAS PRIMARIAS",
                    "RESULTADO PRIMARIO",
                    "RECEITAS FINANCEIRAS",
                    "DESPESAS FINANCEIRAS",
                    "RESULTADO ORCAMENTARIO",
                ],
            ),
        ),
        (
            "Tabela 7 - Receitas correntes e saldo primario corrente de Campo Grande",
            "cg_tabela_7",
            variation_rows(
                cg1,
                [
                    "RECEITA CORRENTE",
                    "Impostos, Taxas e Contribuicoes de Melhoria",
                    "ISS",
                    "IPTU",
                    "Transferencias Correntes",
                    "Cota-Parte do FPM",
                    "Receitas Financeiras Correntes",
                    "Demais Receitas Correntes",
                    "SALDO (A) - Receitas Primarias Correntes",
                ],
            ),
        ),
        (
            "Tabela 8 - Receitas de capital e receita primaria total de Campo Grande",
            "cg_tabela_8",
            variation_rows(
                cg1,
                [
                    "RECEITA DE CAPITAL",
                    "Receitas Financeiras de Capital",
                    "Operacoes de Credito",
                    "Alienacao de Bens",
                    "Amortizacao de Emprestimos",
                    "Transferencias de Capital",
                    "SALDO (B) - Receitas Primarias de Capital",
                    "RECEITA PRIMARIA TOTAL (C = A + B)",
                ],
            ),
        ),
        (
            "Tabela 9 - Despesas correntes de Campo Grande",
            "cg_tabela_9",
            variation_rows(
                cg1,
                [
                    "DESPESA CORRENTE",
                    "Pessoal e Encargos Sociais",
                    "Juros e Encargos da Divida",
                    "Outras Despesas Correntes",
                    "SALDO (D) - Despesas Primarias Correntes",
                ],
            ),
        ),
        (
            "Tabela 10 - Despesas de capital e despesa primaria total de Campo Grande",
            "cg_tabela_10",
            variation_rows(
                cg1,
                [
                    "DESPESA DE CAPITAL",
                    "Investimentos",
                    "Demais Inversoes",
                    "Despesas Financeiras de Capital",
                    "Amortizacao da Divida",
                    "SALDO (E) - Despesas Primarias de Capital",
                    "DESPESA PRIMARIA TOTAL (F = D + E)",
                ],
            ),
        ),
        (
            "Tabela 11 - Consolidacao do resultado primario de Campo Grande",
            "cg_tabela_11",
            variation_rows(
                cg2,
                ["RECEITAS PRIMARIAS", "DESPESAS PRIMARIAS", "RESULTADO PRIMARIO"],
            ),
        ),
        (
            "Tabela 12 - Segundo demonstrativo de Campo Grande",
            "cg_tabela_12",
            variation_rows(
                cg2,
                [
                    "RECEITAS PRIMARIAS",
                    "DESPESAS PRIMARIAS",
                    "RESULTADO PRIMARIO",
                    "RECEITAS FINANCEIRAS",
                    "DESPESAS FINANCEIRAS",
                    "RESULTADO ORCAMENTARIO",
                ],
            ),
        ),
        (
            "Tabela 13 - Cenarios prospectivos e projecao analitica",
            "cenarios",
            [
                ["Cenario", "Condicoes e gatilhos", "Implicacoes fiscais esperadas"],
                [
                    "Reequilibrio gradual",
                    "Crescimento economico moderado, manutencao de arrecadacao corrente em linha com o PIB nominal, contencao seletiva de despesas correntes, revisao de contratos e maior priorizacao do investimento de maior retorno social e economico.",
                    "No Mato Grosso do Sul, o deficit primario tenderia a diminuir progressivamente e o resultado orcamentario poderia se aproximar do equilibrio. Em Campo Grande, o deficit seguiria presente no curto prazo, mas com reducao gradual e recuperacao parcial da capacidade de investimento prioritario.",
                ],
                [
                    "Inercia fiscal",
                    "Desaceleracao da receita apos o pico recente, ausencia de mudancas estruturais na gestao do gasto, manutencao da pressao de pessoal e de custeio e execucao de investimentos sem filtro mais rigoroso de prioridade.",
                    "Os dois entes preservariam deficits recorrentes. O Estado manteria dificuldade de converter crescimento de receita em poupanca corrente, enquanto a capital seguiria com desequilibrio estrutural, maior dependencia de receitas extraordinarias e menor espaco para absorver choques.",
                ],
                [
                    "Estresse adverso",
                    "Choque negativo sobre agronegocio, ICMS ou transferencias, juros elevados, restricoes de credito e eventos climaticos que pressionem receita e gasto ao mesmo tempo, sem resposta gerencial suficientemente rapida.",
                    "A deterioracao fiscal se aprofundaria entre 2025 e 2027, com risco de contingenciamento linear, postergacao de investimentos essenciais, aumento de restos a pagar e piora na qualidade de servicos publicos e de manutencao urbana, sobretudo no municipio.",
                ],
            ],
        ),
        (
            "Tabela 14 - Estrategias e plano de acao para sustentabilidade fiscal",
            "plano_acao",
            [
                ["Frente estrategica", "Mato Grosso do Sul", "Campo Grande", "Horizonte e evidencia de sucesso"],
                [
                    "1. Revisao de gasto corrente",
                    "Instituir revisao periodica de contratos, beneficios, custeio administrativo e crescimento vegetativo da folha, preservando saude, educacao e seguranca como nucleos essenciais.",
                    "Adotar programa de ajuste de despesa corrente com foco em pessoal, horas extras, contratos de servicos e custeio de unidades, com metas gerenciais por secretaria.",
                    "Horizonte de 6 a 12 meses. Sucesso medido por desaceleracao da despesa primaria corrente e melhora da poupanca corrente.",
                ],
                [
                    "2. Gestao de pessoal e encargos",
                    "Tratar a despesa com pessoal como variavel estrategica, com planejamento de reposicoes, ganhos de produtividade e avaliacao de carreiras, evitando expansao inercial da folha.",
                    "Priorizar trajetoria de retorno sustentavel ao limite legal, com gestao ativa da forca de trabalho e reavaliacao de estruturas administrativas intensivas em despesa permanente.",
                    "Horizonte de 12 a 24 meses. Sucesso medido por reducao gradual da pressao da folha sobre a receita corrente liquida.",
                ],
                [
                    "3. Qualidade da receita",
                    "Fortalecer inteligencia tributaria e monitoramento setorial do ICMS, reduzindo dependencia excessiva de bases mais volateis e aprimorando a previsao de arrecadacao.",
                    "Expandir acoes sobre ISS, IPTU, ITBI e divida ativa com uso de dados, cobranca mais eficiente e simplificacao para elevar conformidade sem aumentar inseguranca ao contribuinte.",
                    "Horizonte de 12 a 24 meses. Sucesso medido por maior previsibilidade da receita propria e menor dependencia de ingressos extraordinarios.",
                ],
                [
                    "4. Priorizacao do investimento",
                    "Organizar carteira de projetos por retorno economico e social, protegendo manutencao, logistica, infraestrutura produtiva e obras com maior efeito multiplicador.",
                    "Hierarquizar pavimentacao, drenagem, escolas, saude e manutencao urbana segundo criticidade social, custo de ciclo de vida e fonte estavel de financiamento.",
                    "Horizonte de 12 meses em diante. Sucesso medido por maior participacao de projetos prioritarios executados e menor dispersao orcamentaria.",
                ],
                [
                    "5. Gestao financeira e contingencia",
                    "Aprimorar fluxo de caixa, cronograma de pagamentos, analise de risco de arrecadacao e gatilhos de contingenciamento nao linear em caso de frustracao de receita.",
                    "Estruturar rotina de tesouraria e matriz de contingencia para fornecedores, folha e investimento, reduzindo necessidade de respostas emergenciais descoordenadas.",
                    "Horizonte imediato. Sucesso medido por menor deterioracao do resultado orcamentario e maior previsibilidade de execucao financeira.",
                ],
                [
                    "6. Governanca e transparencia",
                    "Criar painel fiscal de acompanhamento com metas trimestrais de receita, despesa, investimento e resultado, articulado ao planejamento de medio prazo.",
                    "Implementar governanca fiscal com monitoramento mensal e comunicacao publica simples para reforcar credibilidade junto a Camara, orgaos de controle e sociedade.",
                    "Horizonte imediato e continuo. Sucesso medido por decisoes tempestivas, transparencia e reducao de desvios entre orcamento e execucao.",
                ],
            ],
        ),
    ]

    index_lines = []
    for title, sheet_name, rows in tables:
        csv_name = f"{slugify(sheet_name)}.csv"
        write_csv(OUTPUT_DIR / csv_name, rows)
        index_lines.append(f"{csv_name} - {title}")

    write_workbook(tables)

    index_path = OUTPUT_DIR / "indice_tabelas.txt"
    index_path.write_text(
        "Tabelas exportadas do relatorio\n\n" + "\n".join(index_lines) + f"\n\ntabelas_relatorio.xlsx - consolidado com todas as tabelas\n",
        encoding="utf-8",
    )

    print(OUTPUT_DIR)


if __name__ == "__main__":
    main()
